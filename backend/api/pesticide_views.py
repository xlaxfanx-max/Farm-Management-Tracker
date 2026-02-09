"""
Pesticide product and application views.
"""
import csv
import io
import re
from datetime import datetime, timedelta, date
from decimal import Decimal
from django.db.models import Sum, Avg, Count, F, Q, Min, Max
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, filters, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .pur_reporting import PURReportGenerator
from .product_import_tool import PesticideProductImporter
from .models import (
    PesticideProduct, PesticideApplication, Farm, Field,
    CROP_VARIETY_CHOICES, GRADE_CHOICES,
)
from .serializers import (
    PesticideProductSerializer, PesticideApplicationSerializer,
)
from .view_helpers import get_user_company, require_company
from .audit_utils import AuditLogMixin
from .permissions import HasCompanyAccess


class PesticideProductViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing pesticide products
    """
    queryset = PesticideProduct.objects.all()
    serializer_class = PesticideProductSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['product_name', 'epa_registration_number', 'manufacturer', 'active_ingredients']
    ordering_fields = ['product_name', 'created_at']

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import products from CSV file"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        csv_file = request.FILES['file']
        update_existing = request.data.get('update_existing', 'true').lower() == 'true'

        importer = PesticideProductImporter()
        result = importer.import_from_csv(csv_file, update_existing)

        return Response({
            'success': result['errors'] == 0,
            'message': f"Imported {result['created']} new products, updated {result['updated']} existing products",
            'statistics': result
        })


    @action(detail=False, methods=['get'])
    def export_csv_template(self, request):
        """Download CSV template for importing products"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row with all fields
        headers = [
            'epa_registration_number',
            'product_name',
            'manufacturer',
            'active_ingredients',
            'formulation_type',
            'restricted_use',
            'product_type',
            'is_fumigant',
            'signal_word',
            'rei_hours',
            'rei_days',
            'phi_days',
            'max_applications_per_season',
            'max_rate_per_application',
            'max_rate_unit',
            'california_registration_number',
            'active_status_california',
            'formulation_code',
            'density_specific_gravity',
            'approved_crops',
            'groundwater_advisory',
            'endangered_species_restrictions',
            'buffer_zone_required',
            'buffer_zone_feet',
            'product_status',
            'unit_size',
            'cost_per_unit',
            'label_url',
            'sds_url',
            'notes',
            'active',
        ]

        writer.writerow(headers)

        # Example row
        example = [
            '12345-678',
            'Example Insecticide 2.5EC',
            'Example Chemical Company',
            'Permethrin 25%',
            'Emulsifiable Concentrate',
            'false',
            'insecticide',
            'false',
            'CAUTION',
            '12',
            '',
            '7',
            '4',
            '1.0',
            'lbs/acre',
            '',
            'true',
            'EC',
            '1.05',
            'Citrus, Almonds, Walnuts',
            'false',
            'false',
            'false',
            '',
            'active',
            '2.5 gallon',
            '125.50',
            'https://example.com/label.pdf',
            'https://example.com/sds.pdf',
            'Apply early morning or late evening',
            'true',
        ]

        writer.writerow(example)

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pesticide_products_template.csv"'
        return response

    @action(detail=False, methods=['get'])
    def export_current_products(self, request):
        """Export current products to CSV"""
        products = PesticideProduct.objects.all()

        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        headers = [
            'epa_registration_number', 'product_name', 'manufacturer',
            'active_ingredients', 'formulation_type', 'restricted_use',
            'product_type', 'is_fumigant', 'signal_word',
            'rei_hours', 'rei_days', 'phi_days',
            'max_applications_per_season', 'max_rate_per_application', 'max_rate_unit',
            'california_registration_number', 'active_status_california',
            'formulation_code', 'approved_crops', 'product_status',
            'unit_size', 'cost_per_unit', 'label_url', 'notes', 'active'
        ]
        writer.writerow(headers)

        # Data rows
        for product in products:
            row = [
                product.epa_registration_number,
                product.product_name,
                product.manufacturer,
                product.active_ingredients,
                product.formulation_type,
                product.restricted_use,
                product.product_type if hasattr(product, 'product_type') else '',
                product.is_fumigant if hasattr(product, 'is_fumigant') else False,
                product.signal_word if hasattr(product, 'signal_word') else '',
                product.rei_hours if hasattr(product, 'rei_hours') else '',
                product.rei_days if hasattr(product, 'rei_days') else '',
                product.phi_days if hasattr(product, 'phi_days') else '',
                product.max_applications_per_season if hasattr(product, 'max_applications_per_season') else '',
                product.max_rate_per_application if hasattr(product, 'max_rate_per_application') else '',
                product.max_rate_unit if hasattr(product, 'max_rate_unit') else '',
                product.california_registration_number if hasattr(product, 'california_registration_number') else '',
                product.active_status_california if hasattr(product, 'active_status_california') else True,
                product.formulation_code if hasattr(product, 'formulation_code') else '',
                product.approved_crops if hasattr(product, 'approved_crops') else '',
                product.product_status if hasattr(product, 'product_status') else 'active',
                product.unit_size if hasattr(product, 'unit_size') else '',
                product.cost_per_unit if hasattr(product, 'cost_per_unit') else '',
                product.label_url if hasattr(product, 'label_url') else '',
                product.notes if hasattr(product, 'notes') else '',
                product.active if hasattr(product, 'active') else True,
            ]
            writer.writerow(row)

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pesticide_products_export.csv"'
        return response


class PesticideApplicationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing pesticide applications.

    RLS NOTES:
    - Applications inherit company through field->farm relationship
    - get_queryset filters by company
    """
    serializer_class = PesticideApplicationSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['field__name', 'product__product_name', 'applicator_name']
    ordering_fields = ['application_date', 'created_at']

    def get_queryset(self):
        """Filter applications by current user's company through field->farm."""
        queryset = PesticideApplication.objects.select_related(
            'field', 'field__farm', 'product'
        )
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(field__farm__company=company)
        return queryset

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get all pending applications"""
        pending = self.get_queryset().filter(status='pending_signature')
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def ready_for_pur(self, request):
        """Get all complete applications ready for PUR submission"""
        ready = self.get_queryset().filter(status='complete', submitted_to_pur=False)
        serializer = self.get_serializer(ready, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_complete(self, request, pk=None):
        """Mark an application as complete"""
        application = self.get_object()
        application.status = 'complete'
        application.save()
        serializer = self.get_serializer(application)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_submitted(self, request, pk=None):
        """Mark an application as submitted to PUR"""
        application = self.get_object()
        application.submitted_to_pur = True
        application.status = 'submitted'
        from django.utils import timezone
        application.pur_submission_date = timezone.now().date()
        application.save()
        serializer = self.get_serializer(application)
        return Response(serializer.data)


    @action(detail=False, methods=['post'])
    def validate_pur(self, request):
        """Validate applications for PUR compliance."""
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        farm_id = request.data.get('farm_id')

        # Filter applications
        queryset = self.get_queryset()

        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)

        # Generate validation report
        generator = PURReportGenerator(queryset)
        validation_result = generator.validate_for_pur()

        return Response(validation_result)

    @action(detail=False, methods=['post'])
    def export_pur_csv(self, request):
        """Export applications as PUR-formatted CSV."""
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        farm_id = request.data.get('farm_id')

        # Filter applications
        queryset = self.get_queryset()

        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)

        # Validate first
        generator = PURReportGenerator(queryset)
        validation = generator.validate_for_pur()

        if not validation['valid']:
            return Response({
                'error': 'Applications contain validation errors',
                'validation': validation
            }, status=status.HTTP_400_BAD_REQUEST)

        # Generate CSV
        csv_content = generator.generate_csv()

        # Create response
        response = HttpResponse(csv_content, content_type='text/csv')
        filename = f"PUR_Report_{start_date}_to_{end_date}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['post'])
    def pur_summary(self, request):
        """Get summary statistics for PUR report period."""
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        farm_id = request.data.get('farm_id')

        # Filter applications
        queryset = self.get_queryset()

        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)

        # Generate summary
        generator = PURReportGenerator(queryset)
        summary = generator.generate_summary_report()
        validation = generator.validate_for_pur()

        return Response({
            'summary': summary,
            'validation': validation
        })

    @action(detail=False, methods=['get'])
    def export_pur(self, request):
        """
        Enhanced PUR export with multiple format options.

        Query Parameters:
        - format: 'csv' (official CA PUR format), 'excel', or 'csv_detailed'
        - start_date: YYYY-MM-DD
        - end_date: YYYY-MM-DD
        - farm_id: Filter by farm
        - status: Filter by status
        - county: Filter by county
        - validate: 'true' to validate before export (default: true)
        """
        # Get parameters
        export_format = request.query_params.get('format', 'csv')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        farm_id = request.query_params.get('farm_id')
        status_filter = request.query_params.get('status')
        county = request.query_params.get('county')
        validate_first = request.query_params.get('validate', 'true').lower() == 'true'

        # Build query
        queryset = self.get_queryset()

        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if county:
            queryset = queryset.filter(field__county__icontains=county)

        # Select related to optimize
        queryset = queryset.select_related('field', 'field__farm', 'product').order_by('application_date')

        # Use existing PURReportGenerator for validation and official CSV
        generator = PURReportGenerator(queryset)

        # Validate if requested
        if validate_first:
            validation = generator.validate_for_pur()
            if not validation['valid'] and export_format == 'csv':
                # For official PUR format, enforce validation
                return Response({
                    'error': 'Applications contain validation errors. Cannot export official PUR format.',
                    'validation': validation
                }, status=status.HTTP_400_BAD_REQUEST)

        # Route to appropriate export method
        if export_format == 'csv':
            # Use existing official CA PUR format
            return self._export_official_pur_csv(generator, start_date, end_date)
        elif export_format == 'excel':
            # New Excel format with summary
            return self._export_pur_excel(queryset, start_date, end_date)
        elif export_format == 'csv_detailed':
            # Detailed CSV with all fields
            return self._export_pur_csv_detailed(queryset, start_date, end_date)
        else:
            return Response({'error': 'Invalid format'}, status=status.HTTP_400_BAD_REQUEST)


    def _export_official_pur_csv(self, generator, start_date, end_date):
        """Export using existing official California PUR format"""
        csv_content = generator.generate_csv()

        response = HttpResponse(csv_content, content_type='text/csv')
        filename = f"PUR_Official_{start_date or 'all'}_to_{end_date or 'all'}_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


    def _export_pur_csv_detailed(self, queryset, start_date, end_date):
        """Export detailed CSV with all available fields"""
        import csv

        response = HttpResponse(content_type='text/csv')

        date_range = ""
        if start_date and end_date:
            date_range = f"_{start_date}_to_{end_date}"

        filename = f"PUR_Detailed{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)

        # Detailed header
        writer.writerow([
            'Application Date',
            'Farm Name',
            'Farm Number',
            'Operator Name',
            'County',
            'Field Name',
            'Field Number',
            'Section',
            'Township',
            'Range',
            'GPS Latitude',
            'GPS Longitude',
            'Acres Treated',
            'Current Crop',
            'EPA Registration Number',
            'Product Name',
            'Active Ingredients',
            'Amount Used',
            'Unit',
            'Application Method',
            'Target Pest',
            'Applicator Name',
            'Start Time',
            'End Time',
            'Temperature (°F)',
            'Wind Speed (mph)',
            'Wind Direction',
            'Restricted Use',
            'Fumigant',
            'REI (hours)',
            'PHI (days)',
            'Signal Word',
            'Status',
            'PUR Submitted',
            'Submission Date',
            'Notes'
        ])

        # Write data
        for app in queryset:
            writer.writerow([
                app.application_date.strftime('%m/%d/%Y'),
                app.field.farm.name,
                app.field.farm.farm_number or '',
                app.field.farm.operator_name or '',
                app.field.county,
                app.field.name,
                app.field.field_number or '',
                app.field.section or '',
                app.field.township or '',
                app.field.range_value or '',
                app.field.gps_lat or '',
                app.field.gps_long or '',
                app.acres_treated,
                app.field.current_crop or '',
                app.product.epa_registration_number,
                app.product.product_name,
                app.product.active_ingredients,
                app.amount_used,
                app.unit_of_measure,
                app.application_method,
                app.target_pest or '',
                app.applicator_name,
                app.start_time.strftime('%H:%M') if app.start_time else '',
                app.end_time.strftime('%H:%M') if app.end_time else '',
                app.temperature or '',
                app.wind_speed or '',
                app.wind_direction or '',
                'Yes' if app.product.restricted_use else 'No',
                'Yes' if app.product.is_fumigant else 'No',
                app.product.rei_hours or '',
                app.product.phi_days or '',
                app.product.signal_word or '',
                app.get_status_display(),
                'Yes' if app.submitted_to_pur else 'No',
                app.pur_submission_date.strftime('%m/%d/%Y') if app.pur_submission_date else '',
                app.notes or ''
            ])

        return response


    def _export_pur_excel(self, queryset, start_date, end_date):
        """Export detailed Excel with formatting and summary"""
        wb = Workbook()

        # === OFFICIAL PUR FORMAT SHEET ===
        ws_pur = wb.active
        ws_pur.title = "Official PUR Format"

        # Use PURReportGenerator for official format
        generator = PURReportGenerator(queryset)

        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Write official PUR headers
        for col_num, field_name in enumerate(generator.PUR_FIELDS, 1):
            cell = ws_pur.cell(row=1, column=col_num)
            cell.value = field_name.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Write official PUR data
        for row_num, app in enumerate(queryset, 2):
            row_data = generator._application_to_pur_row(app)
            for col_num, field_name in enumerate(generator.PUR_FIELDS, 1):
                cell = ws_pur.cell(row=row_num, column=col_num)
                cell.value = row_data.get(field_name, '')
                cell.border = border

        # Auto-size columns for PUR sheet
        for col in range(1, len(generator.PUR_FIELDS) + 1):
            ws_pur.column_dimensions[chr(64 + col)].width = 15

        ws_pur.freeze_panes = 'A2'

        # === DETAILED DATA SHEET ===
        ws_detail = wb.create_sheet("Detailed Data")

        detail_headers = [
            'Date', 'Farm', 'Farm #', 'County', 'Field', 'Field #',
            'Acres', 'Crop', 'EPA #', 'Product', 'Active Ingredients',
            'Amount', 'Unit', 'Method', 'Target Pest', 'Applicator',
            'Start', 'End', 'Temp', 'Wind Speed', 'Wind Dir',
            'Restricted', 'REI', 'PHI', 'Status', 'Notes'
        ]

        # Write headers
        for col_num, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Write data
        for row_num, app in enumerate(queryset, 2):
            data = [
                app.application_date.strftime('%m/%d/%Y'),
                app.field.farm.name,
                app.field.farm.farm_number or '',
                app.field.county,
                app.field.name,
                app.field.field_number or '',
                float(app.acres_treated),
                app.field.current_crop or '',
                app.product.epa_registration_number,
                app.product.product_name,
                app.product.active_ingredients,
                float(app.amount_used),
                app.unit_of_measure,
                app.application_method,
                app.target_pest or '',
                app.applicator_name,
                app.start_time.strftime('%H:%M') if app.start_time else '',
                app.end_time.strftime('%H:%M') if app.end_time else '',
                app.temperature or '',
                app.wind_speed or '',
                app.wind_direction or '',
                'Yes' if app.product.restricted_use else 'No',
                app.product.rei_hours or '',
                app.product.phi_days or '',
                app.get_status_display(),
                app.notes or ''
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws_detail.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

        # Auto-size columns
        for col_num in range(1, len(detail_headers) + 1):
            col_letter = chr(64 + col_num)
            ws_detail.column_dimensions[col_letter].width = 12

        ws_detail.freeze_panes = 'A2'

        # === SUMMARY SHEET ===
        ws_summary = wb.create_sheet("Summary & Validation")
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # Title
        ws_summary['A1'] = "PUR Report Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)

        row = 3

        # Report Info
        ws_summary[f'A{row}'] = "Report Generated:"
        ws_summary[f'B{row}'] = datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        row += 1

        if start_date:
            ws_summary[f'A{row}'] = "Start Date:"
            ws_summary[f'B{row}'] = start_date
            row += 1

        if end_date:
            ws_summary[f'A{row}'] = "End Date:"
            ws_summary[f'B{row}'] = end_date
            row += 1

        row += 1

        # Statistics
        ws_summary[f'A{row}'] = "STATISTICS"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws_summary[f'A{row}'] = "Total Applications:"
        ws_summary[f'B{row}'] = queryset.count()
        row += 1

        ws_summary[f'A{row}'] = "Total Acres Treated:"
        ws_summary[f'B{row}'] = queryset.aggregate(Sum('acres_treated'))['acres_treated__sum'] or 0
        row += 1

        ws_summary[f'A{row}'] = "Unique Farms:"
        ws_summary[f'B{row}'] = queryset.values('field__farm').distinct().count()
        row += 1

        ws_summary[f'A{row}'] = "Unique Fields:"
        ws_summary[f'B{row}'] = queryset.values('field').distinct().count()
        row += 1

        ws_summary[f'A{row}'] = "Unique Products:"
        ws_summary[f'B{row}'] = queryset.values('product').distinct().count()
        row += 2

        # Validation Results
        validation = generator.validate_for_pur()

        ws_summary[f'A{row}'] = "VALIDATION RESULTS"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws_summary[f'A{row}'] = "Ready for PUR Submission:"
        ws_summary[f'B{row}'] = "YES" if validation['valid'] else "NO"
        ws_summary[f'B{row}'].font = Font(
            color="00FF00" if validation['valid'] else "FF0000",
            bold=True
        )
        row += 2

        if validation['errors']:
            ws_summary[f'A{row}'] = "ERRORS (Must Fix):"
            ws_summary[f'A{row}'].font = Font(bold=True, color="FF0000")
            row += 1
            for error in validation['errors']:
                ws_summary[f'A{row}'] = error
                ws_summary[f'A{row}'].font = Font(color="FF0000")
                row += 1
            row += 1

        if validation['warnings']:
            ws_summary[f'A{row}'] = "WARNINGS (Recommended):"
            ws_summary[f'A{row}'].font = Font(bold=True, color="FFA500")
            row += 1
            for warning in validation['warnings']:
                ws_summary[f'A{row}'] = warning
                ws_summary[f'A{row}'].font = Font(color="FFA500")
                row += 1

        # Status breakdown
        row += 2
        ws_summary[f'A{row}'] = "STATUS BREAKDOWN"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        status_counts = queryset.values('status').annotate(count=Count('id'))
        for status_data in status_counts:
            ws_summary[f'A{row}'] = status_data['status'].replace('_', ' ').title()
            ws_summary[f'B{row}'] = status_data['count']
            row += 1

        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        date_range = ""
        if start_date and end_date:
            date_range = f"_{start_date}_to_{end_date}"

        filename = f"PUR_Report{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


    # Also add the validation and summary endpoints from existing file
    @action(detail=False, methods=['post', 'get'])
    def validate_pur(self, request):
        """
        Validate applications for PUR compliance.

        Supports both GET and POST methods for flexibility.
        """
        # Get parameters from either query params or body
        if request.method == 'POST':
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            farm_id = request.data.get('farm_id')
        else:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            farm_id = request.query_params.get('farm_id')

        # Filter applications
        queryset = self.get_queryset()

        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)

        queryset = queryset.select_related('field', 'field__farm', 'product')

        # Generate validation report
        generator = PURReportGenerator(queryset)
        validation_result = generator.validate_for_pur()

        return Response(validation_result)


    @action(detail=False, methods=['post', 'get'])
    def pur_summary(self, request):
        """
        Get summary statistics for PUR report period.

        Supports both GET and POST methods.
        """
        # Get parameters
        if request.method == 'POST':
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            farm_id = request.data.get('farm_id')
        else:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            farm_id = request.query_params.get('farm_id')

        # Filter applications
        queryset = self.get_queryset()

        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)

        queryset = queryset.select_related('field', 'field__farm', 'product')

        # Generate summary and validation
        generator = PURReportGenerator(queryset)
        summary = generator.generate_summary_report()
        validation = generator.validate_for_pur()

        return Response({
            'summary': summary,
            'validation': validation
        })

    # =========================================================================
    # SERVICE-BASED ENDPOINTS
    # These endpoints use the new services layer for business logic
    # =========================================================================

    @action(detail=False, methods=['post'])
    def validate_proposed(self, request):
        """
        Validate a proposed pesticide application before creating it.

        Uses the PesticideComplianceService for comprehensive validation
        including PHI, REI, rate limits, NOI requirements, and weather.

        POST body:
        {
            "field_id": 1,
            "product_id": 5,
            "application_date": "2024-12-20",
            "rate_per_acre": 2.5,
            "application_method": "Ground Spray",
            "acres_treated": 10.0,
            "applicator_name": "John Smith",  // optional but needed for restricted
            "applicator_license": "12345",    // optional
            "check_weather": true,            // optional, default true
            "check_quarantine": true          // optional, default true
        }

        Returns comprehensive validation result with issues, warnings,
        NOI requirements, and recommended actions.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        # Extract parameters
        field_id = request.data.get('field_id')
        product_id = request.data.get('product_id')
        application_date_str = request.data.get('application_date')
        rate_per_acre = request.data.get('rate_per_acre')
        application_method = request.data.get('application_method')
        acres_treated = request.data.get('acres_treated')
        applicator_name = request.data.get('applicator_name')
        applicator_license = request.data.get('applicator_license')
        check_weather = request.data.get('check_weather', True)
        check_quarantine = request.data.get('check_quarantine', True)

        # Validate required fields
        if not all([field_id, product_id, application_date_str, rate_per_acre,
                   application_method, acres_treated]):
            return Response({
                'error': 'Missing required fields',
                'required': ['field_id', 'product_id', 'application_date',
                           'rate_per_acre', 'application_method', 'acres_treated']
            }, status=status.HTTP_400_BAD_REQUEST)

        # Parse date
        try:
            application_date = datetime.strptime(application_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get company for RLS
        company = get_user_company(request.user)
        company_id = company.id if company else None

        # Use compliance service
        service = PesticideComplianceService(company_id=company_id)
        result = service.validate_proposed_application(
            field_id=int(field_id),
            product_id=int(product_id),
            application_date=application_date,
            rate_per_acre=float(rate_per_acre),
            application_method=application_method,
            acres_treated=float(acres_treated),
            applicator_name=applicator_name,
            applicator_license=applicator_license,
            check_weather=check_weather,
            check_quarantine=check_quarantine,
        )

        return Response(result.to_dict())

    @action(detail=False, methods=['get', 'post'])
    def phi_clearance(self, request):
        """
        Check PHI clearance status for fields.

        GET: Check a single field
            Query params: field_id, proposed_harvest_date (optional)

        POST: Check multiple fields or a farm
            Body: {"farm_id": 1, "proposed_harvest_date": "2024-12-20"}
            Or: {"field_ids": [1, 2, 3], "proposed_harvest_date": "2024-12-20"}

        Returns PHI clearance status including earliest harvest date.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        company = get_user_company(request.user)
        company_id = company.id if company else None
        service = PesticideComplianceService(company_id=company_id)

        if request.method == 'GET':
            field_id = request.query_params.get('field_id')
            proposed_date_str = request.query_params.get('proposed_harvest_date')

            if not field_id:
                return Response({
                    'error': 'field_id is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            proposed_date = None
            if proposed_date_str:
                try:
                    proposed_date = datetime.strptime(proposed_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response({
                        'error': 'Invalid date format. Use YYYY-MM-DD'
                    }, status=status.HTTP_400_BAD_REQUEST)

            result = service.calculate_phi_clearance(
                field_id=int(field_id),
                proposed_harvest_date=proposed_date
            )
            return Response(result.to_dict())

        else:  # POST
            farm_id = request.data.get('farm_id')
            field_ids = request.data.get('field_ids')
            proposed_date_str = request.data.get('proposed_harvest_date')

            proposed_date = None
            if proposed_date_str:
                try:
                    proposed_date = datetime.strptime(proposed_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response({
                        'error': 'Invalid date format. Use YYYY-MM-DD'
                    }, status=status.HTTP_400_BAD_REQUEST)

            if farm_id:
                results = service.calculate_phi_for_all_fields(
                    farm_id=int(farm_id),
                    proposed_harvest_date=proposed_date
                )
            elif field_ids:
                results = [
                    service.calculate_phi_clearance(
                        field_id=int(fid),
                        proposed_harvest_date=proposed_date
                    )
                    for fid in field_ids
                ]
            else:
                return Response({
                    'error': 'Either farm_id or field_ids is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'fields': [r.to_dict() for r in results],
                'summary': {
                    'total_fields': len(results),
                    'clear_for_harvest': len([r for r in results if r.is_clear]),
                    'blocked': len([r for r in results if not r.is_clear]),
                }
            })

    @action(detail=False, methods=['get'])
    def rei_status(self, request):
        """
        Get REI (Restricted Entry Interval) status for a field.

        Query params:
            field_id: Required
            check_datetime: Optional ISO datetime (default: now)

        Returns whether field is safe for worker entry.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        field_id = request.query_params.get('field_id')
        check_datetime_str = request.query_params.get('check_datetime')

        if not field_id:
            return Response({
                'error': 'field_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        check_datetime = None
        if check_datetime_str:
            try:
                check_datetime = datetime.fromisoformat(check_datetime_str)
            except ValueError:
                return Response({
                    'error': 'Invalid datetime format. Use ISO format.'
                }, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = PesticideComplianceService(company_id=company_id)
        result = service.get_rei_status(
            field_id=int(field_id),
            check_datetime=check_datetime
        )

        return Response(result.to_dict())

    @action(detail=False, methods=['get', 'post'])
    def noi_requirements(self, request):
        """
        Get Notice of Intent requirements for a product application.

        GET params or POST body:
            product_id: Required
            application_date: Required (YYYY-MM-DD)
            county: Required

        Returns NOI requirements including deadline and submission info.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        if request.method == 'POST':
            product_id = request.data.get('product_id')
            application_date_str = request.data.get('application_date')
            county = request.data.get('county')
        else:
            product_id = request.query_params.get('product_id')
            application_date_str = request.query_params.get('application_date')
            county = request.query_params.get('county')

        if not all([product_id, application_date_str, county]):
            return Response({
                'error': 'product_id, application_date, and county are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            application_date = datetime.strptime(application_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)

        service = PesticideComplianceService()
        result = service.get_noi_requirements(
            product_id=int(product_id),
            application_date=application_date,
            county=county
        )

        return Response(result)

    @action(detail=False, methods=['get'])
    def spray_windows(self, request):
        """
        Find suitable spray windows based on weather forecast.

        Query params:
            farm_id: Required
            days_ahead: Optional (default: 7, max: 7)
            application_method: Optional ('ground' or 'aerial', default: 'ground')

        Returns list of optimal spray windows.
        """
        from api.services.operations import SprayPlanningService

        farm_id = request.query_params.get('farm_id')
        days_ahead = request.query_params.get('days_ahead', '7')
        application_method = request.query_params.get('application_method', 'ground')

        if not farm_id:
            return Response({
                'error': 'farm_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = SprayPlanningService(company_id=company_id)
        windows = service.find_spray_windows(
            farm_id=int(farm_id),
            days_ahead=min(int(days_ahead), 7),
            application_method=application_method
        )

        return Response({
            'farm_id': int(farm_id),
            'windows': [w.to_dict() for w in windows],
            'summary': {
                'total_windows': len(windows),
                'good_windows': len([w for w in windows if w.rating == 'good']),
                'fair_windows': len([w for w in windows if w.rating == 'fair']),
            }
        })

    @action(detail=False, methods=['get'])
    def spray_conditions(self, request):
        """
        Evaluate current spray conditions for a farm.

        Query params:
            farm_id: Required

        Returns current conditions assessment and recommendations.
        """
        from api.services.operations import SprayPlanningService

        farm_id = request.query_params.get('farm_id')

        if not farm_id:
            return Response({
                'error': 'farm_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = SprayPlanningService(company_id=company_id)
        result = service.evaluate_spray_conditions(farm_id=int(farm_id))

        return Response(result.to_dict())

    @action(detail=False, methods=['post'])
    def recommend_timing(self, request):
        """
        Get recommended application timing for a product on a field.

        POST body:
        {
            "field_id": 1,
            "product_id": 5,
            "urgency": "normal"  // 'urgent', 'normal', or 'flexible'
        }

        Returns timing recommendation considering weather, PHI, and REI.
        """
        from api.services.operations import SprayPlanningService

        field_id = request.data.get('field_id')
        product_id = request.data.get('product_id')
        urgency = request.data.get('urgency', 'normal')

        if not all([field_id, product_id]):
            return Response({
                'error': 'field_id and product_id are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        if urgency not in ('urgent', 'normal', 'flexible'):
            urgency = 'normal'

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = SprayPlanningService(company_id=company_id)
        result = service.recommend_application_timing(
            field_id=int(field_id),
            product_id=int(product_id),
            urgency=urgency
        )

        return Response(result.to_dict())

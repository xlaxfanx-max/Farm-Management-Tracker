"""
Audit Log Views for Farm Management Tracker

Add this file to: backend/api/audit_views.py

This module provides API endpoints for viewing and exporting audit logs.
Supports filtering by date range, user, action type, and model name.
"""

from datetime import datetime, timedelta
from io import BytesIO

from django.db import models as db_models
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import AuditLog, User


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_log_list(request):
    """
    List audit logs with filtering and pagination.
    
    Query Parameters:
        - start_date: ISO date string (e.g., "2025-01-01")
        - end_date: ISO date string (e.g., "2025-12-31")
        - user_id: Filter by specific user ID
        - action: Filter by action type (create, update, delete, etc.)
        - model_name: Filter by model/record type (Farm, Field, etc.)
        - search: Search in object_repr field
        - page: Page number (default: 1)
        - page_size: Items per page (default: 25, max: 100)
        - ordering: Field to order by (default: -timestamp)
    
    Returns:
        Paginated list of audit log entries with metadata.
    """
    user = request.user
    company = user.current_company
    
    if not company:
        return Response(
            {"error": "No company associated with user"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Base queryset - filter by company (RLS provides backup protection)
    queryset = AuditLog.objects.filter(company=company).select_related('user')
    
    # Apply filters
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    user_id = request.query_params.get('user_id')
    action = request.query_params.get('action')
    model_name = request.query_params.get('model_name')
    search = request.query_params.get('search')
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            if timezone.is_naive(start):
                start = timezone.make_aware(start)
            queryset = queryset.filter(timestamp__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if timezone.is_naive(end):
                end = timezone.make_aware(end)
            # Include the entire end date
            end = end + timedelta(days=1)
            queryset = queryset.filter(timestamp__lt=end)
        except ValueError:
            pass
    
    if user_id:
        queryset = queryset.filter(user_id=user_id)
    
    if action:
        queryset = queryset.filter(action=action)
    
    if model_name:
        queryset = queryset.filter(model_name__iexact=model_name)
    
    if search:
        queryset = queryset.filter(object_repr__icontains=search)
    
    # Ordering
    ordering = request.query_params.get('ordering', '-timestamp')
    valid_orderings = ['timestamp', '-timestamp', 'user__email', '-user__email', 
                       'action', '-action', 'model_name', '-model_name']
    if ordering in valid_orderings:
        queryset = queryset.order_by(ordering)
    else:
        queryset = queryset.order_by('-timestamp')
    
    # Pagination
    try:
        page = int(request.query_params.get('page', 1))
        page_size = min(int(request.query_params.get('page_size', 25)), 100)
    except ValueError:
        page = 1
        page_size = 25
    
    total_count = queryset.count()
    total_pages = (total_count + page_size - 1) // page_size if total_count > 0 else 1
    
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    
    logs = queryset[start_idx:end_idx]
    
    # Serialize the results
    results = []
    for log in logs:
        results.append({
            'id': log.id,
            'timestamp': log.timestamp.isoformat(),
            'user': {
                'id': log.user.id if log.user else None,
                'email': log.user.email if log.user else 'System',
                'first_name': log.user.first_name if log.user else '',
                'last_name': log.user.last_name if log.user else '',
            } if log.user else None,
            'action': log.action,
            'action_display': dict(AuditLog.ACTION_TYPES).get(log.action, log.action),
            'model_name': log.model_name,
            'object_id': log.object_id,
            'object_repr': log.object_repr,
            'changes': log.changes,
            'ip_address': log.ip_address,
            'user_agent': log.user_agent[:100] if log.user_agent else None,
        })
    
    return Response({
        'results': results,
        'count': total_count,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'has_next': page < total_pages,
        'has_previous': page > 1,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_log_detail(request, pk):
    """
    Get detailed information about a specific audit log entry.
    """
    user = request.user
    company = user.current_company
    
    if not company:
        return Response(
            {"error": "No company associated with user"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        log = AuditLog.objects.select_related('user').get(id=pk, company=company)
    except AuditLog.DoesNotExist:
        return Response(
            {"error": "Audit log entry not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    
    return Response({
        'id': log.id,
        'timestamp': log.timestamp.isoformat(),
        'user': {
            'id': log.user.id if log.user else None,
            'email': log.user.email if log.user else 'System',
            'first_name': log.user.first_name if log.user else '',
            'last_name': log.user.last_name if log.user else '',
        } if log.user else None,
        'action': log.action,
        'action_display': dict(AuditLog.ACTION_TYPES).get(log.action, log.action),
        'model_name': log.model_name,
        'object_id': log.object_id,
        'object_repr': log.object_repr,
        'changes': log.changes,
        'ip_address': log.ip_address,
        'user_agent': log.user_agent,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_log_filters(request):
    """
    Get available filter options for the audit log viewer.
    """
    user = request.user
    company = user.current_company
    
    if not company:
        return Response(
            {"error": "No company associated with user"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    queryset = AuditLog.objects.filter(company=company)
    
    # Get unique users with audit entries
    user_ids = queryset.values_list('user_id', flat=True).distinct()
    users = User.objects.filter(id__in=user_ids).values('id', 'email', 'first_name', 'last_name')
    
    # Get unique action types
    actions = queryset.values_list('action', flat=True).distinct()
    action_choices = [
        {'value': action, 'label': dict(AuditLog.ACTION_TYPES).get(action, action)}
        for action in actions if action
    ]
    
    # Get unique model names
    model_names = queryset.values_list('model_name', flat=True).distinct()
    model_choices = [{'value': name, 'label': name} for name in model_names if name]
    
    # Get date range of audit entries
    date_range = queryset.aggregate(
        earliest=db_models.Min('timestamp'),
        latest=db_models.Max('timestamp')
    )
    
    return Response({
        'users': list(users),
        'actions': sorted(action_choices, key=lambda x: x['label']),
        'model_names': sorted(model_choices, key=lambda x: x['label']),
        'date_range': {
            'earliest': date_range['earliest'].isoformat() if date_range['earliest'] else None,
            'latest': date_range['latest'].isoformat() if date_range['latest'] else None,
        }
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_log_export(request):
    """
    Export audit logs to Excel format.
    """
    user = request.user
    company = user.current_company
    
    if not company:
        return Response(
            {"error": "No company associated with user"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Base queryset
    queryset = AuditLog.objects.filter(company=company).select_related('user')
    
    # Apply filters
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    user_id = request.query_params.get('user_id')
    action = request.query_params.get('action')
    model_name = request.query_params.get('model_name')
    search = request.query_params.get('search')
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            if timezone.is_naive(start):
                start = timezone.make_aware(start)
            queryset = queryset.filter(timestamp__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if timezone.is_naive(end):
                end = timezone.make_aware(end)
            end = end + timedelta(days=1)
            queryset = queryset.filter(timestamp__lt=end)
        except ValueError:
            pass
    
    if user_id:
        queryset = queryset.filter(user_id=user_id)
    
    if action:
        queryset = queryset.filter(action=action)
    
    if model_name:
        queryset = queryset.filter(model_name__iexact=model_name)
    
    if search:
        queryset = queryset.filter(object_repr__icontains=search)
    
    queryset = queryset.order_by('-timestamp')
    
    # Limit export to 10,000 records
    logs = list(queryset[:10000])
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Timestamp",
        "User Email",
        "User Name",
        "Action",
        "Record Type",
        "Record ID",
        "Description",
        "IP Address",
        "Changes Summary"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, log in enumerate(logs, 2):
        changes_summary = ""
        if log.changes:
            if isinstance(log.changes, dict):
                changes_list = []
                for field, change in log.changes.items():
                    if isinstance(change, dict) and 'old' in change and 'new' in change:
                        changes_list.append(f"{field}: {change['old']} → {change['new']}")
                    else:
                        changes_list.append(f"{field}: {change}")
                changes_summary = "; ".join(changes_list[:5])
                if len(log.changes) > 5:
                    changes_summary += f" (+{len(log.changes) - 5} more)"
        
        user_name = ""
        if log.user:
            user_name = f"{log.user.first_name} {log.user.last_name}".strip()
        
        row_data = [
            log.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            log.user.email if log.user else "System",
            user_name,
            dict(AuditLog.ACTION_TYPES).get(log.action, log.action),
            log.model_name,
            log.object_id,
            log.object_repr,
            log.ip_address or "",
            changes_summary
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    # Column widths
    column_widths = [20, 30, 25, 15, 20, 15, 40, 15, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    ws.freeze_panes = "A2"
    
    # Metadata sheet
    ws_meta = wb.create_sheet("Export Info")
    ws_meta['A1'] = "Export Date:"
    ws_meta['B1'] = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_meta['A2'] = "Company:"
    ws_meta['B2'] = company.name
    ws_meta['A3'] = "Exported By:"
    ws_meta['B3'] = user.email
    ws_meta['A4'] = "Total Records:"
    ws_meta['B4'] = len(logs)
    ws_meta['A5'] = "Filters Applied:"
    
    filters_applied = []
    if start_date:
        filters_applied.append(f"From: {start_date}")
    if end_date:
        filters_applied.append(f"To: {end_date}")
    if user_id:
        filters_applied.append(f"User ID: {user_id}")
    if action:
        filters_applied.append(f"Action: {action}")
    if model_name:
        filters_applied.append(f"Record Type: {model_name}")
    if search:
        filters_applied.append(f"Search: {search}")
    
    ws_meta['B5'] = "; ".join(filters_applied) if filters_applied else "None"
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"audit_log_export_{timestamp}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Log the export
    AuditLog.objects.create(
        user=user,
        company=company,
        action='export',
        model_name='AuditLog',
        object_repr=f'Exported {len(logs)} audit log entries',
        changes={'filters': filters_applied, 'record_count': len(logs)}
    )
    
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_log_statistics(request):
    """
    Get statistics about audit log entries.
    """
    user = request.user
    company = user.current_company
    
    if not company:
        return Response(
            {"error": "No company associated with user"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    queryset = AuditLog.objects.filter(company=company)
    
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            if timezone.is_naive(start):
                start = timezone.make_aware(start)
            queryset = queryset.filter(timestamp__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            if timezone.is_naive(end):
                end = timezone.make_aware(end)
            end = end + timedelta(days=1)
            queryset = queryset.filter(timestamp__lt=end)
        except ValueError:
            pass
    
    action_counts = queryset.values('action').annotate(
        count=db_models.Count('id')
    ).order_by('-count')
    
    action_stats = [
        {
            'action': item['action'],
            'action_display': dict(AuditLog.ACTION_TYPES).get(item['action'], item['action']),
            'count': item['count']
        }
        for item in action_counts
    ]
    
    model_counts = queryset.values('model_name').annotate(
        count=db_models.Count('id')
    ).order_by('-count')[:10]
    
    user_counts = queryset.values('user__email', 'user__first_name', 'user__last_name').annotate(
        count=db_models.Count('id')
    ).order_by('-count')[:10]
    
    user_stats = [
        {
            'email': item['user__email'] or 'System',
            'name': f"{item['user__first_name'] or ''} {item['user__last_name'] or ''}".strip() or 'System',
            'count': item['count']
        }
        for item in user_counts
    ]
    
    seven_days_ago = timezone.now() - timedelta(days=7)
    daily_counts = queryset.filter(timestamp__gte=seven_days_ago).extra(
        select={'day': "date(timestamp)"}
    ).values('day').annotate(count=db_models.Count('id')).order_by('day')
    
    return Response({
        'total_count': queryset.count(),
        'by_action': action_stats,
        'by_model': list(model_counts),
        'by_user': user_stats,
        'daily_activity': list(daily_counts),
    })

"""
Compliance Module Views

This module provides API endpoints for compliance management including:
- Compliance Profile (company settings)
- Compliance Deadlines (tracking and management)
- Compliance Alerts (system-generated notifications)
- Licenses (tracking and renewal)
- WPS Training Records
- Central Posting Locations
- REI Posting Records
- Compliance Reports
- Incident Reports
- Notification Preferences
"""

from datetime import date, timedelta, datetime
from decimal import Decimal
from django.db.models import Count, Q, Sum, Avg, Max, Min
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import serializers as drf_serializers

from .permissions import HasCompanyAccess
from .audit_utils import AuditLogMixin

from .models import (
    ComplianceProfile, ComplianceDeadline, ComplianceAlert,
    License, WPSTrainingRecord, CentralPostingLocation, REIPostingRecord,
    ComplianceReport, IncidentReport, NOISubmission, NotificationPreference,
    NotificationLog, Company, PesticideApplication
)

from .serializers import (
    ComplianceProfileSerializer,
    ComplianceDeadlineSerializer, ComplianceDeadlineListSerializer, ComplianceDeadlineCompleteSerializer,
    ComplianceAlertSerializer, ComplianceAlertListSerializer,
    LicenseSerializer, LicenseListSerializer,
    WPSTrainingRecordSerializer, WPSTrainingRecordListSerializer,
    CentralPostingLocationSerializer,
    REIPostingRecordSerializer, REIPostingRecordListSerializer,
    ComplianceReportSerializer, ComplianceReportListSerializer,
    IncidentReportSerializer, IncidentReportListSerializer,
    NotificationPreferenceSerializer, NotificationLogSerializer,
    ComplianceDashboardSerializer,
)


from .view_helpers import get_user_company, require_company


# =============================================================================
# COMPLIANCE PROFILE VIEWSET
# =============================================================================

class ComplianceProfileViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing company compliance profiles.

    Each company has exactly one compliance profile that defines
    which regulatory frameworks apply to them.
    """
    serializer_class = ComplianceProfileSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    http_method_names = ['get', 'put', 'patch']  # No create/delete - auto-created with company

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if company:
            return ComplianceProfile.objects.filter(company=company)
        return ComplianceProfile.objects.none()

    def get_object(self):
        """Get or create profile for current company."""
        company = require_company(self.request.user)
        profile, created = ComplianceProfile.objects.get_or_create(company=company)
        return profile

    def list(self, request, *args, **kwargs):
        """Return the single profile for the current company."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def generate_deadlines(self, request):
        """
        Generate recurring deadlines based on compliance profile settings.
        Creates deadlines for the next 12 months.
        """
        profile = self.get_object()
        company = profile.company
        created_count = 0
        today = date.today()

        # PUR Monthly Reports (CA only)
        if profile.requires_pur_reporting:
            for month_offset in range(12):
                # Due on 10th of each month for previous month's applications
                due_month = today.month + month_offset
                due_year = today.year + (due_month - 1) // 12
                due_month = ((due_month - 1) % 12) + 1
                due_date = date(due_year, due_month, 10)

                if due_date > today:
                    _, created = ComplianceDeadline.objects.get_or_create(
                        company=company,
                        name=f"PUR Report - {due_date.strftime('%B %Y')}",
                        due_date=due_date,
                        defaults={
                            'description': f"Submit Pesticide Use Report for {(due_date.replace(day=1) - timedelta(days=1)).strftime('%B %Y')}",
                            'category': 'reporting',
                            'regulation': 'CA PUR',
                            'frequency': 'monthly',
                            'warning_days': 7,
                            'auto_generated': True,
                            'action_url': '/reports/pur',
                        }
                    )
                    if created:
                        created_count += 1

        # WPS Annual Training Renewal Reminder (if WPS compliance required)
        if profile.requires_wps_compliance:
            annual_due = date(today.year + 1, 1, 1)
            _, created = ComplianceDeadline.objects.get_or_create(
                company=company,
                name="WPS Training Review",
                due_date=annual_due,
                defaults={
                    'description': "Review all WPS training records and schedule renewals for expiring certifications",
                    'category': 'training',
                    'regulation': 'EPA WPS',
                    'frequency': 'annual',
                    'warning_days': 30,
                    'auto_generated': True,
                    'action_url': '/compliance/wps',
                }
            )
            if created:
                created_count += 1

        # SGMA Semi-Annual Reports
        if profile.requires_sgma_reporting:
            # Oct-Mar report due April 1, Apr-Sep report due October 1
            sgma_dates = [
                (date(today.year, 4, 1), "Oct-Mar"),
                (date(today.year, 10, 1), "Apr-Sep"),
                (date(today.year + 1, 4, 1), "Oct-Mar"),
            ]
            for due_date, period in sgma_dates:
                if due_date > today:
                    _, created = ComplianceDeadline.objects.get_or_create(
                        company=company,
                        name=f"SGMA Extraction Report - {period}",
                        due_date=due_date,
                        defaults={
                            'description': f"Submit groundwater extraction report for {period} period",
                            'category': 'reporting',
                            'regulation': 'SGMA',
                            'frequency': 'semi_annual',
                            'warning_days': 30,
                            'auto_generated': True,
                            'action_url': '/water/sgma',
                        }
                    )
                    if created:
                        created_count += 1

        # ILRP Annual Report
        if profile.requires_ilrp_reporting:
            ilrp_due = date(today.year + 1, 3, 1)  # Typically due March 1
            _, created = ComplianceDeadline.objects.get_or_create(
                company=company,
                name=f"ILRP Nitrogen Report - {today.year}",
                due_date=ilrp_due,
                defaults={
                    'description': f"Submit Irrigated Lands Regulatory Program nitrogen management report for {today.year}",
                    'category': 'reporting',
                    'regulation': 'ILRP',
                    'frequency': 'annual',
                    'warning_days': 30,
                    'auto_generated': True,
                    'action_url': '/reports/nitrogen',
                }
            )
            if created:
                created_count += 1

        return Response({
            'message': f'Generated {created_count} new compliance deadlines',
            'created_count': created_count,
        })


# =============================================================================
# COMPLIANCE DEADLINE VIEWSET
# =============================================================================

class ComplianceDeadlineViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing compliance deadlines.

    Supports filtering by status, category, date range, and regulation.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description', 'regulation']
    ordering_fields = ['due_date', 'status', 'category', 'created_at']
    ordering = ['due_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return ComplianceDeadlineListSerializer
        return ComplianceDeadlineSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return ComplianceDeadline.objects.none()

        queryset = ComplianceDeadline.objects.filter(company=company)

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)

        # Filter by regulation
        regulation = self.request.query_params.get('regulation')
        if regulation:
            queryset = queryset.filter(regulation__icontains=regulation)

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(due_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(due_date__lte=end_date)

        # Filter overdue only
        overdue_only = self.request.query_params.get('overdue')
        if overdue_only and overdue_only.lower() == 'true':
            queryset = queryset.filter(status='overdue')

        # Filter due soon only
        due_soon = self.request.query_params.get('due_soon')
        if due_soon and due_soon.lower() == 'true':
            queryset = queryset.filter(status='due_soon')

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark a deadline as completed."""
        deadline = self.get_object()
        serializer = ComplianceDeadlineCompleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        deadline.mark_complete(
            user=request.user,
            notes=serializer.validated_data.get('notes', '')
        )

        # Deactivate any related alerts
        ComplianceAlert.objects.filter(
            related_deadline=deadline,
            is_active=True
        ).update(is_active=False)

        return Response(ComplianceDeadlineSerializer(deadline).data)

    @action(detail=True, methods=['post'])
    def skip(self, request, pk=None):
        """Mark a deadline as skipped/not applicable."""
        deadline = self.get_object()
        deadline.status = 'skipped'
        deadline.completion_notes = request.data.get('notes', 'Marked as not applicable')
        deadline.save()

        return Response(ComplianceDeadlineSerializer(deadline).data)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """Get deadlines due in the next 30 days."""
        company = require_company(request.user)
        end_date = date.today() + timedelta(days=30)

        deadlines = ComplianceDeadline.objects.filter(
            company=company,
            due_date__lte=end_date,
            status__in=['upcoming', 'due_soon']
        ).order_by('due_date')[:20]

        serializer = ComplianceDeadlineListSerializer(deadlines, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get all overdue deadlines."""
        company = require_company(request.user)

        deadlines = ComplianceDeadline.objects.filter(
            company=company,
            status='overdue'
        ).order_by('due_date')

        serializer = ComplianceDeadlineListSerializer(deadlines, many=True)
        return Response(serializer.data)


# =============================================================================
# COMPLIANCE ALERT VIEWSET
# =============================================================================

class ComplianceAlertViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing compliance alerts.

    Alerts are system-generated notifications about compliance issues.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    http_method_names = ['get', 'post']  # No direct create/update/delete
    filter_backends = [filters.OrderingFilter]
    ordering = ['-priority', '-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return ComplianceAlertListSerializer
        return ComplianceAlertSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return ComplianceAlert.objects.none()

        queryset = ComplianceAlert.objects.filter(company=company)

        # By default, only show active alerts
        active_only = self.request.query_params.get('active', 'true')
        if active_only.lower() == 'true':
            queryset = queryset.filter(is_active=True)

        # Filter by priority
        priority = self.request.query_params.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        # Filter by type
        alert_type = self.request.query_params.get('type')
        if alert_type:
            queryset = queryset.filter(alert_type=alert_type)

        # Filter acknowledged
        acknowledged = self.request.query_params.get('acknowledged')
        if acknowledged:
            queryset = queryset.filter(is_acknowledged=acknowledged.lower() == 'true')

        return queryset

    @action(detail=True, methods=['post'])
    def acknowledge(self, request, pk=None):
        """Acknowledge an alert."""
        alert = self.get_object()
        alert.acknowledge(request.user)
        return Response(ComplianceAlertSerializer(alert).data)

    @action(detail=True, methods=['post'])
    def dismiss(self, request, pk=None):
        """Dismiss/deactivate an alert."""
        alert = self.get_object()
        alert.dismiss()
        return Response(ComplianceAlertSerializer(alert).data)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get alert summary counts by priority and type."""
        company = require_company(request.user)

        active_alerts = ComplianceAlert.objects.filter(
            company=company,
            is_active=True
        )

        summary = {
            'total_active': active_alerts.count(),
            'by_priority': {
                'critical': active_alerts.filter(priority='critical').count(),
                'high': active_alerts.filter(priority='high').count(),
                'medium': active_alerts.filter(priority='medium').count(),
                'low': active_alerts.filter(priority='low').count(),
            },
            'unacknowledged': active_alerts.filter(is_acknowledged=False).count(),
        }

        return Response(summary)


# =============================================================================
# LICENSE VIEWSET
# =============================================================================

class LicenseViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing licenses and certificates.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['license_number', 'name_on_license', 'issuing_authority']
    ordering_fields = ['expiration_date', 'license_type', 'status', 'created_at']
    ordering = ['expiration_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return LicenseListSerializer
        return LicenseSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return License.objects.none()

        queryset = License.objects.filter(company=company)

        # Filter by license type
        license_type = self.request.query_params.get('type')
        if license_type:
            queryset = queryset.filter(license_type=license_type)

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by user
        user_id = self.request.query_params.get('user')
        if user_id:
            if user_id == 'company':
                queryset = queryset.filter(user__isnull=True)
            else:
                queryset = queryset.filter(user_id=user_id)

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company)

    @action(detail=False, methods=['get'])
    def expiring(self, request):
        """Get licenses expiring within 90 days."""
        company = require_company(request.user)
        cutoff = date.today() + timedelta(days=90)

        licenses = License.objects.filter(
            company=company,
            expiration_date__lte=cutoff,
            status__in=['active', 'expiring_soon']
        ).order_by('expiration_date')

        serializer = LicenseListSerializer(licenses, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def start_renewal(self, request, pk=None):
        """Mark that renewal process has started."""
        license_obj = self.get_object()
        license_obj.renewal_in_progress = True
        license_obj.renewal_submitted_date = request.data.get('submitted_date')
        license_obj.renewal_notes = request.data.get('notes', '')
        license_obj.save()
        return Response(LicenseSerializer(license_obj).data)


# =============================================================================
# WPS TRAINING VIEWSET
# =============================================================================

class WPSTrainingRecordViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing WPS training records.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['trainee_name', 'trainer_name', 'trainee_employee_id']
    ordering_fields = ['training_date', 'expiration_date', 'training_type']
    ordering = ['-training_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return WPSTrainingRecordListSerializer
        return WPSTrainingRecordSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return WPSTrainingRecord.objects.none()

        queryset = WPSTrainingRecord.objects.filter(company=company)

        # Filter by training type
        training_type = self.request.query_params.get('type')
        if training_type:
            queryset = queryset.filter(training_type=training_type)

        # Filter by trainee name
        trainee = self.request.query_params.get('trainee')
        if trainee:
            queryset = queryset.filter(trainee_name__icontains=trainee)

        # Filter by valid/expired
        valid_only = self.request.query_params.get('valid')
        if valid_only and valid_only.lower() == 'true':
            queryset = queryset.filter(expiration_date__gte=date.today())
        elif valid_only and valid_only.lower() == 'false':
            queryset = queryset.filter(expiration_date__lt=date.today())

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company)

    @action(detail=False, methods=['get'])
    def expiring(self, request):
        """Get training records expiring within 90 days."""
        company = require_company(request.user)
        cutoff = date.today() + timedelta(days=90)

        records = WPSTrainingRecord.objects.filter(
            company=company,
            expiration_date__lte=cutoff,
            expiration_date__gte=date.today()
        ).order_by('expiration_date')

        serializer = WPSTrainingRecordListSerializer(records, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_worker(self, request):
        """Get training records grouped by worker."""
        company = require_company(request.user)

        # Get unique workers with their latest training per type
        workers = WPSTrainingRecord.objects.filter(
            company=company
        ).values('trainee_name').annotate(
            total_trainings=Count('id'),
            latest_training=Max('training_date'),
            next_expiration=Min('expiration_date', filter=Q(expiration_date__gte=date.today()))
        ).order_by('trainee_name')

        return Response(list(workers))

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get WPS compliance dashboard data."""
        company = require_company(request.user)
        today = date.today()

        total_records = WPSTrainingRecord.objects.filter(company=company)
        valid_records = total_records.filter(expiration_date__gte=today)
        expired_records = total_records.filter(expiration_date__lt=today)
        expiring_soon = total_records.filter(
            expiration_date__gte=today,
            expiration_date__lte=today + timedelta(days=90)
        )

        # Unique workers
        unique_workers = total_records.values('trainee_name').distinct().count()
        workers_with_valid = valid_records.values('trainee_name').distinct().count()

        return Response({
            'total_records': total_records.count(),
            'valid_records': valid_records.count(),
            'expired_records': expired_records.count(),
            'expiring_soon': expiring_soon.count(),
            'unique_workers': unique_workers,
            'workers_with_valid_training': workers_with_valid,
            'by_type': list(
                valid_records.values('training_type').annotate(
                    count=Count('id')
                ).order_by('training_type')
            ),
        })


# =============================================================================
# CENTRAL POSTING LOCATION VIEWSET
# =============================================================================

class CentralPostingLocationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing WPS central posting locations.
    """
    serializer_class = CentralPostingLocationSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return CentralPostingLocation.objects.none()

        queryset = CentralPostingLocation.objects.filter(
            company=company,
            active=True
        ).select_related('farm', 'last_verified_by')

        # Filter by farm
        farm_id = self.request.query_params.get('farm')
        if farm_id:
            queryset = queryset.filter(farm_id=farm_id)

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company)

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Mark location as verified."""
        location = self.get_object()
        notes = request.data.get('notes', '')
        location.verify(request.user, notes)
        return Response(CentralPostingLocationSerializer(location).data)


# =============================================================================
# REI POSTING RECORD VIEWSET
# =============================================================================

class REIPostingRecordViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing REI posting records.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    http_method_names = ['get', 'post']

    def get_serializer_class(self):
        if self.action == 'list':
            return REIPostingRecordListSerializer
        return REIPostingRecordSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return REIPostingRecord.objects.none()

        queryset = REIPostingRecord.objects.filter(
            application__field__farm__company=company
        ).select_related(
            'application', 'application__field', 'application__product',
            'posted_by', 'removed_by'
        )

        # Filter active only
        active_only = self.request.query_params.get('active')
        if active_only and active_only.lower() == 'true':
            queryset = queryset.filter(rei_end_datetime__gt=timezone.now())

        return queryset.order_by('-rei_end_datetime')

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get currently active REIs."""
        company = require_company(request.user)

        active_reis = REIPostingRecord.objects.filter(
            application__field__farm__company=company,
            rei_end_datetime__gt=timezone.now()
        ).select_related(
            'application', 'application__field', 'application__product'
        ).order_by('rei_end_datetime')

        serializer = REIPostingRecordListSerializer(active_reis, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_posted(self, request, pk=None):
        """Mark REI signs as posted."""
        record = self.get_object()
        record.mark_posted(request.user)
        return Response(REIPostingRecordSerializer(record).data)

    @action(detail=True, methods=['post'])
    def mark_removed(self, request, pk=None):
        """Mark REI signs as removed."""
        record = self.get_object()
        record.mark_removed(request.user)
        return Response(REIPostingRecordSerializer(record).data)


# =============================================================================
# COMPLIANCE REPORT VIEWSET
# =============================================================================

class ComplianceReportViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing compliance reports.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-reporting_period_end', '-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return ComplianceReportListSerializer
        return ComplianceReportSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return ComplianceReport.objects.none()

        queryset = ComplianceReport.objects.filter(company=company)

        # Filter by report type
        report_type = self.request.query_params.get('type')
        if report_type:
            queryset = queryset.filter(report_type=report_type)

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company, created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        """Run validation on a report."""
        report = self.get_object()

        # TODO: Implement actual validation logic based on report type
        # For now, just mark as validated
        report.validation_run_at = timezone.now()
        report.validation_errors = []
        report.validation_warnings = []
        report.is_valid = True
        report.status = 'ready'
        report.save()

        return Response(ComplianceReportSerializer(report).data)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Mark report as submitted."""
        report = self.get_object()

        if not report.can_submit:
            return Response(
                {'error': 'Report is not ready for submission. Please validate first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        report.status = 'submitted'
        report.submitted_at = timezone.now()
        report.submitted_by = request.user
        report.submission_method = request.data.get('method', '')
        report.submission_reference = request.data.get('reference', '')
        report.save()

        # Mark related deadline as complete if exists
        if report.related_deadline:
            report.related_deadline.mark_complete(request.user, f"Report submitted: {report.submission_reference}")

        return Response(ComplianceReportSerializer(report).data)

    @action(detail=False, methods=['post'])
    def generate_pur(self, request):
        """Auto-generate a PUR report from existing pesticide application records."""
        company = require_company(request.user)
        period_start_str = request.data.get('period_start')
        period_end_str = request.data.get('period_end')

        if not period_start_str or not period_end_str:
            return Response(
                {'error': 'period_start and period_end are required (YYYY-MM-DD format)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            period_start = datetime.strptime(period_start_str, '%Y-%m-%d').date()
            period_end = datetime.strptime(period_end_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        applications = PesticideApplication.objects.filter(
            field__farm__company=company,
            application_date__gte=period_start,
            application_date__lte=period_end,
        ).select_related('field', 'field__farm', 'product', 'product__active_ingredient')

        app_data = []
        validation_errors = []
        validation_warnings = []

        for app in applications:
            row = {
                'id': app.id,
                'date': app.application_date.isoformat(),
                'field_name': str(app.field),
                'farm_name': str(app.field.farm),
                'product_name': getattr(app.product, 'product_name', 'Unknown'),
                'epa_reg_number': getattr(app.product, 'epa_registration_number', '') or '',
                'active_ingredient': getattr(getattr(app.product, 'active_ingredient', None), 'name', '') or '',
                'amount_applied': str(app.amount_used or ''),
                'unit': app.unit_of_measure or '',
                'acres_treated': str(app.acres_treated or ''),
                'applicator_name': app.applicator_name or '',
                'applicator_license': app.applicator_license_no or '',
                'start_time': app.start_time.strftime('%H:%M') if app.start_time else '',
                'end_time': app.end_time.strftime('%H:%M') if app.end_time else '',
            }
            # Flag missing required fields
            missing = []
            if not row['epa_reg_number']:
                missing.append('EPA registration number')
            if not row['applicator_license']:
                missing.append('applicator license')
            if not row['acres_treated']:
                missing.append('acres treated')
            if missing:
                validation_warnings.append(f"Application {app.id} ({row['date']} - {row['field_name']}): missing {', '.join(missing)}")
                row['has_warnings'] = True
                row['missing_fields'] = missing
            else:
                row['has_warnings'] = False
                row['missing_fields'] = []
            app_data.append(row)

        return Response({
            'period_start': period_start_str,
            'period_end': period_end_str,
            'application_count': len(app_data),
            'applications': app_data,
            'validation_errors': validation_errors,
            'validation_warnings': validation_warnings,
            'ready_for_report': len(validation_errors) == 0,
        })


# =============================================================================

# INCIDENT REPORT VIEWSET
# =============================================================================

class IncidentReportViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing incident reports.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'description', 'location_description']
    ordering_fields = ['incident_date', 'severity', 'status']
    ordering = ['-incident_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return IncidentReportListSerializer
        return IncidentReportSerializer

    def get_queryset(self):
        company = get_user_company(self.request.user)
        if not company:
            return IncidentReport.objects.none()

        queryset = IncidentReport.objects.filter(company=company)

        # Filter by type
        incident_type = self.request.query_params.get('type')
        if incident_type:
            queryset = queryset.filter(incident_type=incident_type)

        # Filter by severity
        severity = self.request.query_params.get('severity')
        if severity:
            queryset = queryset.filter(severity=severity)

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company, reported_by=self.request.user)

    @action(detail=True, methods=['post'])
    def start_investigation(self, request, pk=None):
        """Start investigation on an incident."""
        incident = self.get_object()
        incident.status = 'investigating'
        incident.investigator = request.user
        incident.save()
        return Response(IncidentReportSerializer(incident).data)

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Mark incident as resolved."""
        incident = self.get_object()
        incident.status = 'resolved'
        incident.resolved_date = date.today()
        incident.resolved_by = request.user
        incident.resolution_summary = request.data.get('summary', '')
        incident.corrective_actions = request.data.get('corrective_actions', incident.corrective_actions)
        incident.preventive_measures = request.data.get('preventive_measures', incident.preventive_measures)
        incident.save()
        return Response(IncidentReportSerializer(incident).data)


# =============================================================================
# NOTIFICATION PREFERENCE VIEWSET
# =============================================================================

class NotificationPreferenceViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing user notification preferences.
    """
    serializer_class = NotificationPreferenceSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'put', 'patch']

    def get_queryset(self):
        return NotificationPreference.objects.filter(user=self.request.user)

    def get_object(self):
        """Get or create preferences for current user."""
        prefs, created = NotificationPreference.objects.get_or_create(user=self.request.user)
        return prefs

    def list(self, request, *args, **kwargs):
        """Return the single preference object for current user."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


# =============================================================================
# COMPLIANCE DASHBOARD VIEW
# =============================================================================

class ComplianceDashboardViewSet(viewsets.ViewSet):
    """
    API endpoint for the compliance dashboard overview.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def list(self, request):
        """Get comprehensive compliance dashboard data."""
        company = require_company(request.user)
        today = date.today()

        # Auto-generate deadlines on first visit if not yet done
        try:
            profile = ComplianceProfile.objects.get(company=company)
            if not profile.deadlines_auto_populated:
                deadline_count = ComplianceDeadline.objects.filter(company=company).count()
                if deadline_count == 0:
                    from .tasks.compliance_tasks import generate_recurring_deadlines
                    generate_recurring_deadlines(company_id=company.id)
                profile.deadlines_auto_populated = True
                profile.save(update_fields=["deadlines_auto_populated"])
        except ComplianceProfile.DoesNotExist:
            pass
        except Exception:
            pass  # Never crash the dashboard due to deadline generation failure

        # Deadline statistics
        deadlines = ComplianceDeadline.objects.filter(company=company)
        deadlines_this_month = deadlines.filter(
            due_date__year=today.year,
            due_date__month=today.month,
            status__in=['upcoming', 'due_soon']
        ).count()
        overdue_deadlines = deadlines.filter(status='overdue').count()

        # License statistics
        licenses = License.objects.filter(company=company)
        expiring_licenses = licenses.filter(
            expiration_date__lte=today + timedelta(days=90),
            status__in=['active', 'expiring_soon']
        ).count()

        # Alert statistics
        active_alerts = ComplianceAlert.objects.filter(
            company=company,
            is_active=True
        ).count()

        # Training statistics
        training = WPSTrainingRecord.objects.filter(company=company)
        expiring_training = training.filter(
            expiration_date__lte=today + timedelta(days=90),
            expiration_date__gte=today
        ).count()

        # Calculate overall status and score
        critical_issues = overdue_deadlines + licenses.filter(status='expired').count()
        warning_issues = expiring_licenses + expiring_training

        if critical_issues > 0:
            overall_status = 'critical'
            score = max(0, 100 - (critical_issues * 20) - (warning_issues * 5))
        elif warning_issues > 0:
            overall_status = 'warning'
            score = max(50, 100 - (warning_issues * 10))
        else:
            overall_status = 'good'
            score = 100

        # Compliance by category
        by_category = {}
        for category, _ in ComplianceDeadline.CATEGORY_CHOICES:
            cat_deadlines = deadlines.filter(category=category)
            by_category[category] = {
                'pending': cat_deadlines.filter(status__in=['upcoming', 'due_soon']).count(),
                'overdue': cat_deadlines.filter(status='overdue').count(),
                'completed': cat_deadlines.filter(status='completed').count(),
            }
            # Determine category status
            if by_category[category]['overdue'] > 0:
                by_category[category]['status'] = 'critical'
            elif by_category[category]['pending'] > 5:
                by_category[category]['status'] = 'warning'
            else:
                by_category[category]['status'] = 'good'

        # Get upcoming deadlines
        upcoming_deadlines = deadlines.filter(
            due_date__lte=today + timedelta(days=30),
            status__in=['upcoming', 'due_soon']
        ).order_by('due_date')[:10]

        # Get active alerts
        alerts = ComplianceAlert.objects.filter(
            company=company,
            is_active=True
        ).order_by('-priority', '-created_at')[:10]

        # Get expiring licenses
        exp_licenses = licenses.filter(
            expiration_date__lte=today + timedelta(days=90),
            status__in=['active', 'expiring_soon']
        ).order_by('expiration_date')[:5]

        # Get expiring training
        exp_training = training.filter(
            expiration_date__lte=today + timedelta(days=90),
            expiration_date__gte=today
        ).order_by('expiration_date')[:5]

        data = {
            'overall_status': overall_status,
            'score': score,
            'summary': {
                'deadlines_this_month': deadlines_this_month,
                'overdue_items': overdue_deadlines,
                'expiring_licenses': expiring_licenses,
                'active_alerts': active_alerts,
                'expiring_training': expiring_training,
            },
            'by_category': by_category,
            'upcoming_deadlines': ComplianceDeadlineListSerializer(upcoming_deadlines, many=True).data,
            'active_alerts': ComplianceAlertListSerializer(alerts, many=True).data,
            'expiring_licenses': LicenseListSerializer(exp_licenses, many=True).data,
            'expiring_training': WPSTrainingRecordListSerializer(exp_training, many=True).data,
        }

        return Response(data)

    @action(detail=False, methods=['get'])
    def calendar(self, request):
        """Get calendar view of compliance deadlines."""
        company = require_company(request.user)

        # Get date range from params (default to current month)
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = date.today().replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # End of next month
            if start_date.month == 12:
                end_date = date(start_date.year + 1, 2, 1) - timedelta(days=1)
            else:
                end_date = date(start_date.year, start_date.month + 2, 1) - timedelta(days=1)

        deadlines = ComplianceDeadline.objects.filter(
            company=company,
            due_date__gte=start_date,
            due_date__lte=end_date
        ).order_by('due_date')

        # Group by date
        calendar_data = {}
        for deadline in deadlines:
            date_key = deadline.due_date.isoformat()
            if date_key not in calendar_data:
                calendar_data[date_key] = {
                    'date': date_key,
                    'deadlines': [],
                }
            calendar_data[date_key]['deadlines'].append(
                ComplianceDeadlineListSerializer(deadline).data
            )

        return Response(list(calendar_data.values()))

    @action(detail=False, methods=['get'])
    def smart_score(self, request):
        """Additive compliance score - starts at 0, earns points for completeness."""
        company = require_company(request.user)
        today = date.today()

        breakdown = []
        total_score = 0

        # Licenses exist (15 pts)
        licenses = License.objects.filter(company=company)
        active_licenses = licenses.filter(status__in=['active', 'expiring_soon'])
        passed = active_licenses.exists()
        earned = 15 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'licenses_exist', 'label': 'Active Licenses',
            'earned': earned, 'possible': 15, 'passed': passed,
            'action': None if passed else 'Add your QAL or PCA license',
            'action_key': None if passed else 'compliance-licenses',
        })

        # Licenses current (10 pts)
        expired_licenses = licenses.filter(status='expired').count()
        passed = active_licenses.exists() and expired_licenses == 0
        earned = 10 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'licenses_current', 'label': 'No Expired Licenses',
            'earned': earned, 'possible': 10, 'passed': passed,
            'action': None if passed else f'{expired_licenses} expired license(s)',
            'action_key': None if passed else 'compliance-licenses',
        })

        # WPS training exists (10 pts)
        training = WPSTrainingRecord.objects.filter(company=company)
        passed = training.exists()
        earned = 10 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'wps_training_exist', 'label': 'WPS Training Records',
            'earned': earned, 'possible': 10, 'passed': passed,
            'action': None if passed else 'Add worker safety training records',
            'action_key': None if passed else 'compliance-wps',
        })

        # WPS training current (10 pts)
        expired_training = training.filter(expiration_date__lt=today).count()
        passed = training.exists() and expired_training == 0
        earned = 10 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'wps_current', 'label': 'WPS Training Current',
            'earned': earned, 'possible': 10, 'passed': passed,
            'action': None if passed else f'{expired_training} expired training record(s)',
            'action_key': None if passed else 'compliance-wps',
        })

        # Deadlines populated (10 pts)
        deadlines = ComplianceDeadline.objects.filter(company=company)
        passed = deadlines.exists()
        earned = 10 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'deadlines_populated', 'label': 'Compliance Calendar',
            'earned': earned, 'possible': 10, 'passed': passed,
            'action': None if passed else 'Generate your compliance calendar',
            'action_key': None if passed else 'compliance-deadlines',
        })

        # No overdue deadlines (15 pts)
        overdue_count = deadlines.filter(status='overdue').count()
        passed = overdue_count == 0
        earned = 15 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'no_overdue', 'label': 'No Overdue Deadlines',
            'earned': earned, 'possible': 15, 'passed': passed,
            'action': None if passed else f'{overdue_count} overdue deadline(s)',
            'action_key': None if passed else 'compliance-deadlines',
        })

        # PUR current - last month submitted (10 pts)
        first_of_month = today.replace(day=1)
        if first_of_month.month == 1:
            last_month_start = date(first_of_month.year - 1, 12, 1)
        else:
            last_month_start = date(first_of_month.year, first_of_month.month - 1, 1)
        last_month_end = first_of_month - timedelta(days=1)
        pur_submitted = ComplianceReport.objects.filter(
            company=company,
            report_type='pur_monthly',
            status__in=['submitted', 'accepted'],
            reporting_period_start__gte=last_month_start,
            reporting_period_end__lte=last_month_end,
        ).exists()
        passed = pur_submitted
        earned = 10 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'pur_current', 'label': 'PUR Reports Current',
            'earned': earned, 'possible': 10, 'passed': passed,
            'action': None if passed else f'{last_month_start.strftime("%B %Y")} PUR not submitted',
            'action_key': None if passed else 'compliance-reports',
        })

        # FSMA setup (5 pts)
        try:
            from .models.fsma import FSMAFacility, WaterAssessment
            has_fsma = FSMAFacility.objects.filter(company=company).exists() or \
                       WaterAssessment.objects.filter(company=company).exists()
        except Exception:
            has_fsma = False
        passed = has_fsma
        earned = 5 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'fsma_setup', 'label': 'FSMA Setup',
            'earned': earned, 'possible': 5, 'passed': passed,
            'action': None if passed else 'Set up FSMA facilities or water assessment',
            'action_key': None if passed else 'compliance-fsma',
        })

        # Water testing current (5 pts)
        try:
            from .models.water import WaterSample
            ninety_days_ago = today - timedelta(days=90)
            has_recent_water = WaterSample.objects.filter(
                farm__company=company,
                sample_date__gte=ninety_days_ago
            ).exists()
        except Exception:
            has_recent_water = False
        passed = has_recent_water
        earned = 5 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'water_testing', 'label': 'Water Testing Current',
            'earned': earned, 'possible': 5, 'passed': passed,
            'action': None if passed else 'No water tests in the last 90 days',
            'action_key': None if passed else 'water-management',
        })

        # Posting locations configured (5 pts)
        has_posting = CentralPostingLocation.objects.filter(company=company).exists()
        passed = has_posting
        earned = 5 if passed else 0
        total_score += earned
        breakdown.append({
            'key': 'posting_locations', 'label': 'Posting Locations',
            'earned': earned, 'possible': 5, 'passed': passed,
            'action': None if passed else 'Set up central WPS posting locations',
            'action_key': None if passed else 'compliance-settings',
        })

        gap_items = [
            {'key': b['key'], 'action': b['action'], 'action_key': b['action_key'], 'points': b['possible']}
            for b in breakdown if not b['passed']
        ]
        gap_items.sort(key=lambda x: -x['points'])

        passed_count = sum(1 for b in breakdown if b['passed'])
        total_count = len(breakdown)

        return Response({
            'score': total_score,
            'setup_completeness': int((passed_count / total_count) * 100),
            'passed_count': passed_count,
            'total_count': total_count,
            'score_breakdown': breakdown,
            'gap_items': gap_items,
        })

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Everything the farmer needs to know and do today."""
        company = require_company(request.user)
        today = date.today()

        # Overdue deadlines
        overdue = ComplianceDeadline.objects.filter(
            company=company, status='overdue'
        ).order_by('due_date')[:10]

        # Due today
        due_today = ComplianceDeadline.objects.filter(
            company=company,
            due_date=today,
            status__in=['upcoming', 'due_soon']
        ).order_by('title')[:10]

        # Due this week
        week_from_now = today + timedelta(days=7)
        due_this_week = ComplianceDeadline.objects.filter(
            company=company,
            due_date__gt=today,
            due_date__lte=week_from_now,
            status__in=['upcoming', 'due_soon']
        ).order_by('due_date')[:10]

        # Expired licenses
        expired_lic = License.objects.filter(
            company=company, status='expired'
        ).order_by('expiration_date')[:5]

        # Expiring training (30 days)
        thirty_days = today + timedelta(days=30)
        expiring_training = WPSTrainingRecord.objects.filter(
            company=company,
            expiration_date__gte=today,
            expiration_date__lte=thirty_days
        ).order_by('expiration_date')[:5]

        # Active REIs
        active_reis = REIPostingRecord.objects.filter(
            company=company,
            rei_end_datetime__gt=timezone.now(),
            removed_at__isnull=True
        ).select_related('application__field', 'application__product')[:10]

        # PHI blocked fields - applications where PHI end date is in future
        from .models import PesticideApplication, Field
        today_dt = timezone.now()
        phi_apps = PesticideApplication.objects.filter(
            field__farm__company=company,
            phi_end_date__gt=today
        ).select_related('field', 'product')[:10]

        # Pending PUR month
        first_of_month = today.replace(day=1)
        if first_of_month.month == 1:
            last_month_start = date(first_of_month.year - 1, 12, 1)
        else:
            last_month_start = date(first_of_month.year, first_of_month.month - 1, 1)
        last_month_end = first_of_month - timedelta(days=1)
        pur_submitted = ComplianceReport.objects.filter(
            company=company,
            report_type='pur_monthly',
            status__in=['submitted', 'accepted'],
            reporting_period_start__gte=last_month_start,
        ).exists()
        pending_pur = None if pur_submitted else last_month_start.strftime('%B %Y')

        # Pending NOI count
        from .models import NOISubmission
        pending_noi_count = NOISubmission.objects.filter(
            company=company, status='pending'
        ).count()

        # Build quick_wins
        quick_wins = []
        if pending_pur:
            quick_wins.append({
                'action': f'Submit {pending_pur} PUR Report',
                'url_key': 'compliance-reports',
                'priority': 'high',
                'icon': 'FileText',
            })
        if expired_lic.exists():
            quick_wins.append({
                'action': 'Renew expired licenses',
                'url_key': 'compliance-licenses',
                'priority': 'high',
                'icon': 'Award',
            })
        if overdue.exists():
            quick_wins.append({
                'action': f'Address {overdue.count()} overdue deadline(s)',
                'url_key': 'compliance-deadlines',
                'priority': 'high',
                'icon': 'AlertTriangle',
            })
        if pending_noi_count > 0:
            quick_wins.append({
                'action': f'File {pending_noi_count} pending NOI(s)',
                'url_key': 'compliance-reports',
                'priority': 'medium',
                'icon': 'ClipboardCheck',
            })

        all_clear = (
            not overdue.exists() and
            not due_today.exists() and
            not expired_lic.exists() and
            not expiring_training.exists() and
            not active_reis.exists() and
            not phi_apps.exists() and
            pending_pur is None and
            pending_noi_count == 0
        )

        def serialize_deadline(d):
            return {
                'id': d.id,
                'title': d.title,
                'due_date': d.due_date.isoformat(),
                'category': d.category,
                'days_overdue': (today - d.due_date).days if d.due_date < today else 0,
            }

        def serialize_rei(r):
            remaining = (r.rei_end_datetime - timezone.now()).total_seconds()
            return {
                'id': r.id,
                'field_name': str(r.application.field) if r.application else 'Unknown',
                'product_name': str(r.application.product) if r.application else 'Unknown',
                'rei_end_datetime': r.rei_end_datetime.isoformat(),
                'time_remaining_seconds': max(0, int(remaining)),
                'is_active': remaining > 0,
            }

        def serialize_phi(app):
            return {
                'field_name': str(app.field),
                'product_name': str(app.product) if hasattr(app, 'product') else 'Unknown',
                'clear_date': app.phi_end_date.isoformat() if app.phi_end_date else None,
                'days_remaining': (app.phi_end_date - today).days if app.phi_end_date else 0,
            }

        return Response({
            'date': today.isoformat(),
            'all_clear': all_clear,
            'overdue_deadlines': [serialize_deadline(d) for d in overdue],
            'due_today': [serialize_deadline(d) for d in due_today],
            'due_this_week': [serialize_deadline(d) for d in due_this_week],
            'expired_licenses': LicenseListSerializer(expired_lic, many=True).data,
            'expiring_training': WPSTrainingRecordListSerializer(expiring_training, many=True).data,
            'active_reis': [serialize_rei(r) for r in active_reis],
            'phi_blocked_fields': [serialize_phi(a) for a in phi_apps],
            'pending_pur_month': pending_pur,
            'pending_noi_count': pending_noi_count,
            'quick_wins': quick_wins,
        })

    @action(detail=False, methods=['get'])
    def suggestions(self, request):
        """Proactive compliance suggestions based on farm data."""
        company = require_company(request.user)
        today = date.today()
        suggestions = []

        # PUR not submitted last month
        first_of_month = today.replace(day=1)
        if first_of_month.month == 1:
            last_month_start = date(first_of_month.year - 1, 12, 1)
        else:
            last_month_start = date(first_of_month.year, first_of_month.month - 1, 1)
        pur_submitted = ComplianceReport.objects.filter(
            company=company,
            report_type='pur_monthly',
            status__in=['submitted', 'accepted'],
            reporting_period_start__gte=last_month_start,
        ).exists()
        if not pur_submitted:
            app_count = PesticideApplication.objects.filter(
                field__farm__company=company,
                application_date__gte=last_month_start,
                application_date__lt=first_of_month,
            ).count()
            if app_count > 0:
                suggestions.append({
                    'id': 'pur_pending',
                    'priority': 'high',
                    'icon': 'FileText',
                    'title': f'{last_month_start.strftime("%B %Y")} PUR Not Submitted',
                    'detail': f'{app_count} application(s) recorded - generate report now',
                    'action_label': 'Generate PUR',
                    'action_key': 'compliance-reports',
                })

        # Expiring licenses in 60 days
        sixty_days = today + timedelta(days=60)
        expiring_lic = License.objects.filter(
            company=company,
            expiration_date__gte=today,
            expiration_date__lte=sixty_days,
            status__in=['active', 'expiring_soon']
        ).order_by('expiration_date')[:3]
        for lic in expiring_lic:
            days = (lic.expiration_date - today).days
            suggestions.append({
                'id': f'license_expiring_{lic.id}',
                'priority': 'high' if days <= 30 else 'medium',
                'icon': 'Award',
                'title': f'{lic.get_license_type_display()} Expiring Soon',
                'detail': f'{lic.holder_name} - expires in {days} days',
                'action_label': 'Renew License',
                'action_key': 'compliance-licenses',
            })

        # WPS training expiring in 30 days
        thirty_days = today + timedelta(days=30)
        expiring_training = WPSTrainingRecord.objects.filter(
            company=company,
            expiration_date__gte=today,
            expiration_date__lte=thirty_days
        ).order_by('expiration_date')[:3]
        for rec in expiring_training:
            days = (rec.expiration_date - today).days
            suggestions.append({
                'id': f'training_expiring_{rec.id}',
                'priority': 'medium',
                'icon': 'Users',
                'title': 'WPS Training Expiring Soon',
                'detail': f'{rec.worker_name} - expires in {days} days',
                'action_label': 'Update Training',
                'action_key': 'compliance-wps',
            })

        # Overdue deadlines
        overdue_count = ComplianceDeadline.objects.filter(
            company=company, status='overdue'
        ).count()
        if overdue_count > 0:
            suggestions.append({
                'id': 'overdue_deadlines',
                'priority': 'high',
                'icon': 'AlertTriangle',
                'title': f'{overdue_count} Overdue Deadline(s)',
                'detail': 'Address overdue compliance items',
                'action_label': 'View Deadlines',
                'action_key': 'compliance-deadlines',
            })

        # Water testing overdue (90+ days)
        try:
            from .models.water import WaterSample
            ninety_days_ago = today - timedelta(days=90)
            stale_sources = WaterSample.objects.filter(
                farm__company=company
            ).values('source_id').annotate(last=Max('sample_date')).filter(last__lt=ninety_days_ago)
            if stale_sources.exists():
                suggestions.append({
                    'id': 'water_overdue',
                    'priority': 'medium',
                    'icon': 'Droplets',
                    'title': 'Water Testing Overdue',
                    'detail': f'{stale_sources.count()} source(s) not tested in 90+ days',
                    'action_label': 'View Water',
                    'action_key': 'water-management',
                })
        except Exception:
            pass

        return Response({'suggestions': suggestions})


# =============================================================================

# INSPECTOR REPORT VIEWSET
# =============================================================================

class InspectorReportViewSet(viewsets.ViewSet):
    """
    Generate inspector-ready compliance reports.
    One endpoint to rule them all - consolidated compliance PDF/JSON.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def list(self, request):
        """Get the full inspector report as JSON."""
        company = require_company(request.user)
        farm_id = request.query_params.get('farm_id')

        from .services.compliance.inspector_report_generator import InspectorReportGenerator
        generator = InspectorReportGenerator(company, farm_id=farm_id)
        data = generator.generate_report_data()
        return Response(data)

    @action(detail=False, methods=['get'])
    def pdf(self, request):
        """Download the inspector report as a PDF."""
        company = require_company(request.user)
        farm_id = request.query_params.get('farm_id')

        from .services.compliance.inspector_report_generator import InspectorReportGenerator
        generator = InspectorReportGenerator(company, farm_id=farm_id)
        pdf_buffer = generator.generate_pdf()

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="compliance_report_{company.name}_{date.today().isoformat()}.pdf"'
        )
        return response

    @action(detail=False, methods=['get'])
    def checklist(self, request):
        company = require_company(request.user)
        today = date.today()
        sections = []

        # SECTION 1: Licenses & Permits
        licenses = License.objects.filter(company=company)
        active_lic = licenses.filter(status__in=['active', 'expiring_soon'])
        expired_lic = licenses.filter(status='expired')
        sections.append({
            'title': 'Licenses & Permits',
            'icon': 'Award',
            'items': [
                {
                    'label': 'Active applicator license on file (QAL/PCA)',
                    'is_passed': active_lic.exists(),
                    'detail': f"{active_lic.count()} active license(s)" if active_lic.exists() else 'No active licenses found',
                    'action_label': 'Add License' if not active_lic.exists() else None,
                    'action_key': 'compliance-licenses',
                },
                {
                    'label': 'No expired licenses',
                    'is_passed': expired_lic.count() == 0,
                    'detail': 'All licenses current' if expired_lic.count() == 0 else f"{expired_lic.count()} expired",
                    'action_label': 'Renew Licenses' if expired_lic.count() > 0 else None,
                    'action_key': 'compliance-licenses',
                },
            ],
        })

        # SECTION 2: Pesticide Use Records
        first_of_month = today.replace(day=1)
        if first_of_month.month == 1:
            last_month_start = date(first_of_month.year - 1, 12, 1)
        else:
            last_month_start = date(first_of_month.year, first_of_month.month - 1, 1)
        last_month_end = first_of_month - timedelta(days=1)
        pur_ok = ComplianceReport.objects.filter(
            company=company,
            report_type='pur_monthly',
            status__in=['submitted', 'accepted'],
            reporting_period_start__gte=last_month_start,
        ).exists()
        apps_missing_epa = PesticideApplication.objects.filter(
            field__farm__company=company,
        ).exclude(product__epa_registration_number__isnull=False).count()
        apps_missing_license = PesticideApplication.objects.filter(
            field__farm__company=company,
            applicator_license_no=''
        ).count()

        sections.append({
            'title': 'Pesticide Use Records',
            'icon': 'FileText',
            'items': [
                {
                    'label': f"PUR submitted for {last_month_start.strftime('%B %Y')}",
                    'is_passed': pur_ok,
                    'detail': 'Submitted' if pur_ok else 'Not submitted',
                    'action_label': 'Generate PUR' if not pur_ok else None,
                    'action_key': 'compliance-reports',
                },
                {
                    'label': 'All applications have EPA registration numbers',
                    'is_passed': apps_missing_epa == 0,
                    'detail': 'All complete' if apps_missing_epa == 0 else f"{apps_missing_epa} application(s) missing",
                    'action_label': 'Review Applications' if apps_missing_epa > 0 else None,
                    'action_key': 'compliance-reports',
                },
                {
                    'label': 'All applications have applicator license numbers',
                    'is_passed': apps_missing_license == 0,
                    'detail': 'All complete' if apps_missing_license == 0 else f"{apps_missing_license} missing",
                    'action_label': 'Review Applications' if apps_missing_license > 0 else None,
                    'action_key': 'compliance-reports',
                },
            ],
        })

        # SECTION 3: Worker Protection
        wps_training = WPSTrainingRecord.objects.filter(company=company)
        expired_wps = wps_training.filter(expiration_date__lt=today).count()
        posting_locs = CentralPostingLocation.objects.filter(company=company)
        sections.append({
            'title': 'Worker Protection Standard',
            'icon': 'Users',
            'items': [
                {
                    'label': 'WPS training records on file',
                    'is_passed': wps_training.exists(),
                    'detail': f"{wps_training.count()} training record(s)" if wps_training.exists() else 'No records found',
                    'action_label': 'Add Training' if not wps_training.exists() else None,
                    'action_key': 'compliance-wps',
                },
                {
                    'label': 'All WPS training current (not expired)',
                    'is_passed': expired_wps == 0,
                    'detail': 'All current' if expired_wps == 0 else f"{expired_wps} expired",
                    'action_label': 'Update Training' if expired_wps > 0 else None,
                    'action_key': 'compliance-wps',
                },
                {
                    'label': 'Central posting locations configured',
                    'is_passed': posting_locs.exists(),
                    'detail': f"{posting_locs.count()} location(s)" if posting_locs.exists() else 'No posting locations',
                    'action_label': 'Add Locations' if not posting_locs.exists() else None,
                    'action_key': 'compliance-settings',
                },
            ],
        })

        # SECTION 4: Deadlines & Records
        overdue_count = ComplianceDeadline.objects.filter(company=company, status='overdue').count()
        has_deadlines = ComplianceDeadline.objects.filter(company=company).exists()
        sections.append({
            'title': 'Deadlines & Records',
            'icon': 'Calendar',
            'items': [
                {
                    'label': 'No overdue compliance deadlines',
                    'is_passed': overdue_count == 0,
                    'detail': 'All deadlines on track' if overdue_count == 0 else f"{overdue_count} overdue",
                    'action_label': 'View Deadlines' if overdue_count > 0 else None,
                    'action_key': 'compliance-deadlines',
                },
                {
                    'label': 'Compliance calendar populated',
                    'is_passed': has_deadlines,
                    'detail': 'Calendar has deadlines',
                    'action_label': 'Generate Calendar' if not has_deadlines else None,
                    'action_key': 'compliance-deadlines',
                },
            ],
        })

        all_items = [item for section in sections for item in section['items']]
        passed_count = sum(1 for item in all_items if item['is_passed'])
        total_count = len(all_items)
        readiness_pct = int((passed_count / total_count) * 100) if total_count > 0 else 0

        return Response({
            'sections': sections,
            'passed_count': passed_count,
            'total_count': total_count,
            'readiness_pct': readiness_pct,
            'all_passed': readiness_pct == 100,
        })


# =============================================================================

# NOI SUBMISSION VIEWSET
# =============================================================================

class NOISubmissionSerializer(drf_serializers.ModelSerializer):
    product_name = drf_serializers.CharField(source='product.product_name', read_only=True)
    field_name = drf_serializers.CharField(source='field.name', read_only=True)
    farm_name = drf_serializers.CharField(source='field.farm.name', read_only=True)
    is_valid = drf_serializers.BooleanField(read_only=True)
    is_overdue = drf_serializers.BooleanField(read_only=True)

    class Meta:
        from .models import NOISubmission
        model = NOISubmission
        fields = '__all__'
        read_only_fields = ['company', 'created_by']


class NOISubmissionViewSet(viewsets.ModelViewSet):
    """
    API endpoint for NOI (Notice of Intent) submission tracking.

    Tracks when farmers file NOIs with the County Ag Commissioner
    for restricted use pesticide applications.
    """
    serializer_class = NOISubmissionSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def get_queryset(self):
        from .models import NOISubmission
        company = require_company(self.request.user)
        queryset = NOISubmission.objects.filter(
            company=company
        ).select_related('product', 'field', 'field__farm', 'pesticide_application')

        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        field_id = self.request.query_params.get('field')
        if field_id:
            queryset = queryset.filter(field_id=field_id)

        return queryset

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company, created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Mark NOI as submitted to county."""
        noi = self.get_object()
        noi.status = 'submitted'
        noi.filed_date = request.data.get('filed_date', date.today())
        noi.submission_method = request.data.get('submission_method', '')
        noi.save()
        return Response(NOISubmissionSerializer(noi).data)

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Mark NOI as confirmed by county."""
        noi = self.get_object()
        noi.status = 'confirmed'
        noi.confirmation_number = request.data.get('confirmation_number', '')
        noi.county_response_date = request.data.get('response_date', date.today())
        noi.county_response_notes = request.data.get('notes', '')
        noi.valid_from = request.data.get('valid_from', noi.planned_application_date)
        noi.valid_until = request.data.get('valid_until')
        noi.conditions = request.data.get('conditions', '')
        noi.save()
        return Response(NOISubmissionSerializer(noi).data)

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get NOIs that still need to be filed."""
        queryset = self.get_queryset().filter(status='pending')
        serializer = NOISubmissionSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get NOIs past their filing deadline."""
        from .models import NOISubmission
        company = require_company(request.user)
        today = date.today()

        overdue = NOISubmission.objects.filter(
            company=company,
            status='pending',
            planned_application_date__lte=today + timedelta(days=1),
        )
        serializer = NOISubmissionSerializer(overdue, many=True)
        return Response(serializer.data)


# =============================================================================
# WATER GM/STV CALCULATION VIEWSET
# =============================================================================

class WaterGMSTVViewSet(viewsets.ViewSet):
    """
    Calculate FSMA-required Geometric Mean (GM) and Statistical Threshold
    Value (STV) for water sources.

    FSMA Produce Safety Rule requires:
    - GM ≤ 126 CFU/100mL (rolling dataset)
    - STV ≤ 410 CFU/100mL (rolling dataset)
    - Minimum 5 samples for valid calculation
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def list(self, request):
        """Get GM/STV for all water sources."""
        import math
        from .models import WaterSource, WaterTest

        company = require_company(request.user)
        farm_id = request.query_params.get('farm_id')

        sources = WaterSource.objects.filter(farm__company=company)
        if farm_id:
            sources = sources.filter(farm_id=farm_id)

        results = []
        for source in sources.select_related('farm'):
            tests = WaterTest.objects.filter(
                water_source=source,
                ecoli_result__isnull=False,
            ).order_by('-test_date')[:20]

            ecoli_values = [float(t.ecoli_result) for t in tests]
            test_dates = [t.test_date.isoformat() for t in tests]

            result = {
                'source_id': source.id,
                'source_name': source.name,
                'source_type': source.source_type,
                'farm': source.farm.name if source.farm else '',
                'sample_count': len(ecoli_values),
                'test_dates': test_dates,
                'ecoli_values': ecoli_values,
                'geometric_mean': None,
                'gm_threshold': 126.0,
                'gm_compliant': None,
                'stv': None,
                'stv_threshold': 410.0,
                'stv_compliant': None,
                'fsma_compliant': None,
                'sufficient_samples': len(ecoli_values) >= 5,
                'samples_needed': max(0, 5 - len(ecoli_values)),
            }

            if len(ecoli_values) >= 5:
                # Geometric mean: exp(mean(ln(x)))
                log_values = [math.log(max(v, 0.1)) for v in ecoli_values]
                gm = math.exp(sum(log_values) / len(log_values))
                result['geometric_mean'] = round(gm, 2)
                result['gm_compliant'] = gm <= 126.0

                # STV: exp(mean(ln) + 0.6745 * std(ln))
                mean_ln = sum(log_values) / len(log_values)
                variance = sum((lv - mean_ln) ** 2 for lv in log_values) / (len(log_values) - 1)
                std_ln = math.sqrt(variance)
                stv = math.exp(mean_ln + 0.6745 * std_ln)
                result['stv'] = round(stv, 2)
                result['stv_compliant'] = stv <= 410.0

                result['fsma_compliant'] = result['gm_compliant'] and result['stv_compliant']

            results.append(result)

        return Response(results)

    @action(detail=False, methods=['get'], url_path='source/(?P<source_id>[0-9]+)')
    def by_source(self, request, source_id=None):
        """Get detailed GM/STV trend for a specific water source."""
        import math
        from .models import WaterSource, WaterTest

        company = require_company(request.user)

        try:
            source = WaterSource.objects.get(
                id=source_id, farm__company=company
            )
        except WaterSource.DoesNotExist:
            return Response({'error': 'Water source not found'}, status=status.HTTP_404_NOT_FOUND)

        tests = WaterTest.objects.filter(
            water_source=source,
            ecoli_result__isnull=False,
        ).order_by('-test_date')

        # Build rolling GM/STV trend (recalculate for each window)
        all_values = [(t.test_date.isoformat(), float(t.ecoli_result)) for t in tests]
        trend = []

        for i in range(len(all_values)):
            window = all_values[i:i+20]  # Use up to 20 most recent from this point
            values = [v[1] for v in window]

            if len(values) >= 5:
                log_values = [math.log(max(v, 0.1)) for v in values]
                gm = math.exp(sum(log_values) / len(log_values))
                mean_ln = sum(log_values) / len(log_values)
                variance = sum((lv - mean_ln) ** 2 for lv in log_values) / (len(log_values) - 1)
                stv = math.exp(mean_ln + 0.6745 * math.sqrt(variance))

                trend.append({
                    'date': all_values[i][0],
                    'ecoli': all_values[i][1],
                    'gm': round(gm, 2),
                    'stv': round(stv, 2),
                    'gm_compliant': gm <= 126.0,
                    'stv_compliant': stv <= 410.0,
                    'sample_count': len(values),
                })

        return Response({
            'source_id': source.id,
            'source_name': source.name,
            'total_samples': len(all_values),
            'trend': trend,
        })


# =============================================================================
# SGMA REPORT EXPORT VIEWSET
# =============================================================================

class SGMAReportExportViewSet(viewsets.ViewSet):
    """
    Export SGMA semi-annual extraction reports.
    Wraps the existing WaterComplianceService.generate_sgma_report_data()
    and provides JSON + Excel export endpoints.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def list(self, request):
        """Get SGMA report data as JSON."""
        company = require_company(request.user)
        farm_id = request.query_params.get('farm_id')
        period = request.query_params.get('period', 'H1')  # H1 or H2

        if not farm_id:
            # Return reports for all farms
            from .models import Farm
            farms = Farm.objects.filter(company=company)
            reports = []
            for farm in farms:
                report = self._generate_report(company.id, farm.id, period)
                reports.append(report)
            return Response(reports)

        report = self._generate_report(company.id, int(farm_id), period)
        return Response(report)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export SGMA report as Excel."""
        company = require_company(request.user)
        farm_id = request.query_params.get('farm_id')
        period = request.query_params.get('period', 'H1')

        if not farm_id:
            return Response(
                {'error': 'farm_id is required for Excel export'},
                status=status.HTTP_400_BAD_REQUEST
            )

        report = self._generate_report(company.id, int(farm_id), period)

        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"SGMA Report {period}"

            # Header
            ws['A1'] = 'SGMA Semi-Annual Extraction Report'
            ws['A2'] = f'Farm: {report["farm_name"]}'
            ws['A3'] = f'Water Year: {report["water_year"]}'
            ws['A4'] = f'Period: {report["report_period"]} ({report["period_start"]} to {report["period_end"]})'
            ws['A5'] = f'Compliance Status: {report["compliance_status"]}'

            # Well data table
            row = 7
            headers = ['Well Name', 'State Well #', 'GSA', 'Flowmeter', 'Extraction (AF)', 'Period Alloc (AF)', 'Annual Alloc (AF)']
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)

            for well in report.get('wells', []):
                row += 1
                ws.cell(row=row, column=1, value=well.get('well_name', ''))
                ws.cell(row=row, column=2, value=well.get('state_well_number', ''))
                ws.cell(row=row, column=3, value=well.get('gsa', ''))
                ws.cell(row=row, column=4, value='Yes' if well.get('has_flowmeter') else 'No')
                ws.cell(row=row, column=5, value=well.get('extraction_af', 0))
                ws.cell(row=row, column=6, value=well.get('period_allocation_af', 0))
                ws.cell(row=row, column=7, value=well.get('annual_allocation_af', 0))

            # Totals
            row += 1
            ws.cell(row=row, column=1, value='TOTAL')
            ws.cell(row=row, column=5, value=report.get('total_extraction_af', 0))
            ws.cell(row=row, column=6, value=report.get('total_allocation_af', 0))

            # Notes
            row += 2
            ws.cell(row=row, column=1, value='Notes:')
            for note in report.get('notes', []):
                row += 1
                ws.cell(row=row, column=1, value=note)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="sgma_report_{report["farm_name"]}_{period}_{report["water_year"]}.xlsx"'
            )
            return response

        except ImportError:
            return Response(
                {'error': 'openpyxl not installed for Excel export'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _generate_report(self, company_id, farm_id, period):
        """Generate SGMA report data using the existing service."""
        from .services.compliance.water_compliance import WaterComplianceService
        from dataclasses import asdict

        service = WaterComplianceService(company_id=company_id)
        report = service.generate_sgma_report_data(farm_id, period)
        return asdict(report)

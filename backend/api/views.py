import csv
import io
import requests
import re
from datetime import datetime, timedelta, date
from decimal import Decimal
from django.db.models import Sum, Avg, Count, F, Q, Min, Max
from django.db.models.functions import Coalesce
from rest_framework import viewsets, filters, status, serializers
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .permissions import HasCompanyAccess
from .pur_reporting import PURReportGenerator
from .product_import_tool import PesticideProductImporter
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import (
    Farm, Field, FarmParcel, PesticideProduct, PesticideApplication, WaterSource, WaterTest,
    Buyer, LaborContractor, Harvest, HarvestLoad, HarvestLabor,
    WellReading, MeterCalibration, WaterAllocation,
    ExtractionReport, IrrigationEvent,
    FertilizerProduct, NutrientApplication, NutrientPlan,
    QuarantineStatus,
    IrrigationZone, CropCoefficientProfile, CIMISDataCache,
    IrrigationRecommendation, SoilMoistureReading,
    Crop, Rootstock, CropCategory,
    CROP_VARIETY_CHOICES, GRADE_CHOICES
)
from .serializers import (
    FarmSerializer, FarmParcelSerializer, FarmParcelListSerializer, FieldSerializer, PesticideProductSerializer, PesticideApplicationSerializer,
    WaterSourceSerializer, WaterSourceListSerializer, WaterTestSerializer,
    BuyerSerializer, BuyerListSerializer,
    LaborContractorSerializer, LaborContractorListSerializer,
    HarvestSerializer, HarvestListSerializer,
    HarvestLoadSerializer, HarvestLaborSerializer,
    PHICheckSerializer, HarvestStatisticsSerializer,
    WellReadingSerializer, WellReadingCreateSerializer, WellReadingListSerializer,
    MeterCalibrationSerializer, MeterCalibrationCreateSerializer,
    WaterAllocationSerializer, WaterAllocationSummarySerializer,
    ExtractionReportSerializer, ExtractionReportCreateSerializer, ExtractionReportListSerializer,
    IrrigationEventSerializer, IrrigationEventCreateSerializer, IrrigationEventListSerializer,
    SGMADashboardSerializer,
    FertilizerProductSerializer, FertilizerProductListSerializer,
    NutrientApplicationSerializer, NutrientApplicationListSerializer,
    NutrientPlanSerializer, NutrientPlanListSerializer,
    QuarantineStatusSerializer,
    IrrigationZoneSerializer, IrrigationZoneListSerializer, IrrigationZoneDetailSerializer,
    IrrigationZoneEventSerializer, IrrigationZoneEventCreateSerializer,
    IrrigationRecommendationSerializer, IrrigationRecommendationListSerializer,
    CropCoefficientProfileSerializer, CIMISDataSerializer, SoilMoistureReadingSerializer,
    IrrigationDashboardSerializer,
    CropSerializer, CropListSerializer, RootstockSerializer, RootstockListSerializer,
)

# Audit logging mixin for automatic activity tracking
from .audit_utils import AuditLogMixin


# =============================================================================
# HELPER: Company validation for RLS
# =============================================================================

def get_user_company(user):
    """Helper to safely get user's current company."""
    if hasattr(user, 'current_company') and user.current_company:
        return user.current_company
    return None


def require_company(user):
    """Raise validation error if user has no company."""
    company = get_user_company(user)
    if not company:
        raise serializers.ValidationError(
            "You must be associated with a company to perform this action."
        )
    return company


class FarmViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing farms.
    
    RLS NOTES:
    - get_queryset filters by company (defense in depth with RLS)
    - perform_create sets company_id (REQUIRED for RLS INSERT)
    """
    serializer_class = FarmSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'farm_number', 'owner_name', 'county']
    ordering_fields = ['name', 'created_at']
    
    def get_queryset(self):
        """Filter farms by current user's company."""
        queryset = Farm.objects.filter(active=True)
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(company=company)
        return queryset
    
    def perform_create(self, serializer):
        """Set company when creating a new farm - REQUIRED FOR RLS."""
        company = require_company(self.request.user)
        serializer.save(company=company)
    
    @action(detail=True, methods=['get'])
    def fields(self, request, pk=None):
        """Get all fields for a specific farm"""
        farm = self.get_object()
        fields = farm.fields.all()
        serializer = FieldSerializer(fields, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get', 'post'])
    def parcels(self, request, pk=None):
        """GET: List parcels, POST: Add parcel"""
        farm = self.get_object()
        
        if request.method == 'GET':
            serializer = FarmParcelSerializer(farm.parcels.all(), many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            data = request.data.copy()
            data['farm'] = farm.id
            if 'apn' in data:
                data['apn'] = FarmParcel.format_apn(data['apn'], farm.county)
            
            serializer = FarmParcelSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='bulk-parcels')
    def bulk_parcels(self, request, pk=None):
        """Bulk add/update parcels"""
        farm = self.get_object()
        parcels_data = request.data.get('parcels', [])
        replace_existing = request.data.get('replace', False)
        
        if replace_existing:
            farm.parcels.all().delete()
        
        created, updated, errors = [], [], []
        
        for parcel_data in parcels_data:
            apn = FarmParcel.format_apn(parcel_data.get('apn', ''), farm.county)
            try:
                parcel, was_created = FarmParcel.objects.update_or_create(
                    farm=farm, apn=apn,
                    defaults={
                        'acreage': parcel_data.get('acreage'),
                        'ownership_type': parcel_data.get('ownership_type', 'owned'),
                        'notes': parcel_data.get('notes', ''),
                    }
                )
                (created if was_created else updated).append(
                    FarmParcelSerializer(parcel).data
                )
            except Exception as e:
                errors.append({'apn': parcel_data.get('apn'), 'error': str(e)})
        
        return Response({
            'created': created, 'updated': updated, 'errors': errors,
            'total_parcels': farm.parcels.count()
        })

    @action(detail=True, methods=['post'], url_path='update-coordinates')
    def update_coordinates(self, request, pk=None):
        """Update only GPS coordinates for a farm - bypasses full serializer validation"""
        farm = self.get_object()

        lat = request.data.get('gps_latitude')
        lng = request.data.get('gps_longitude')

        if lat is None or lng is None:
            return Response(
                {'error': 'Both gps_latitude and gps_longitude are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            farm.gps_latitude = float(lat)
            farm.gps_longitude = float(lng)
            farm.save(update_fields=['gps_latitude', 'gps_longitude'])

            return Response({
                'success': True,
                'gps_latitude': farm.gps_latitude,
                'gps_longitude': farm.gps_longitude
            })
        except (ValueError, TypeError) as e:
            return Response(
                {'error': f'Invalid coordinate values: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

class FieldViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing fields.
    
    RLS NOTES:
    - Fields inherit company from their Farm
    - get_queryset filters by company through farm relationship
    """
    serializer_class = FieldSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'field_number', 'current_crop', 'county']
    ordering_fields = ['name', 'total_acres', 'created_at']
    
    def get_queryset(self):
        """Filter fields by current user's company through farm."""
        queryset = Field.objects.filter(active=True).select_related('farm')
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(farm__company=company)
        return queryset
    
    @action(detail=True, methods=['get'])
    def applications(self, request, pk=None):
        """Get all applications for a specific field"""
        field = self.get_object()
        applications = field.applications.all()
        serializer = PesticideApplicationSerializer(applications, many=True)
        return Response(serializer.data)


# =============================================================================
# CROP & ROOTSTOCK VIEWSETS
# =============================================================================

class CropViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing crops.

    System defaults (company=null) are read-only.
    Companies can create custom crops.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'variety', 'scientific_name']
    ordering_fields = ['name', 'category', 'created_at']
    ordering = ['category', 'name']

    def get_queryset(self):
        """Return system defaults + company-specific crops."""
        user = self.request.user
        company = get_user_company(user)
        queryset = Crop.objects.filter(active=True)

        if company:
            # System defaults (company=null) + company's custom crops
            queryset = queryset.filter(Q(company__isnull=True) | Q(company=company))
        else:
            # Only system defaults
            queryset = queryset.filter(company__isnull=True)

        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)

        # Filter by crop type
        crop_type = self.request.query_params.get('crop_type')
        if crop_type:
            queryset = queryset.filter(crop_type=crop_type)

        return queryset

    def get_serializer_class(self):
        if self.action == 'list':
            return CropListSerializer
        return CropSerializer

    def perform_create(self, serializer):
        """Assign company to new crops."""
        company = require_company(self.request.user)
        serializer.save(company=company)

    def perform_update(self, serializer):
        """Prevent editing system defaults."""
        if serializer.instance.company is None:
            raise serializers.ValidationError(
                'System default crops cannot be modified. Create a custom crop instead.'
            )
        serializer.save()

    def perform_destroy(self, instance):
        """Prevent deleting system defaults; soft-delete custom crops."""
        if instance.company is None:
            raise serializers.ValidationError(
                'System default crops cannot be deleted.'
            )
        # Check if crop is in use
        if instance.fields.exists():
            raise serializers.ValidationError(
                'Cannot delete crop that is assigned to fields.'
            )
        instance.active = False
        instance.save()

    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Return list of crop categories."""
        return Response([
            {'value': choice.value, 'label': choice.label}
            for choice in CropCategory
        ])

    @action(detail=False, methods=['get'])
    def search(self, request):
        """Quick search for autocomplete."""
        q = request.query_params.get('q', '')
        if len(q) < 2:
            return Response([])
        queryset = self.get_queryset().filter(
            Q(name__icontains=q) | Q(variety__icontains=q)
        )[:20]
        return Response(CropListSerializer(queryset, many=True).data)


class RootstockViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing rootstocks.

    System defaults (company=null) are read-only.
    Companies can create custom rootstocks.
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'primary_category']
    ordering = ['primary_category', 'name']

    def get_queryset(self):
        user = self.request.user
        company = get_user_company(user)
        queryset = Rootstock.objects.filter(active=True)

        if company:
            queryset = queryset.filter(Q(company__isnull=True) | Q(company=company))
        else:
            queryset = queryset.filter(company__isnull=True)

        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(primary_category=category)

        # Filter by compatible crop
        crop_id = self.request.query_params.get('crop')
        if crop_id:
            queryset = queryset.filter(compatible_crops__id=crop_id)

        return queryset

    def get_serializer_class(self):
        if self.action == 'list':
            return RootstockListSerializer
        return RootstockSerializer

    def perform_create(self, serializer):
        company = require_company(self.request.user)
        serializer.save(company=company)

    def perform_update(self, serializer):
        if serializer.instance.company is None:
            raise serializers.ValidationError(
                'System default rootstocks cannot be modified.'
            )
        serializer.save()

    def perform_destroy(self, instance):
        if instance.company is None:
            raise serializers.ValidationError(
                'System default rootstocks cannot be deleted.'
            )
        if instance.fields.exists():
            raise serializers.ValidationError(
                'Cannot delete rootstock that is assigned to fields.'
            )
        instance.active = False
        instance.save()

    @action(detail=False, methods=['get'])
    def for_crop(self, request):
        """Get rootstocks compatible with a specific crop."""
        crop_id = request.query_params.get('crop_id')
        if not crop_id:
            return Response({'error': 'crop_id required'}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset().filter(compatible_crops__id=crop_id)
        return Response(RootstockListSerializer(queryset, many=True).data)


class FarmParcelViewSet(AuditLogMixin, viewsets.ModelViewSet):
    serializer_class = FarmParcelSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_queryset(self):
        user = self.request.user
        if hasattr(user, 'current_company') and user.current_company:
            queryset = FarmParcel.objects.filter(
                farm__company=user.current_company
            ).select_related('farm')
        else:
            queryset = FarmParcel.objects.all().select_related('farm')
        
        farm_id = self.request.query_params.get('farm')
        if farm_id:
            queryset = queryset.filter(farm_id=farm_id)
        
        return queryset


class PesticideProductViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing pesticide products
    """
    queryset = PesticideProduct.objects.all()
    serializer_class = PesticideProductSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['product_name', 'epa_registration_number', 'manufacturer', 'active_ingredients']
    ordering_fields = ['product_name', 'created_at']

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import products from CSV file"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        csv_file = request.FILES['file']
        update_existing = request.data.get('update_existing', 'true').lower() == 'true'
        
        importer = PesticideProductImporter()
        result = importer.import_from_csv(csv_file, update_existing)
        
        return Response({
            'success': result['errors'] == 0,
            'message': f"Imported {result['created']} new products, updated {result['updated']} existing products",
            'statistics': result
        })

    
    @action(detail=False, methods=['get'])
    def export_csv_template(self, request):
        """Download CSV template for importing products"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header row with all fields
        headers = [
            'epa_registration_number',
            'product_name',
            'manufacturer',
            'active_ingredients',
            'formulation_type',
            'restricted_use',
            'product_type',
            'is_fumigant',
            'signal_word',
            'rei_hours',
            'rei_days',
            'phi_days',
            'max_applications_per_season',
            'max_rate_per_application',
            'max_rate_unit',
            'california_registration_number',
            'active_status_california',
            'formulation_code',
            'density_specific_gravity',
            'approved_crops',
            'groundwater_advisory',
            'endangered_species_restrictions',
            'buffer_zone_required',
            'buffer_zone_feet',
            'product_status',
            'unit_size',
            'cost_per_unit',
            'label_url',
            'sds_url',
            'notes',
            'active',
        ]
        
        writer.writerow(headers)
        
        # Example row
        example = [
            '12345-678',
            'Example Insecticide 2.5EC',
            'Example Chemical Company',
            'Permethrin 25%',
            'Emulsifiable Concentrate',
            'false',
            'insecticide',
            'false',
            'CAUTION',
            '12',
            '',
            '7',
            '4',
            '1.0',
            'lbs/acre',
            '',
            'true',
            'EC',
            '1.05',
            'Citrus, Almonds, Walnuts',
            'false',
            'false',
            'false',
            '',
            'active',
            '2.5 gallon',
            '125.50',
            'https://example.com/label.pdf',
            'https://example.com/sds.pdf',
            'Apply early morning or late evening',
            'true',
        ]
        
        writer.writerow(example)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pesticide_products_template.csv"'
        return response

    @action(detail=False, methods=['get'])
    def export_current_products(self, request):
        """Export current products to CSV"""
        products = PesticideProduct.objects.all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'epa_registration_number', 'product_name', 'manufacturer',
            'active_ingredients', 'formulation_type', 'restricted_use',
            'product_type', 'is_fumigant', 'signal_word',
            'rei_hours', 'rei_days', 'phi_days',
            'max_applications_per_season', 'max_rate_per_application', 'max_rate_unit',
            'california_registration_number', 'active_status_california',
            'formulation_code', 'approved_crops', 'product_status',
            'unit_size', 'cost_per_unit', 'label_url', 'notes', 'active'
        ]
        writer.writerow(headers)
        
        # Data rows
        for product in products:
            row = [
                product.epa_registration_number,
                product.product_name,
                product.manufacturer,
                product.active_ingredients,
                product.formulation_type,
                product.restricted_use,
                product.product_type if hasattr(product, 'product_type') else '',
                product.is_fumigant if hasattr(product, 'is_fumigant') else False,
                product.signal_word if hasattr(product, 'signal_word') else '',
                product.rei_hours if hasattr(product, 'rei_hours') else '',
                product.rei_days if hasattr(product, 'rei_days') else '',
                product.phi_days if hasattr(product, 'phi_days') else '',
                product.max_applications_per_season if hasattr(product, 'max_applications_per_season') else '',
                product.max_rate_per_application if hasattr(product, 'max_rate_per_application') else '',
                product.max_rate_unit if hasattr(product, 'max_rate_unit') else '',
                product.california_registration_number if hasattr(product, 'california_registration_number') else '',
                product.active_status_california if hasattr(product, 'active_status_california') else True,
                product.formulation_code if hasattr(product, 'formulation_code') else '',
                product.approved_crops if hasattr(product, 'approved_crops') else '',
                product.product_status if hasattr(product, 'product_status') else 'active',
                product.unit_size if hasattr(product, 'unit_size') else '',
                product.cost_per_unit if hasattr(product, 'cost_per_unit') else '',
                product.label_url if hasattr(product, 'label_url') else '',
                product.notes if hasattr(product, 'notes') else '',
                product.active if hasattr(product, 'active') else True,
            ]
            writer.writerow(row)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="pesticide_products_export.csv"'
        return response


class PesticideApplicationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing pesticide applications.
    
    RLS NOTES:
    - Applications inherit company through field->farm relationship
    - get_queryset filters by company
    """
    serializer_class = PesticideApplicationSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['field__name', 'product__product_name', 'applicator_name']
    ordering_fields = ['application_date', 'created_at']
    
    def get_queryset(self):
        """Filter applications by current user's company through field->farm."""
        queryset = PesticideApplication.objects.select_related(
            'field', 'field__farm', 'product'
        )
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(field__farm__company=company)
        return queryset
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get all pending applications"""
        pending = self.get_queryset().filter(status='pending_signature')
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def ready_for_pur(self, request):
        """Get all complete applications ready for PUR submission"""
        ready = self.get_queryset().filter(status='complete', submitted_to_pur=False)
        serializer = self.get_serializer(ready, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def mark_complete(self, request, pk=None):
        """Mark an application as complete"""
        application = self.get_object()
        application.status = 'complete'
        application.save()
        serializer = self.get_serializer(application)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def mark_submitted(self, request, pk=None):
        """Mark an application as submitted to PUR"""
        application = self.get_object()
        application.submitted_to_pur = True
        application.status = 'submitted'
        from django.utils import timezone
        application.pur_submission_date = timezone.now().date()
        application.save()
        serializer = self.get_serializer(application)
        return Response(serializer.data)
    
        
    @action(detail=False, methods=['post'])
    def validate_pur(self, request):
        """Validate applications for PUR compliance."""
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        farm_id = request.data.get('farm_id')
        
        # Filter applications
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        # Generate validation report
        generator = PURReportGenerator(queryset)
        validation_result = generator.validate_for_pur()
        
        return Response(validation_result)
    
    @action(detail=False, methods=['post'])
    def export_pur_csv(self, request):
        """Export applications as PUR-formatted CSV."""
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        farm_id = request.data.get('farm_id')
        
        # Filter applications
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        # Validate first
        generator = PURReportGenerator(queryset)
        validation = generator.validate_for_pur()
        
        if not validation['valid']:
            return Response({
                'error': 'Applications contain validation errors',
                'validation': validation
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate CSV
        csv_content = generator.generate_csv()
        
        # Create response
        response = HttpResponse(csv_content, content_type='text/csv')
        filename = f"PUR_Report_{start_date}_to_{end_date}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @action(detail=False, methods=['post'])
    def pur_summary(self, request):
        """Get summary statistics for PUR report period."""
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        farm_id = request.data.get('farm_id')
        
        # Filter applications
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        # Generate summary
        generator = PURReportGenerator(queryset)
        summary = generator.generate_summary_report()
        validation = generator.validate_for_pur()
        
        return Response({
            'summary': summary,
            'validation': validation
        })
    
    @action(detail=False, methods=['get'])
    def export_pur(self, request):
        """
        Enhanced PUR export with multiple format options.
        
        Query Parameters:
        - format: 'csv' (official CA PUR format), 'excel', or 'csv_detailed'
        - start_date: YYYY-MM-DD
        - end_date: YYYY-MM-DD
        - farm_id: Filter by farm
        - status: Filter by status
        - county: Filter by county
        - validate: 'true' to validate before export (default: true)
        """
        # Get parameters
        export_format = request.query_params.get('format', 'csv')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        farm_id = request.query_params.get('farm_id')
        status_filter = request.query_params.get('status')
        county = request.query_params.get('county')
        validate_first = request.query_params.get('validate', 'true').lower() == 'true'
        
        # Build query
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if county:
            queryset = queryset.filter(field__county__icontains=county)
        
        # Select related to optimize
        queryset = queryset.select_related('field', 'field__farm', 'product').order_by('application_date')
        
        # Use existing PURReportGenerator for validation and official CSV
        generator = PURReportGenerator(queryset)
        
        # Validate if requested
        if validate_first:
            validation = generator.validate_for_pur()
            if not validation['valid'] and export_format == 'csv':
                # For official PUR format, enforce validation
                return Response({
                    'error': 'Applications contain validation errors. Cannot export official PUR format.',
                    'validation': validation
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Route to appropriate export method
        if export_format == 'csv':
            # Use existing official CA PUR format
            return self._export_official_pur_csv(generator, start_date, end_date)
        elif export_format == 'excel':
            # New Excel format with summary
            return self._export_pur_excel(queryset, start_date, end_date)
        elif export_format == 'csv_detailed':
            # Detailed CSV with all fields
            return self._export_pur_csv_detailed(queryset, start_date, end_date)
        else:
            return Response({'error': 'Invalid format'}, status=status.HTTP_400_BAD_REQUEST)


    def _export_official_pur_csv(self, generator, start_date, end_date):
        """Export using existing official California PUR format"""
        csv_content = generator.generate_csv()
        
        response = HttpResponse(csv_content, content_type='text/csv')
        filename = f"PUR_Official_{start_date or 'all'}_to_{end_date or 'all'}_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


    def _export_pur_csv_detailed(self, queryset, start_date, end_date):
        """Export detailed CSV with all available fields"""
        import csv
        
        response = HttpResponse(content_type='text/csv')
        
        date_range = ""
        if start_date and end_date:
            date_range = f"_{start_date}_to_{end_date}"
        
        filename = f"PUR_Detailed{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Detailed header
        writer.writerow([
            'Application Date',
            'Farm Name',
            'Farm Number',
            'Operator Name',
            'County',
            'Field Name',
            'Field Number',
            'Section',
            'Township',
            'Range',
            'GPS Latitude',
            'GPS Longitude',
            'Acres Treated',
            'Current Crop',
            'EPA Registration Number',
            'Product Name',
            'Active Ingredients',
            'Amount Used',
            'Unit',
            'Application Method',
            'Target Pest',
            'Applicator Name',
            'Start Time',
            'End Time',
            'Temperature (°F)',
            'Wind Speed (mph)',
            'Wind Direction',
            'Restricted Use',
            'Fumigant',
            'REI (hours)',
            'PHI (days)',
            'Signal Word',
            'Status',
            'PUR Submitted',
            'Submission Date',
            'Notes'
        ])
        
        # Write data
        for app in queryset:
            writer.writerow([
                app.application_date.strftime('%m/%d/%Y'),
                app.field.farm.name,
                app.field.farm.farm_number or '',
                app.field.farm.operator_name or '',
                app.field.county,
                app.field.name,
                app.field.field_number or '',
                app.field.section or '',
                app.field.township or '',
                app.field.range_value or '',
                app.field.gps_lat or '',
                app.field.gps_long or '',
                app.acres_treated,
                app.field.current_crop or '',
                app.product.epa_registration_number,
                app.product.product_name,
                app.product.active_ingredients,
                app.amount_used,
                app.unit_of_measure,
                app.application_method,
                app.target_pest or '',
                app.applicator_name,
                app.start_time.strftime('%H:%M') if app.start_time else '',
                app.end_time.strftime('%H:%M') if app.end_time else '',
                app.temperature or '',
                app.wind_speed or '',
                app.wind_direction or '',
                'Yes' if app.product.restricted_use else 'No',
                'Yes' if app.product.is_fumigant else 'No',
                app.product.rei_hours or '',
                app.product.phi_days or '',
                app.product.signal_word or '',
                app.get_status_display(),
                'Yes' if app.submitted_to_pur else 'No',
                app.pur_submission_date.strftime('%m/%d/%Y') if app.pur_submission_date else '',
                app.notes or ''
            ])
        
        return response


    def _export_pur_excel(self, queryset, start_date, end_date):
        """Export detailed Excel with formatting and summary"""
        wb = Workbook()
        
        # === OFFICIAL PUR FORMAT SHEET ===
        ws_pur = wb.active
        ws_pur.title = "Official PUR Format"
        
        # Use PURReportGenerator for official format
        generator = PURReportGenerator(queryset)
        
        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Write official PUR headers
        for col_num, field_name in enumerate(generator.PUR_FIELDS, 1):
            cell = ws_pur.cell(row=1, column=col_num)
            cell.value = field_name.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write official PUR data
        for row_num, app in enumerate(queryset, 2):
            row_data = generator._application_to_pur_row(app)
            for col_num, field_name in enumerate(generator.PUR_FIELDS, 1):
                cell = ws_pur.cell(row=row_num, column=col_num)
                cell.value = row_data.get(field_name, '')
                cell.border = border
        
        # Auto-size columns for PUR sheet
        for col in range(1, len(generator.PUR_FIELDS) + 1):
            ws_pur.column_dimensions[chr(64 + col)].width = 15
        
        ws_pur.freeze_panes = 'A2'
        
        # === DETAILED DATA SHEET ===
        ws_detail = wb.create_sheet("Detailed Data")
        
        detail_headers = [
            'Date', 'Farm', 'Farm #', 'County', 'Field', 'Field #',
            'Acres', 'Crop', 'EPA #', 'Product', 'Active Ingredients',
            'Amount', 'Unit', 'Method', 'Target Pest', 'Applicator',
            'Start', 'End', 'Temp', 'Wind Speed', 'Wind Dir',
            'Restricted', 'REI', 'PHI', 'Status', 'Notes'
        ]
        
        # Write headers
        for col_num, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_num, app in enumerate(queryset, 2):
            data = [
                app.application_date.strftime('%m/%d/%Y'),
                app.field.farm.name,
                app.field.farm.farm_number or '',
                app.field.county,
                app.field.name,
                app.field.field_number or '',
                float(app.acres_treated),
                app.field.current_crop or '',
                app.product.epa_registration_number,
                app.product.product_name,
                app.product.active_ingredients,
                float(app.amount_used),
                app.unit_of_measure,
                app.application_method,
                app.target_pest or '',
                app.applicator_name,
                app.start_time.strftime('%H:%M') if app.start_time else '',
                app.end_time.strftime('%H:%M') if app.end_time else '',
                app.temperature or '',
                app.wind_speed or '',
                app.wind_direction or '',
                'Yes' if app.product.restricted_use else 'No',
                app.product.rei_hours or '',
                app.product.phi_days or '',
                app.get_status_display(),
                app.notes or ''
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws_detail.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
        
        # Auto-size columns
        for col_num in range(1, len(detail_headers) + 1):
            col_letter = chr(64 + col_num)
            ws_detail.column_dimensions[col_letter].width = 12
        
        ws_detail.freeze_panes = 'A2'
        
        # === SUMMARY SHEET ===
        ws_summary = wb.create_sheet("Summary & Validation")
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # Title
        ws_summary['A1'] = "PUR Report Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Report Info
        ws_summary[f'A{row}'] = "Report Generated:"
        ws_summary[f'B{row}'] = datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        row += 1
        
        if start_date:
            ws_summary[f'A{row}'] = "Start Date:"
            ws_summary[f'B{row}'] = start_date
            row += 1
        
        if end_date:
            ws_summary[f'A{row}'] = "End Date:"
            ws_summary[f'B{row}'] = end_date
            row += 1
        
        row += 1
        
        # Statistics
        ws_summary[f'A{row}'] = "STATISTICS"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws_summary[f'A{row}'] = "Total Applications:"
        ws_summary[f'B{row}'] = queryset.count()
        row += 1
        
        ws_summary[f'A{row}'] = "Total Acres Treated:"
        ws_summary[f'B{row}'] = queryset.aggregate(Sum('acres_treated'))['acres_treated__sum'] or 0
        row += 1
        
        ws_summary[f'A{row}'] = "Unique Farms:"
        ws_summary[f'B{row}'] = queryset.values('field__farm').distinct().count()
        row += 1
        
        ws_summary[f'A{row}'] = "Unique Fields:"
        ws_summary[f'B{row}'] = queryset.values('field').distinct().count()
        row += 1
        
        ws_summary[f'A{row}'] = "Unique Products:"
        ws_summary[f'B{row}'] = queryset.values('product').distinct().count()
        row += 2
        
        # Validation Results
        validation = generator.validate_for_pur()
        
        ws_summary[f'A{row}'] = "VALIDATION RESULTS"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws_summary[f'A{row}'] = "Ready for PUR Submission:"
        ws_summary[f'B{row}'] = "YES" if validation['valid'] else "NO"
        ws_summary[f'B{row}'].font = Font(
            color="00FF00" if validation['valid'] else "FF0000",
            bold=True
        )
        row += 2
        
        if validation['errors']:
            ws_summary[f'A{row}'] = "ERRORS (Must Fix):"
            ws_summary[f'A{row}'].font = Font(bold=True, color="FF0000")
            row += 1
            for error in validation['errors']:
                ws_summary[f'A{row}'] = error
                ws_summary[f'A{row}'].font = Font(color="FF0000")
                row += 1
            row += 1
        
        if validation['warnings']:
            ws_summary[f'A{row}'] = "WARNINGS (Recommended):"
            ws_summary[f'A{row}'].font = Font(bold=True, color="FFA500")
            row += 1
            for warning in validation['warnings']:
                ws_summary[f'A{row}'] = warning
                ws_summary[f'A{row}'].font = Font(color="FFA500")
                row += 1
        
        # Status breakdown
        row += 2
        ws_summary[f'A{row}'] = "STATUS BREAKDOWN"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        status_counts = queryset.values('status').annotate(count=Count('id'))
        for status_data in status_counts:
            ws_summary[f'A{row}'] = status_data['status'].replace('_', ' ').title()
            ws_summary[f'B{row}'] = status_data['count']
            row += 1
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        date_range = ""
        if start_date and end_date:
            date_range = f"_{start_date}_to_{end_date}"
        
        filename = f"PUR_Report{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


    # Also add the validation and summary endpoints from existing file
    @action(detail=False, methods=['post', 'get'])
    def validate_pur(self, request):
        """
        Validate applications for PUR compliance.
        
        Supports both GET and POST methods for flexibility.
        """
        # Get parameters from either query params or body
        if request.method == 'POST':
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            farm_id = request.data.get('farm_id')
        else:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            farm_id = request.query_params.get('farm_id')
        
        # Filter applications
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        queryset = queryset.select_related('field', 'field__farm', 'product')
        
        # Generate validation report
        generator = PURReportGenerator(queryset)
        validation_result = generator.validate_for_pur()
        
        return Response(validation_result)


    @action(detail=False, methods=['post', 'get'])
    def pur_summary(self, request):
        """
        Get summary statistics for PUR report period.
        
        Supports both GET and POST methods.
        """
        # Get parameters
        if request.method == 'POST':
            start_date = request.data.get('start_date')
            end_date = request.data.get('end_date')
            farm_id = request.data.get('farm_id')
        else:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            farm_id = request.query_params.get('farm_id')
        
        # Filter applications
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        queryset = queryset.select_related('field', 'field__farm', 'product')
        
        # Generate summary and validation
        generator = PURReportGenerator(queryset)
        summary = generator.generate_summary_report()
        validation = generator.validate_for_pur()
        
        return Response({
            'summary': summary,
            'validation': validation
        })

    # =========================================================================
    # SERVICE-BASED ENDPOINTS
    # These endpoints use the new services layer for business logic
    # =========================================================================

    @action(detail=False, methods=['post'])
    def validate_proposed(self, request):
        """
        Validate a proposed pesticide application before creating it.

        Uses the PesticideComplianceService for comprehensive validation
        including PHI, REI, rate limits, NOI requirements, and weather.

        POST body:
        {
            "field_id": 1,
            "product_id": 5,
            "application_date": "2024-12-20",
            "rate_per_acre": 2.5,
            "application_method": "Ground Spray",
            "acres_treated": 10.0,
            "applicator_name": "John Smith",  // optional but needed for restricted
            "applicator_license": "12345",    // optional
            "check_weather": true,            // optional, default true
            "check_quarantine": true          // optional, default true
        }

        Returns comprehensive validation result with issues, warnings,
        NOI requirements, and recommended actions.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        # Extract parameters
        field_id = request.data.get('field_id')
        product_id = request.data.get('product_id')
        application_date_str = request.data.get('application_date')
        rate_per_acre = request.data.get('rate_per_acre')
        application_method = request.data.get('application_method')
        acres_treated = request.data.get('acres_treated')
        applicator_name = request.data.get('applicator_name')
        applicator_license = request.data.get('applicator_license')
        check_weather = request.data.get('check_weather', True)
        check_quarantine = request.data.get('check_quarantine', True)

        # Validate required fields
        if not all([field_id, product_id, application_date_str, rate_per_acre,
                   application_method, acres_treated]):
            return Response({
                'error': 'Missing required fields',
                'required': ['field_id', 'product_id', 'application_date',
                           'rate_per_acre', 'application_method', 'acres_treated']
            }, status=status.HTTP_400_BAD_REQUEST)

        # Parse date
        try:
            application_date = datetime.strptime(application_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get company for RLS
        company = get_user_company(request.user)
        company_id = company.id if company else None

        # Use compliance service
        service = PesticideComplianceService(company_id=company_id)
        result = service.validate_proposed_application(
            field_id=int(field_id),
            product_id=int(product_id),
            application_date=application_date,
            rate_per_acre=float(rate_per_acre),
            application_method=application_method,
            acres_treated=float(acres_treated),
            applicator_name=applicator_name,
            applicator_license=applicator_license,
            check_weather=check_weather,
            check_quarantine=check_quarantine,
        )

        return Response(result.to_dict())

    @action(detail=False, methods=['get', 'post'])
    def phi_clearance(self, request):
        """
        Check PHI clearance status for fields.

        GET: Check a single field
            Query params: field_id, proposed_harvest_date (optional)

        POST: Check multiple fields or a farm
            Body: {"farm_id": 1, "proposed_harvest_date": "2024-12-20"}
            Or: {"field_ids": [1, 2, 3], "proposed_harvest_date": "2024-12-20"}

        Returns PHI clearance status including earliest harvest date.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        company = get_user_company(request.user)
        company_id = company.id if company else None
        service = PesticideComplianceService(company_id=company_id)

        if request.method == 'GET':
            field_id = request.query_params.get('field_id')
            proposed_date_str = request.query_params.get('proposed_harvest_date')

            if not field_id:
                return Response({
                    'error': 'field_id is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            proposed_date = None
            if proposed_date_str:
                try:
                    proposed_date = datetime.strptime(proposed_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response({
                        'error': 'Invalid date format. Use YYYY-MM-DD'
                    }, status=status.HTTP_400_BAD_REQUEST)

            result = service.calculate_phi_clearance(
                field_id=int(field_id),
                proposed_harvest_date=proposed_date
            )
            return Response(result.to_dict())

        else:  # POST
            farm_id = request.data.get('farm_id')
            field_ids = request.data.get('field_ids')
            proposed_date_str = request.data.get('proposed_harvest_date')

            proposed_date = None
            if proposed_date_str:
                try:
                    proposed_date = datetime.strptime(proposed_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response({
                        'error': 'Invalid date format. Use YYYY-MM-DD'
                    }, status=status.HTTP_400_BAD_REQUEST)

            if farm_id:
                results = service.calculate_phi_for_all_fields(
                    farm_id=int(farm_id),
                    proposed_harvest_date=proposed_date
                )
            elif field_ids:
                results = [
                    service.calculate_phi_clearance(
                        field_id=int(fid),
                        proposed_harvest_date=proposed_date
                    )
                    for fid in field_ids
                ]
            else:
                return Response({
                    'error': 'Either farm_id or field_ids is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'fields': [r.to_dict() for r in results],
                'summary': {
                    'total_fields': len(results),
                    'clear_for_harvest': len([r for r in results if r.is_clear]),
                    'blocked': len([r for r in results if not r.is_clear]),
                }
            })

    @action(detail=False, methods=['get'])
    def rei_status(self, request):
        """
        Get REI (Restricted Entry Interval) status for a field.

        Query params:
            field_id: Required
            check_datetime: Optional ISO datetime (default: now)

        Returns whether field is safe for worker entry.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        field_id = request.query_params.get('field_id')
        check_datetime_str = request.query_params.get('check_datetime')

        if not field_id:
            return Response({
                'error': 'field_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        check_datetime = None
        if check_datetime_str:
            try:
                check_datetime = datetime.fromisoformat(check_datetime_str)
            except ValueError:
                return Response({
                    'error': 'Invalid datetime format. Use ISO format.'
                }, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = PesticideComplianceService(company_id=company_id)
        result = service.get_rei_status(
            field_id=int(field_id),
            check_datetime=check_datetime
        )

        return Response(result.to_dict())

    @action(detail=False, methods=['get', 'post'])
    def noi_requirements(self, request):
        """
        Get Notice of Intent requirements for a product application.

        GET params or POST body:
            product_id: Required
            application_date: Required (YYYY-MM-DD)
            county: Required

        Returns NOI requirements including deadline and submission info.
        """
        from api.services.compliance import PesticideComplianceService
        from datetime import datetime

        if request.method == 'POST':
            product_id = request.data.get('product_id')
            application_date_str = request.data.get('application_date')
            county = request.data.get('county')
        else:
            product_id = request.query_params.get('product_id')
            application_date_str = request.query_params.get('application_date')
            county = request.query_params.get('county')

        if not all([product_id, application_date_str, county]):
            return Response({
                'error': 'product_id, application_date, and county are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            application_date = datetime.strptime(application_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)

        service = PesticideComplianceService()
        result = service.get_noi_requirements(
            product_id=int(product_id),
            application_date=application_date,
            county=county
        )

        return Response(result)

    @action(detail=False, methods=['get'])
    def spray_windows(self, request):
        """
        Find suitable spray windows based on weather forecast.

        Query params:
            farm_id: Required
            days_ahead: Optional (default: 7, max: 7)
            application_method: Optional ('ground' or 'aerial', default: 'ground')

        Returns list of optimal spray windows.
        """
        from api.services.operations import SprayPlanningService

        farm_id = request.query_params.get('farm_id')
        days_ahead = request.query_params.get('days_ahead', '7')
        application_method = request.query_params.get('application_method', 'ground')

        if not farm_id:
            return Response({
                'error': 'farm_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = SprayPlanningService(company_id=company_id)
        windows = service.find_spray_windows(
            farm_id=int(farm_id),
            days_ahead=min(int(days_ahead), 7),
            application_method=application_method
        )

        return Response({
            'farm_id': int(farm_id),
            'windows': [w.to_dict() for w in windows],
            'summary': {
                'total_windows': len(windows),
                'good_windows': len([w for w in windows if w.rating == 'good']),
                'fair_windows': len([w for w in windows if w.rating == 'fair']),
            }
        })

    @action(detail=False, methods=['get'])
    def spray_conditions(self, request):
        """
        Evaluate current spray conditions for a farm.

        Query params:
            farm_id: Required

        Returns current conditions assessment and recommendations.
        """
        from api.services.operations import SprayPlanningService

        farm_id = request.query_params.get('farm_id')

        if not farm_id:
            return Response({
                'error': 'farm_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = SprayPlanningService(company_id=company_id)
        result = service.evaluate_spray_conditions(farm_id=int(farm_id))

        return Response(result.to_dict())

    @action(detail=False, methods=['post'])
    def recommend_timing(self, request):
        """
        Get recommended application timing for a product on a field.

        POST body:
        {
            "field_id": 1,
            "product_id": 5,
            "urgency": "normal"  // 'urgent', 'normal', or 'flexible'
        }

        Returns timing recommendation considering weather, PHI, and REI.
        """
        from api.services.operations import SprayPlanningService

        field_id = request.data.get('field_id')
        product_id = request.data.get('product_id')
        urgency = request.data.get('urgency', 'normal')

        if not all([field_id, product_id]):
            return Response({
                'error': 'field_id and product_id are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        if urgency not in ('urgent', 'normal', 'flexible'):
            urgency = 'normal'

        company = get_user_company(request.user)
        company_id = company.id if company else None

        service = SprayPlanningService(company_id=company_id)
        result = service.recommend_application_timing(
            field_id=int(field_id),
            product_id=int(product_id),
            urgency=urgency
        )

        return Response(result.to_dict())


class WaterSourceViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing water sources.
    
    RLS NOTES:
    - Water sources belong to farms, which have company FK
    - get_queryset filters by company through farm
    """
    serializer_class = WaterSourceSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'source_type', 'farm__name']
    ordering_fields = ['name', 'created_at']
    
    def get_queryset(self):
        """Filter water sources by current user's company through farm."""
        queryset = WaterSource.objects.filter(active=True).select_related('farm')
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(farm__company=company)
        return queryset
    
    @action(detail=True, methods=['get'])
    def tests(self, request, pk=None):
        """Get all tests for a specific water source"""
        water_source = self.get_object()
        tests = water_source.water_tests.all()
        serializer = WaterTestSerializer(tests, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get all water sources with overdue tests"""
        overdue_sources = [ws for ws in self.get_queryset() if ws.is_test_overdue()]
        serializer = self.get_serializer(overdue_sources, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def gsa_fee_defaults(self, request):
        """Get default fee rates for each GSA."""
        from .models import GSA_FEE_DEFAULTS, GSA_CHOICES
        # Convert Decimal to float for JSON serialization
        result = {}
        for gsa_code, defaults in GSA_FEE_DEFAULTS.items():
            result[gsa_code] = {
                k: float(v) if v is not None else None
                for k, v in defaults.items()
            }
        # Add GSA display names
        gsa_names = dict(GSA_CHOICES)
        for gsa_code in result:
            result[gsa_code]['display_name'] = gsa_names.get(gsa_code, gsa_code)
        return Response(result)


class WaterTestViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing water tests.
    
    RLS NOTES:
    - Water tests inherit company through water_source->farm relationship
    """
    serializer_class = WaterTestSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['water_source__name', 'lab_name', 'status']
    ordering_fields = ['test_date', 'created_at']

    def get_queryset(self):
        """Filter water tests by company through water_source->farm."""
        queryset = WaterTest.objects.select_related('water_source', 'water_source__farm')
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(water_source__farm__company=company)
        
        water_source_id = self.request.query_params.get('water_source', None)
        if water_source_id:
            queryset = queryset.filter(water_source_id=water_source_id)
        return queryset

    def perform_create(self, serializer):
        """Auto-determine status when creating"""
        instance = serializer.save()
        if instance.status == 'pending' and instance.ecoli_result is not None:
            instance.status = instance.auto_determine_status()
            instance.save()

    @action(detail=False, methods=['get'])
    def failed(self, request):
        """Get all failed tests"""
        failed = self.get_queryset().filter(status='fail')
        serializer = self.get_serializer(failed, many=True)
        return Response(serializer.data)
    
from rest_framework.decorators import api_view

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def report_statistics(request):
    """Get statistics for the reports dashboard"""
    from django.db.models import Sum, Count
    
    # Get query parameters
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    farm_id = request.query_params.get('farm_id')
    
    # Build base queryset with company filter
    company = get_user_company(request.user)
    queryset = PesticideApplication.objects.all()
    if company:
        queryset = queryset.filter(field__farm__company=company)
    
    if start_date:
        queryset = queryset.filter(application_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(application_date__lte=end_date)
    if farm_id:
        queryset = queryset.filter(field__farm_id=farm_id)
    
    # Calculate statistics
    stats = {
        'total_applications': queryset.count(),
        'total_acres': float(queryset.aggregate(Sum('acres_treated'))['acres_treated__sum'] or 0),
        'unique_farms': queryset.values('field__farm').distinct().count(),
        'unique_fields': queryset.values('field').distinct().count(),
        'unique_products': queryset.values('product').distinct().count(),
        'status_breakdown': {},
        'by_county': {},
        'by_month': {},
        'restricted_use_count': 0,
        'submitted_to_pur': queryset.filter(submitted_to_pur=True).count(),
        'pending_signature': queryset.filter(status='pending_signature').count(),
    }
    
    # Status breakdown
    status_counts = queryset.values('status').annotate(count=Count('id'))
    for item in status_counts:
        stats['status_breakdown'][item['status']] = item['count']
    
    # By county
    county_counts = queryset.values('field__county').annotate(
        count=Count('id'),
        acres=Sum('acres_treated')
    ).order_by('-count')[:10]
    
    for item in county_counts:
        county = item['field__county'] or 'Unknown'
        stats['by_county'][county] = {
            'applications': item['count'],
            'acres': float(item['acres'] or 0)
        }
    
    # By month (last 12 months)
    twelve_months_ago = datetime.now() - timedelta(days=365)
    monthly_data = queryset.filter(
        application_date__gte=twelve_months_ago
    ).extra(
        select={'month': "strftime('%%Y-%%m', application_date)"}
    ).values('month').annotate(
        count=Count('id'),
        acres=Sum('acres_treated')
    ).order_by('month')
    
    for item in monthly_data:
        if item['month']:
            stats['by_month'][item['month']] = {
                'applications': item['count'],
                'acres': float(item['acres'] or 0)
            }
    
    # Restricted use products
    stats['restricted_use_count'] = queryset.filter(
        product__restricted_use=True
    ).count()
    
    return Response(stats)

# -----------------------------------------------------------------------------
# BUYER VIEWSET
# -----------------------------------------------------------------------------

class BuyerViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    CRUD operations for Buyers (packing houses, processors, etc.)

    RLS NOTES:
    - Buyers are scoped by company for multi-tenant isolation
    - Uses RLS policy: buyer_company_isolation
    """
    queryset = Buyer.objects.all()
    serializer_class = BuyerSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def get_queryset(self):
        queryset = Buyer.objects.all()

        # Filter by company (multi-tenancy)
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(company=company)

        # Filter by active status
        active = self.request.query_params.get('active')
        if active is not None:
            queryset = queryset.filter(active=active.lower() == 'true')

        # Filter by buyer type
        buyer_type = self.request.query_params.get('buyer_type')
        if buyer_type:
            queryset = queryset.filter(buyer_type=buyer_type)

        # Search by name
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)

        return queryset.order_by('name')

    def perform_create(self, serializer):
        """Auto-assign company when creating buyer."""
        company = get_user_company(self.request.user)
        serializer.save(company=company)
    
    def get_serializer_class(self):
        if self.action == 'list' and self.request.query_params.get('simple') == 'true':
            return BuyerListSerializer
        return BuyerSerializer
    
    @action(detail=True, methods=['get'])
    def load_history(self, request, pk=None):
        """Get all loads sent to this buyer."""
        buyer = self.get_object()
        loads = HarvestLoad.objects.filter(buyer=buyer).select_related(
            'harvest', 'harvest__field', 'harvest__field__farm'
        ).order_by('-harvest__harvest_date')
        
        serializer = HarvestLoadSerializer(loads, many=True)
        
        # Calculate summary stats
        summary = loads.aggregate(
            total_loads=Count('id'),
            total_bins=Sum('bins'),
            total_revenue=Sum('total_revenue'),
            pending_revenue=Sum('total_revenue', filter=Q(payment_status='pending'))
        )
        
        return Response({
            'buyer': BuyerSerializer(buyer).data,
            'summary': summary,
            'loads': serializer.data
        })

    @action(detail=True, methods=['get'])
    def performance(self, request, pk=None):
        """
        Get performance metrics for this buyer.
        Returns: historical pricing, payment timeliness, quality grades, volume
        """
        from django.db.models import Avg, Count, Sum, Q, F, ExpressionWrapper, DurationField
        from django.db.models.functions import Coalesce
        from datetime import date

        buyer = self.get_object()

        # Get all loads for this buyer
        loads = HarvestLoad.objects.filter(buyer=buyer).select_related(
            'harvest', 'harvest__field'
        )

        # Overall statistics
        overall_stats = loads.aggregate(
            total_loads=Count('id'),
            total_bins=Sum('bins'),
            total_revenue=Sum('total_price'),
            avg_price_per_bin=Avg(F('total_price') / F('bins'), output_field=DecimalField()),
            paid_count=Count('id', filter=Q(payment_status='paid')),
            pending_count=Count('id', filter=Q(payment_status='pending')),
            late_count=Count('id', filter=Q(payment_status='pending', payment_due_date__lt=date.today())),
        )

        # Calculate payment timeliness for paid loads
        paid_loads = loads.filter(payment_status='paid', payment_due_date__isnull=False)

        # This is a simplified calculation - in production you'd track actual payment dates
        # For now, we'll use the updated_at timestamp as a proxy
        avg_days_to_pay = None
        if paid_loads.exists():
            # Calculate average days between harvest and payment
            payment_times = []
            for load in paid_loads:
                days_diff = (load.updated_at.date() - load.harvest.harvest_date).days
                payment_times.append(days_diff)
            avg_days_to_pay = sum(payment_times) / len(payment_times) if payment_times else None

        # Pricing by crop variety
        pricing_by_crop = []
        for crop_choice in HARVEST_CONSTANTS['CROP_VARIETIES']:
            crop_value = crop_choice['value']
            crop_loads = loads.filter(harvest__crop_variety=crop_value)
            crop_stats = crop_loads.aggregate(
                load_count=Count('id'),
                total_bins=Sum('bins'),
                avg_price=Avg(F('total_price') / F('bins'), output_field=DecimalField()),
                min_price=Min(F('total_price') / F('bins'), output_field=DecimalField()),
                max_price=Max(F('total_price') / F('bins'), output_field=DecimalField()),
            )

            if crop_stats['load_count'] > 0:
                pricing_by_crop.append({
                    'crop_variety': crop_value,
                    'crop_variety_display': crop_choice['label'],
                    'load_count': crop_stats['load_count'],
                    'total_bins': crop_stats['total_bins'] or 0,
                    'avg_price_per_bin': float(crop_stats['avg_price'] or 0),
                    'min_price_per_bin': float(crop_stats['min_price'] or 0),
                    'max_price_per_bin': float(crop_stats['max_price'] or 0),
                })

        # Quality grade distribution
        quality_distribution = []
        for grade_value, grade_label in GRADE_CHOICES:
            grade_count = loads.filter(grade=grade_value).count()
            if grade_count > 0:
                quality_distribution.append({
                    'grade': grade_value,
                    'grade_display': grade_label,
                    'count': grade_count,
                    'percentage': (grade_count / overall_stats['total_loads']) * 100 if overall_stats['total_loads'] else 0
                })

        # Recent loads (last 10)
        recent_loads = loads.order_by('-harvest__harvest_date')[:10]
        recent_loads_data = HarvestLoadSerializer(recent_loads, many=True).data

        return Response({
            'buyer': BuyerSerializer(buyer).data,
            'overall_stats': {
                'total_loads': overall_stats['total_loads'],
                'total_bins': overall_stats['total_bins'] or 0,
                'total_revenue': float(overall_stats['total_revenue'] or 0),
                'avg_price_per_bin': float(overall_stats['avg_price_per_bin'] or 0),
                'paid_count': overall_stats['paid_count'],
                'pending_count': overall_stats['pending_count'],
                'late_count': overall_stats['late_count'],
                'payment_rate': (overall_stats['paid_count'] / overall_stats['total_loads'] * 100) if overall_stats['total_loads'] else 0,
                'avg_days_to_pay': avg_days_to_pay,
            },
            'pricing_by_crop': pricing_by_crop,
            'quality_distribution': quality_distribution,
            'recent_loads': recent_loads_data
        })


# -----------------------------------------------------------------------------
# LABOR CONTRACTOR VIEWSET
# -----------------------------------------------------------------------------

class LaborContractorViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    CRUD operations for Labor Contractors.

    RLS NOTES:
    - Labor contractors are scoped by company for multi-tenant isolation
    - Uses RLS policy: laborcontractor_company_isolation
    """
    queryset = LaborContractor.objects.all()
    serializer_class = LaborContractorSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def get_queryset(self):
        queryset = LaborContractor.objects.all()

        # Filter by company (multi-tenancy)
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(company=company)

        # Filter by active status
        active = self.request.query_params.get('active')
        if active is not None:
            queryset = queryset.filter(active=active.lower() == 'true')

        # Filter by valid license
        valid_license = self.request.query_params.get('valid_license')
        if valid_license == 'true':
            queryset = queryset.filter(license_expiration__gte=date.today())

        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(company_name__icontains=search)

        return queryset.order_by('company_name')

    def perform_create(self, serializer):
        """Auto-assign company when creating labor contractor."""
        company = get_user_company(self.request.user)
        serializer.save(company=company)
    
    def get_serializer_class(self):
        if self.action == 'list' and self.request.query_params.get('simple') == 'true':
            return LaborContractorListSerializer
        return LaborContractorSerializer
    
    @action(detail=True, methods=['get'])
    def job_history(self, request, pk=None):
        """Get all harvest jobs for this contractor."""
        contractor = self.get_object()
        labor_records = HarvestLabor.objects.filter(
            contractor=contractor
        ).select_related(
            'harvest', 'harvest__field', 'harvest__field__farm'
        ).order_by('-harvest__harvest_date')
        
        serializer = HarvestLaborSerializer(labor_records, many=True)
        
        # Calculate summary stats
        summary = labor_records.aggregate(
            total_jobs=Count('id'),
            total_bins=Sum('bins_picked'),
            total_cost=Sum('total_labor_cost'),
            total_hours=Sum('total_hours')
        )
        
        return Response({
            'contractor': LaborContractorSerializer(contractor).data,
            'summary': summary,
            'jobs': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get contractors with licenses/insurance expiring in next 30 days."""
        threshold = date.today() + timedelta(days=30)

        expiring = LaborContractor.objects.filter(
            Q(license_expiration__lte=threshold) |
            Q(insurance_expiration__lte=threshold) |
            Q(workers_comp_expiration__lte=threshold),
            active=True
        )

        serializer = LaborContractorSerializer(expiring, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def performance(self, request, pk=None):
        """
        Get performance metrics for this contractor.
        Returns: productivity (bins/hour), cost per bin, compliance history, reliability
        """
        from django.db.models import DecimalField

        contractor = self.get_object()

        # Get all labor records for this contractor
        labor_records = HarvestLabor.objects.filter(
            contractor=contractor
        ).select_related('harvest', 'harvest__field')

        # Overall statistics
        overall_stats = labor_records.aggregate(
            total_jobs=Count('id'),
            total_harvests=Count('harvest', distinct=True),
            total_bins=Sum('bins_picked'),
            total_hours=Sum('hours_worked'),
            total_cost=Sum('total_cost'),
            avg_bins_per_hour=Avg(F('bins_picked') / F('hours_worked'), output_field=DecimalField()),
            avg_cost_per_bin=Avg(F('total_cost') / F('bins_picked'), output_field=DecimalField()),
            avg_hourly_rate=Avg('rate', filter=Q(pay_type='hourly')),
            avg_piece_rate=Avg('rate', filter=Q(pay_type='piece_rate')),
        )

        # Performance by crop variety
        performance_by_crop = []
        for crop_value, crop_label in CROP_VARIETY_CHOICES:
            crop_choice = {'value': crop_value, 'label': crop_label}
            crop_records = labor_records.filter(harvest__crop_variety=crop_value)
            crop_stats = crop_records.aggregate(
                job_count=Count('id'),
                total_bins=Sum('bins_picked'),
                total_hours=Sum('hours_worked'),
                total_cost=Sum('total_cost'),
                avg_bins_per_hour=Avg(F('bins_picked') / F('hours_worked'), output_field=DecimalField()),
                avg_cost_per_bin=Avg(F('total_cost') / F('bins_picked'), output_field=DecimalField()),
            )

            if crop_stats['job_count'] > 0:
                performance_by_crop.append({
                    'crop_variety': crop_value,
                    'crop_variety_display': crop_choice['label'],
                    'job_count': crop_stats['job_count'],
                    'total_bins': crop_stats['total_bins'] or 0,
                    'total_hours': float(crop_stats['total_hours'] or 0),
                    'total_cost': float(crop_stats['total_cost'] or 0),
                    'avg_bins_per_hour': float(crop_stats['avg_bins_per_hour'] or 0),
                    'avg_cost_per_bin': float(crop_stats['avg_cost_per_bin'] or 0),
                })

        # Compliance status
        today = date.today()
        compliance_status = {
            'license_valid': contractor.license_expiration >= today if contractor.license_expiration else False,
            'license_expiration': contractor.license_expiration,
            'insurance_valid': contractor.insurance_expiration >= today if contractor.insurance_expiration else False,
            'insurance_expiration': contractor.insurance_expiration,
            'workers_comp_valid': contractor.workers_comp_expiration >= today if contractor.workers_comp_expiration else False,
            'workers_comp_expiration': contractor.workers_comp_expiration,
            'food_safety_training_current': contractor.food_safety_training_current,
            'overall_compliant': all([
                contractor.license_expiration >= today if contractor.license_expiration else False,
                contractor.insurance_expiration >= today if contractor.insurance_expiration else False,
                contractor.workers_comp_expiration >= today if contractor.workers_comp_expiration else False,
                contractor.food_safety_training_current
            ])
        }

        # Recent jobs (last 10)
        recent_jobs = labor_records.order_by('-harvest__harvest_date')[:10]
        recent_jobs_data = HarvestLaborSerializer(recent_jobs, many=True).data

        # Reliability metrics
        # Calculate consistency - standard deviation of bins per hour over recent jobs
        recent_productivity = []
        for record in labor_records.order_by('-harvest__harvest_date')[:20]:
            if record.hours_worked and record.hours_worked > 0:
                bins_per_hour = record.bins_picked / record.hours_worked
                recent_productivity.append(bins_per_hour)

        productivity_consistency = None
        if len(recent_productivity) > 1:
            import statistics
            avg = statistics.mean(recent_productivity)
            std_dev = statistics.stdev(recent_productivity)
            # Coefficient of variation (lower is more consistent)
            productivity_consistency = (std_dev / avg * 100) if avg > 0 else None

        return Response({
            'contractor': LaborContractorSerializer(contractor).data,
            'overall_stats': {
                'total_jobs': overall_stats['total_jobs'],
                'total_harvests': overall_stats['total_harvests'],
                'total_bins': overall_stats['total_bins'] or 0,
                'total_hours': float(overall_stats['total_hours'] or 0),
                'total_cost': float(overall_stats['total_cost'] or 0),
                'avg_bins_per_hour': float(overall_stats['avg_bins_per_hour'] or 0),
                'avg_cost_per_bin': float(overall_stats['avg_cost_per_bin'] or 0),
                'avg_hourly_rate': float(overall_stats['avg_hourly_rate'] or 0),
                'avg_piece_rate': float(overall_stats['avg_piece_rate'] or 0),
                'productivity_consistency': productivity_consistency,
            },
            'performance_by_crop': performance_by_crop,
            'compliance_status': compliance_status,
            'recent_jobs': recent_jobs_data
        })


# -----------------------------------------------------------------------------
# HARVEST VIEWSET
# -----------------------------------------------------------------------------

class HarvestViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    CRUD operations for Harvests with PHI checking and statistics.
    
    RLS NOTES:
    - Harvests inherit company through field->farm relationship
    - get_queryset filters by company
    """
    serializer_class = HarvestSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_queryset(self):
        from django.db.models import Prefetch

        # Optimize loads and labor_records prefetch with select_related
        loads_prefetch = Prefetch(
            'loads',
            queryset=HarvestLoad.objects.select_related('buyer').order_by('load_number')
        )
        labor_prefetch = Prefetch(
            'labor_records',
            queryset=HarvestLabor.objects.select_related('contractor').order_by('-start_time')
        )

        queryset = Harvest.objects.select_related(
            'field', 'field__farm'
        ).prefetch_related(loads_prefetch, labor_prefetch)

        # Filter by company
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(field__farm__company=company)
        
        # Filter by field
        field_id = self.request.query_params.get('field')
        if field_id:
            queryset = queryset.filter(field_id=field_id)
        
        # Filter by farm
        farm_id = self.request.query_params.get('farm')
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(harvest_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(harvest_date__lte=end_date)
        
        # Filter by crop variety
        crop = self.request.query_params.get('crop_variety')
        if crop:
            queryset = queryset.filter(crop_variety=crop)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by PHI compliance
        phi_compliant = self.request.query_params.get('phi_compliant')
        if phi_compliant == 'true':
            queryset = queryset.filter(phi_compliant=True)
        elif phi_compliant == 'false':
            queryset = queryset.filter(phi_compliant=False)
        
        # Season/date filter
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        season = self.request.query_params.get('season')

        if start_date and end_date:
            queryset = queryset.filter(harvest_date__gte=start_date, harvest_date__lte=end_date)
        elif season:
            if '-' in season:
                from .services.season_service import SeasonService
                try:
                    s_start, s_end = SeasonService().get_season_date_range(season, crop_category='citrus')
                    queryset = queryset.filter(harvest_date__gte=s_start, harvest_date__lte=s_end)
                except Exception:
                    pass
            else:
                try:
                    queryset = queryset.filter(harvest_date__year=int(season))
                except (ValueError, TypeError):
                    pass

        return queryset.order_by('-harvest_date', '-created_at')
    
    def get_serializer_class(self):
        if self.action == 'list':
            return HarvestListSerializer
        return HarvestSerializer
    
    @action(detail=False, methods=['post'])
    def check_phi(self, request):
        """
        Check PHI compliance before creating a harvest.
        POST: { "field_id": 1, "proposed_harvest_date": "2024-12-20" }
        """
        field_id = request.data.get('field_id')
        proposed_date_str = request.data.get('proposed_harvest_date')
        
        if not field_id or not proposed_date_str:
            return Response(
                {'error': 'field_id and proposed_harvest_date required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            proposed_date = datetime.strptime(proposed_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get most recent application
        last_app = PesticideApplication.objects.filter(
            field_id=field_id
        ).select_related('product').order_by('-application_date').first()
        
        if not last_app:
            return Response({
                'field_id': field_id,
                'proposed_harvest_date': proposed_date_str,
                'last_application_date': None,
                'last_application_product': None,
                'phi_required_days': None,
                'days_since_application': None,
                'is_compliant': True,
                'warning_message': 'No pesticide applications found for this field.'
            })
        
        days_since = (proposed_date - last_app.application_date).days
        phi_required = last_app.product.phi_days if last_app.product else None
        is_compliant = days_since >= phi_required if phi_required else None
        
        warning_message = None
        if is_compliant is False:
            warning_message = (
                f"PHI VIOLATION: Only {days_since} days since application of "
                f"{last_app.product.product_name}. Required: {phi_required} days. "
                f"Earliest safe harvest date: {last_app.application_date + timedelta(days=phi_required)}"
            )
        elif is_compliant is True:
            warning_message = f"PHI compliant. {days_since} days since last application (required: {phi_required})."
        
        return Response({
            'field_id': field_id,
            'proposed_harvest_date': proposed_date_str,
            'last_application_date': last_app.application_date,
            'last_application_product': last_app.product.product_name if last_app.product else None,
            'phi_required_days': phi_required,
            'days_since_application': days_since,
            'is_compliant': is_compliant,
            'warning_message': warning_message
        })
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get harvest statistics with optional filters.
        """
        queryset = self.get_queryset()
        
        # Base aggregations
        stats = queryset.aggregate(
            total_harvests=Count('id'),
            total_bins=Coalesce(Sum('total_bins'), 0),
            total_weight_lbs=Coalesce(Sum('estimated_weight_lbs'), Decimal('0')),
            total_acres_harvested=Coalesce(Sum('acres_harvested'), Decimal('0')),
        )
        
        # Revenue from loads
        load_stats = HarvestLoad.objects.filter(
            harvest__in=queryset
        ).aggregate(
            total_revenue=Coalesce(Sum('total_revenue'), Decimal('0')),
            pending_payments=Coalesce(
                Sum('total_revenue', filter=Q(payment_status='pending')),
                Decimal('0')
            ),
            avg_price_per_bin=Avg('price_per_unit', filter=Q(price_unit='per_bin'))
        )
        
        # Labor costs
        labor_stats = HarvestLabor.objects.filter(
            harvest__in=queryset
        ).aggregate(
            total_labor_cost=Coalesce(Sum('total_labor_cost'), Decimal('0'))
        )
        
        # PHI violations
        phi_violations = queryset.filter(phi_compliant=False).count()
        
        # Calculate yield per acre
        if stats['total_acres_harvested'] and stats['total_acres_harvested'] > 0:
            avg_yield = float(stats['total_bins']) / float(stats['total_acres_harvested'])
        else:
            avg_yield = 0
        
        # By crop breakdown
        by_crop = list(queryset.values('crop_variety').annotate(
            count=Count('id'),
            bins=Sum('total_bins'),
            acres=Sum('acres_harvested')
        ).order_by('-bins'))
        
        # By buyer breakdown
        by_buyer = list(HarvestLoad.objects.filter(
            harvest__in=queryset
        ).values('buyer__name').annotate(
            loads=Count('id'),
            bins=Sum('bins'),
            revenue=Sum('total_revenue')
        ).order_by('-revenue'))
        
        return Response({
            'total_harvests': stats['total_harvests'],
            'total_bins': stats['total_bins'],
            'total_weight_lbs': stats['total_weight_lbs'],
            'total_acres_harvested': stats['total_acres_harvested'],
            'total_revenue': load_stats['total_revenue'],
            'total_labor_cost': labor_stats['total_labor_cost'],
            'avg_yield_per_acre': round(avg_yield, 1),
            'avg_price_per_bin': load_stats['avg_price_per_bin'],
            'pending_payments': load_stats['pending_payments'],
            'phi_violations': phi_violations,
            'by_crop': by_crop,
            'by_buyer': by_buyer
        })
    
    @action(detail=True, methods=['post'])
    def mark_complete(self, request, pk=None):
        """Mark harvest as complete."""
        harvest = self.get_object()
        harvest.status = 'complete'
        harvest.save()
        return Response(HarvestSerializer(harvest).data)
    
    @action(detail=True, methods=['post'])
    def mark_verified(self, request, pk=None):
        """Mark harvest as verified (for GAP/GHP)."""
        harvest = self.get_object()
        
        # Check GAP/GHP requirements
        warnings = []
        if not harvest.phi_verified:
            warnings.append("PHI verification not checked")
        if not harvest.equipment_cleaned:
            warnings.append("Equipment cleaning not verified")
        if not harvest.no_contamination_observed:
            warnings.append("Contamination check not verified")
        
        harvest.status = 'verified'
        harvest.save()
        
        return Response({
            'harvest': HarvestSerializer(harvest).data,
            'warnings': warnings
        })
    
    @action(detail=False, methods=['get'])
    def by_field(self, request):
        """Get harvests grouped by field."""
        queryset = self.get_queryset()
        
        fields_data = {}
        for harvest in queryset:
            field_id = harvest.field_id
            if field_id not in fields_data:
                fields_data[field_id] = {
                    'field_id': field_id,
                    'field_name': harvest.field.name,
                    'farm_name': harvest.field.farm.name if harvest.field.farm else None,
                    'harvests': []
                }
            fields_data[field_id]['harvests'].append(
                HarvestListSerializer(harvest).data
            )
        
        return Response(list(fields_data.values()))

    @action(detail=False, methods=['get'])
    def cost_analysis(self, request):
        """
        Get cost analysis metrics for harvests.
        Returns: cost per acre, cost per bin, revenue per acre, labor efficiency metrics
        Filters: season, field, crop_variety, start_date, end_date
        """
        from django.db.models import Sum, Avg, Count, F, Q, DecimalField
        from django.db.models.functions import Coalesce

        queryset = self.get_queryset()

        # Apply date/season filters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        season = request.query_params.get('season')

        if start_date and end_date:
            queryset = queryset.filter(harvest_date__gte=start_date, harvest_date__lte=end_date)
        elif season:
            if '-' in season:
                # Cross-year season (e.g. "2024-2025") - resolve via SeasonService
                from .services.season_service import SeasonService
                try:
                    s_start, s_end = SeasonService().get_season_date_range(season, crop_category='citrus')
                    queryset = queryset.filter(harvest_date__gte=s_start, harvest_date__lte=s_end)
                except Exception:
                    pass
            else:
                try:
                    queryset = queryset.filter(harvest_date__year=int(season))
                except (ValueError, TypeError):
                    pass

        # Get aggregated metrics
        harvests_with_metrics = queryset.annotate(
            total_revenue=Coalesce(Sum('loads__total_price'), 0, output_field=DecimalField()),
            total_labor_cost=Coalesce(Sum('labor_records__total_cost'), 0, output_field=DecimalField()),
            total_labor_hours=Coalesce(Sum('labor_records__hours_worked'), 0, output_field=DecimalField()),
        )

        # Calculate overall metrics
        total_harvests = harvests_with_metrics.count()
        if total_harvests == 0:
            return Response({
                'total_harvests': 0,
                'metrics': {},
                'by_crop': [],
                'by_field': [],
                'by_contractor': []
            })

        aggregates = harvests_with_metrics.aggregate(
            total_acres=Sum('acres_harvested'),
            total_bins=Sum('total_bins'),
            total_revenue=Sum('total_revenue'),
            total_labor_cost=Sum('total_labor_cost'),
            total_labor_hours=Sum('total_labor_hours'),
            avg_cost_per_acre=Avg(F('total_labor_cost') / F('acres_harvested'), output_field=DecimalField()),
            avg_cost_per_bin=Avg(F('total_labor_cost') / F('total_bins'), output_field=DecimalField()),
            avg_revenue_per_acre=Avg(F('total_revenue') / F('acres_harvested'), output_field=DecimalField()),
        )

        # Calculate derived metrics
        metrics = {
            'total_acres': float(aggregates['total_acres'] or 0),
            'total_bins': aggregates['total_bins'] or 0,
            'total_revenue': float(aggregates['total_revenue'] or 0),
            'total_labor_cost': float(aggregates['total_labor_cost'] or 0),
            'total_profit': float((aggregates['total_revenue'] or 0) - (aggregates['total_labor_cost'] or 0)),
            'total_labor_hours': float(aggregates['total_labor_hours'] or 0),
            'avg_cost_per_acre': float(aggregates['avg_cost_per_acre'] or 0),
            'avg_cost_per_bin': float(aggregates['avg_cost_per_bin'] or 0),
            'avg_revenue_per_acre': float(aggregates['avg_revenue_per_acre'] or 0),
            'avg_bins_per_hour': float(aggregates['total_bins'] / aggregates['total_labor_hours']) if aggregates['total_labor_hours'] else 0,
            'profit_margin': float(((aggregates['total_revenue'] or 0) - (aggregates['total_labor_cost'] or 0)) / (aggregates['total_revenue'] or 1)) * 100 if aggregates['total_revenue'] else 0,
        }

        # Breakdown by crop variety
        by_crop = []
        for crop_value, crop_label in CROP_VARIETY_CHOICES:
            crop_choice = {'value': crop_value, 'label': crop_label}
            crop_harvests = harvests_with_metrics.filter(crop_variety=crop_value)
            crop_aggregates = crop_harvests.aggregate(
                count=Count('id'),
                total_acres=Sum('acres_harvested'),
                total_bins=Sum('total_bins'),
                total_revenue=Sum('total_revenue'),
                total_labor_cost=Sum('total_labor_cost'),
                avg_cost_per_bin=Avg(F('total_labor_cost') / F('total_bins'), output_field=DecimalField()),
            )

            if crop_aggregates['count'] > 0:
                by_crop.append({
                    'crop_variety': crop_value,
                    'crop_variety_display': crop_choice['label'],
                    'harvest_count': crop_aggregates['count'],
                    'total_acres': float(crop_aggregates['total_acres'] or 0),
                    'total_bins': crop_aggregates['total_bins'] or 0,
                    'total_revenue': float(crop_aggregates['total_revenue'] or 0),
                    'total_labor_cost': float(crop_aggregates['total_labor_cost'] or 0),
                    'profit': float((crop_aggregates['total_revenue'] or 0) - (crop_aggregates['total_labor_cost'] or 0)),
                    'avg_cost_per_bin': float(crop_aggregates['avg_cost_per_bin'] or 0),
                })

        # Breakdown by field
        from collections import defaultdict
        by_field_data = defaultdict(lambda: {
            'total_revenue': 0,
            'total_labor_cost': 0,
            'total_acres': 0,
            'total_bins': 0,
            'harvest_count': 0
        })

        for harvest in harvests_with_metrics:
            field_id = harvest.field_id
            by_field_data[field_id]['field_id'] = field_id
            by_field_data[field_id]['field_name'] = harvest.field.name
            by_field_data[field_id]['farm_name'] = harvest.field.farm.name if harvest.field.farm else None
            by_field_data[field_id]['total_revenue'] += float(harvest.total_revenue)
            by_field_data[field_id]['total_labor_cost'] += float(harvest.total_labor_cost)
            by_field_data[field_id]['total_acres'] += float(harvest.acres_harvested or 0)
            by_field_data[field_id]['total_bins'] += harvest.total_bins or 0
            by_field_data[field_id]['harvest_count'] += 1

        by_field = []
        for field_data in by_field_data.values():
            field_data['profit'] = field_data['total_revenue'] - field_data['total_labor_cost']
            field_data['revenue_per_acre'] = field_data['total_revenue'] / field_data['total_acres'] if field_data['total_acres'] > 0 else 0
            field_data['cost_per_bin'] = field_data['total_labor_cost'] / field_data['total_bins'] if field_data['total_bins'] > 0 else 0
            by_field.append(field_data)

        # Sort by profit (descending)
        by_field.sort(key=lambda x: x['profit'], reverse=True)

        # Breakdown by contractor (labor efficiency)
        from .models import LaborContractor
        company = get_user_company(request.user)
        contractors = LaborContractor.objects.filter(company=company, active=True)

        by_contractor = []
        for contractor in contractors:
            contractor_labor = HarvestLabor.objects.filter(
                contractor=contractor,
                harvest__in=queryset
            ).aggregate(
                total_bins=Sum('bins_picked'),
                total_hours=Sum('hours_worked'),
                total_cost=Sum('total_cost'),
                harvest_count=Count('harvest', distinct=True)
            )

            if contractor_labor['harvest_count'] and contractor_labor['total_bins']:
                by_contractor.append({
                    'contractor_id': contractor.id,
                    'contractor_name': contractor.name,
                    'harvest_count': contractor_labor['harvest_count'],
                    'total_bins': contractor_labor['total_bins'] or 0,
                    'total_hours': float(contractor_labor['total_hours'] or 0),
                    'total_cost': float(contractor_labor['total_cost'] or 0),
                    'bins_per_hour': float(contractor_labor['total_bins'] / contractor_labor['total_hours']) if contractor_labor['total_hours'] else 0,
                    'cost_per_bin': float(contractor_labor['total_cost'] / contractor_labor['total_bins']) if contractor_labor['total_bins'] else 0,
                })

        # Sort by bins per hour (descending) - most efficient first
        by_contractor.sort(key=lambda x: x['bins_per_hour'], reverse=True)

        return Response({
            'total_harvests': total_harvests,
            'metrics': metrics,
            'by_crop': by_crop,
            'by_field': by_field,
            'by_contractor': by_contractor
        })


# -----------------------------------------------------------------------------
# HARVEST LOAD VIEWSET
# -----------------------------------------------------------------------------

class HarvestLoadViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    CRUD operations for Harvest Loads.
    
    RLS NOTES:
    - Loads inherit company through harvest->field->farm relationship
    """
    serializer_class = HarvestLoadSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_queryset(self):
        queryset = HarvestLoad.objects.select_related(
            'harvest', 'harvest__field', 'harvest__field__farm', 'buyer'
        )
        
        # Filter by company
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(harvest__field__farm__company=company)
        
        # Filter by harvest
        harvest_id = self.request.query_params.get('harvest')
        if harvest_id:
            queryset = queryset.filter(harvest_id=harvest_id)
        
        # Filter by buyer
        buyer_id = self.request.query_params.get('buyer')
        if buyer_id:
            queryset = queryset.filter(buyer_id=buyer_id)
        
        # Filter by payment status
        payment_status = self.request.query_params.get('payment_status')
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
        
        # Filter by grade
        grade = self.request.query_params.get('grade')
        if grade:
            queryset = queryset.filter(grade=grade)
        
        # Date range (based on harvest date)
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(harvest__harvest_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(harvest__harvest_date__lte=end_date)
        
        return queryset.order_by('-harvest__harvest_date', 'load_number')
    
    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        """Mark load as paid."""
        load = self.get_object()
        load.payment_status = 'paid'
        load.payment_date = request.data.get('payment_date', date.today())
        load.save()
        return Response(HarvestLoadSerializer(load).data)
    
    @action(detail=False, methods=['get'])
    def pending_payments(self, request):
        """Get all loads with pending payments."""
        loads = self.get_queryset().filter(
            payment_status__in=['pending', 'invoiced']
        ).order_by('harvest__harvest_date')

        total_pending = loads.aggregate(total=Sum('total_revenue'))['total'] or 0

        return Response({
            'total_pending': total_pending,
            'loads': HarvestLoadSerializer(loads, many=True).data
        })

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get all loads with overdue payments."""
        from datetime import date

        loads = self.get_queryset().filter(
            payment_status__in=['pending', 'invoiced'],
            payment_due_date__isnull=False,
            payment_due_date__lt=date.today()
        ).order_by('payment_due_date')

        total_overdue = loads.aggregate(total=Sum('total_revenue'))['total'] or 0

        return Response({
            'total_overdue': total_overdue,
            'count': loads.count(),
            'loads': HarvestLoadSerializer(loads, many=True).data
        })


# -----------------------------------------------------------------------------
# HARVEST LABOR VIEWSET
# -----------------------------------------------------------------------------

class HarvestLaborViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    CRUD operations for Harvest Labor records.
    
    RLS NOTES:
    - Labor records inherit company through harvest->field->farm relationship
    """
    serializer_class = HarvestLaborSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_queryset(self):
        queryset = HarvestLabor.objects.select_related(
            'harvest', 'harvest__field', 'harvest__field__farm', 'contractor'
        )
        
        # Filter by company
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(harvest__field__farm__company=company)
        
        # Filter by harvest
        harvest_id = self.request.query_params.get('harvest')
        if harvest_id:
            queryset = queryset.filter(harvest_id=harvest_id)
        
        # Filter by contractor
        contractor_id = self.request.query_params.get('contractor')
        if contractor_id:
            queryset = queryset.filter(contractor_id=contractor_id)
        
        # Date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(harvest__harvest_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(harvest__harvest_date__lte=end_date)
        
        return queryset.order_by('-harvest__harvest_date')
    
    @action(detail=False, methods=['get'])
    def cost_analysis(self, request):
        """Get labor cost analysis."""
        queryset = self.get_queryset()
        
        analysis = queryset.aggregate(
            total_cost=Sum('total_labor_cost'),
            total_hours=Sum('total_hours'),
            total_bins=Sum('bins_picked'),
            avg_hourly_rate=Avg('rate', filter=Q(pay_type='hourly')),
            avg_piece_rate=Avg('rate', filter=Q(pay_type='piece_rate'))
        )
        
        # Calculate cost per bin
        if analysis['total_bins'] and analysis['total_cost']:
            analysis['cost_per_bin'] = float(analysis['total_cost']) / analysis['total_bins']
        else:
            analysis['cost_per_bin'] = None
        
        # By contractor breakdown
        by_contractor = list(queryset.values(
            'contractor__company_name'
        ).annotate(
            jobs=Count('id'),
            bins=Sum('bins_picked'),
            cost=Sum('total_labor_cost'),
            hours=Sum('total_hours')
        ).order_by('-cost'))
        
        return Response({
            **analysis,
            'by_contractor': by_contractor
        })

# =============================================================================
# MAP FEATURE ENDPOINTS
# =============================================================================

def get_plss_from_coordinates(lat, lng):
    """
    Get Section, Township, Range from GPS coordinates using BLM PLSS service.
    Returns dict with section, township, range (or None values if not found).
    """
    import requests as req
    import re
    
    result = {
        'section': None,
        'township': None,
        'range': None
    }
    
    try:
        # Use identify endpoint
        identify_url = "https://gis.blm.gov/arcgis/rest/services/Cadastral/BLM_Natl_PLSS_CadNSDI/MapServer/identify"
        
        buffer = 0.001
        identify_params = {
            'geometry': f'{lng},{lat}',
            'geometryType': 'esriGeometryPoint',
            'sr': '4326',
            'layers': 'all',
            'tolerance': '1',
            'mapExtent': f'{lng-buffer},{lat-buffer},{lng+buffer},{lat+buffer}',
            'imageDisplay': '100,100,96',
            'returnGeometry': 'false',
            'f': 'json'
        }
        
        response = req.get(identify_url, params=identify_params, timeout=15,
                          headers={'User-Agent': 'FarmManagementTracker/1.0'})
        
        if response.status_code == 200:
            data = response.json()
            
            if 'results' in data:
                for item in data['results']:
                    attrs = item.get('attributes', {})
                    
                    # Look for section number from multiple possible fields
                    if result['section'] is None:
                        # Try FRSTDIVNO first
                        if 'FRSTDIVNO' in attrs and attrs['FRSTDIVNO']:
                            result['section'] = str(attrs['FRSTDIVNO'])
                        # Try SECDIVNO
                        elif 'SECDIVNO' in attrs and attrs['SECDIVNO']:
                            result['section'] = str(attrs['SECDIVNO'])
                        # Try to extract from 'First Division Identifier' (format: CA210140S0200E0SN030)
                        elif 'First Division Identifier' in attrs and attrs['First Division Identifier']:
                            fdi = str(attrs['First Division Identifier'])
                            # Look for SN followed by digits (Section Number)
                            sec_match = re.search(r'SN0*(\d+)', fdi)
                            if sec_match:
                                result['section'] = sec_match.group(1)
                        # Try FRSTDIVID field
                        elif 'FRSTDIVID' in attrs and attrs['FRSTDIVID']:
                            fdi = str(attrs['FRSTDIVID'])
                            sec_match = re.search(r'SN0*(\d+)', fdi)
                            if sec_match:
                                result['section'] = sec_match.group(1)
                    
                    # Look for township/range in TWNSHPLAB
                    # Format can be: "T14S R20E" or "14S 20E" or "T14S-R20E" etc.
                    if 'TWNSHPLAB' in attrs and attrs['TWNSHPLAB']:
                        label = str(attrs['TWNSHPLAB'])
                        
                        # Try to find township (number followed by N or S)
                        if result['township'] is None:
                            twp_match = re.search(r'T?\.?\s*(\d+)\s*([NS])', label, re.IGNORECASE)
                            if twp_match:
                                result['township'] = f"{twp_match.group(1)}{twp_match.group(2).upper()}"
                        
                        # Try to find range (number followed by E or W)
                        if result['range'] is None:
                            rng_match = re.search(r'R?\.?\s*(\d+)\s*([EW])', label, re.IGNORECASE)
                            if rng_match:
                                result['range'] = f"{rng_match.group(1)}{rng_match.group(2).upper()}"
                    
                    # Also check individual fields as backup
                    if result['township'] is None and 'TWNSHPNO' in attrs and attrs['TWNSHPNO']:
                        twp_dir = attrs.get('TWNSHPDIR', 'N')
                        result['township'] = f"{attrs['TWNSHPNO']}{twp_dir}"
                    
                    if result['range'] is None and 'RANGENO' in attrs and attrs['RANGENO']:
                        rng_dir = attrs.get('RANGEDIR', 'E')
                        result['range'] = f"{attrs['RANGENO']}{rng_dir}"
        
        return result
        
    except Exception as e:
        print(f"PLSS lookup error: {e}")
        return result


@api_view(['POST'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def geocode_address(request):
    """
    Geocode an address to GPS coordinates using Nominatim (OpenStreetMap).
    Also returns PLSS data (Section, Township, Range) if available.

    Tries multiple address formats to improve success rate for rural addresses.

    POST /api/geocode/
    {
        "address": "123 Main St, Fresno, CA",
        "county": "Fresno",  # Optional - used to try additional search strategies
        "city": "Fresno"     # Optional - used to try additional search strategies
    }

    Returns:
    {
        "lat": 36.7378,
        "lng": -119.7871,
        "display_name": "...",
        "search_query": "...",  # The query that succeeded
        "section": "10",
        "township": "15S",
        "range": "22E",
        "alternatives": [...]  # Other potential matches if found
    }
    """
    address = request.data.get('address', '')
    county = request.data.get('county', '')
    city = request.data.get('city', '')

    if not address and not county:
        return Response({'error': 'Address or county is required'}, status=status.HTTP_400_BAD_REQUEST)

    def try_geocode(query):
        """Try to geocode a single query, return results or empty list."""
        try:
            response = requests.get(
                'https://nominatim.openstreetmap.org/search',
                params={
                    'q': query,
                    'format': 'json',
                    'limit': 5,  # Get multiple results for comparison
                    'countrycodes': 'us',
                    'addressdetails': 1  # Include address breakdown
                },
                headers={'User-Agent': 'FarmManagementTracker/1.0'},
                timeout=10
            )
            return response.json() if response.status_code == 200 else []
        except Exception:
            return []

    # Build list of address queries to try (in order of preference)
    queries_to_try = []

    # 1. Full address as provided
    if address:
        queries_to_try.append(address)

    # 2. Address with explicit California state
    if address and 'california' not in address.lower() and ', ca' not in address.lower():
        queries_to_try.append(f"{address}, California, USA")

    # 3. If county provided, try address + county + state
    if address and county:
        queries_to_try.append(f"{address}, {county} County, California, USA")
        # Also try without "County" suffix
        queries_to_try.append(f"{address}, {county}, California, USA")

    # 4. If city provided, try address + city + state
    if address and city:
        queries_to_try.append(f"{address}, {city}, California, USA")

    # 5. Try just county center (fallback for when address fails)
    if county:
        queries_to_try.append(f"{county} County, California, USA")

    # 6. Try structured query with street parsing (for rural addresses)
    if address:
        # Try extracting just the road/street name for rural areas
        import re
        # Match patterns like "1234 Some Road" or "12345 W Highway 99"
        road_match = re.search(r'\d+\s+(.+)', address.split(',')[0])
        if road_match:
            road_name = road_match.group(1).strip()
            if county:
                queries_to_try.append(f"{road_name}, {county} County, California, USA")
            if city:
                queries_to_try.append(f"{road_name}, {city}, California, USA")

    # Remove duplicates while preserving order
    seen = set()
    unique_queries = []
    for q in queries_to_try:
        q_lower = q.lower()
        if q_lower not in seen:
            seen.add(q_lower)
            unique_queries.append(q)

    # Try each query until we get a result
    best_result = None
    successful_query = None
    all_alternatives = []

    for query in unique_queries:
        results = try_geocode(query)

        if results:
            # Filter results to prefer California results
            ca_results = [r for r in results if 'california' in r.get('display_name', '').lower()]

            if ca_results:
                # Use the first California result as best
                if not best_result:
                    best_result = ca_results[0]
                    successful_query = query

                # Collect alternatives (excluding the best result)
                for r in ca_results[1:]:
                    if r not in all_alternatives:
                        all_alternatives.append(r)
            elif not best_result and results:
                # If no CA results, still use what we found
                best_result = results[0]
                successful_query = query

        # If we found a good result, don't try more queries
        if best_result:
            break

    if best_result:
        lat = float(best_result['lat'])
        lng = float(best_result['lon'])

        # Also get PLSS data (Section/Township/Range)
        plss_data = get_plss_from_coordinates(lat, lng)

        # Format alternatives for response
        formatted_alternatives = []
        for alt in all_alternatives[:4]:  # Limit to 4 alternatives
            formatted_alternatives.append({
                'lat': float(alt['lat']),
                'lng': float(alt['lon']),
                'display_name': alt.get('display_name', ''),
            })

        return Response({
            'lat': lat,
            'lng': lng,
            'display_name': best_result.get('display_name', ''),
            'search_query': successful_query,
            'section': plss_data.get('section'),
            'township': plss_data.get('township'),
            'range': plss_data.get('range'),
            'alternatives': formatted_alternatives,
        })
    else:
        # Return helpful error message
        tried_queries = ', '.join([f'"{q}"' for q in unique_queries[:3]])
        return Response({
            'error': 'Address not found',
            'tried_queries': unique_queries,
            'suggestion': 'Try entering a more specific address, nearby city name, or use the manual coordinate entry option.'
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
def update_field_boundary(request, field_id):
    """
    Update a field's boundary from a drawn polygon.
    Also updates total_acres and fetches PLSS data.
    
    POST /api/fields/{id}/boundary/
    {
        "boundary_geojson": {...},
        "calculated_acres": 45.5
    }
    """
    from .models import Field  # Import here to avoid circular imports
    
    print(f"[Boundary] update_field_boundary called for field_id: {field_id}")
    
    try:
        field = Field.objects.get(id=field_id)
        print(f"[Boundary] Found field: {field.name}")
    except Field.DoesNotExist:
        return Response({'error': 'Field not found'}, status=status.HTTP_404_NOT_FOUND)
    
    boundary = request.data.get('boundary_geojson')
    acres = request.data.get('calculated_acres')
    
    print(f"[Boundary] Received boundary: {boundary is not None}")
    print(f"[Boundary] Received acres: {acres}")
    
    if boundary:
        field.boundary_geojson = boundary
        
        # Calculate centroid of the polygon to get PLSS data
        try:
            coords = boundary.get('coordinates', [[]])[0]
            print(f"[Boundary] Coordinates count: {len(coords) if coords else 0}")
            if coords:
                # Calculate centroid
                lats = [c[1] for c in coords]
                lngs = [c[0] for c in coords]
                centroid_lat = sum(lats) / len(lats)
                centroid_lng = sum(lngs) / len(lngs)
                
                print(f"[Boundary] Centroid: {centroid_lat}, {centroid_lng}")
                
                # Update field GPS coordinates to centroid
                field.gps_lat = centroid_lat
                field.gps_long = centroid_lng
                
                # Get PLSS data
                print(f"[Boundary] Calling get_plss_from_coordinates...")
                plss_data = get_plss_from_coordinates(centroid_lat, centroid_lng)
                print(f"[Boundary] PLSS result: {plss_data}")
                
                if plss_data.get('section'):
                    field.section = plss_data['section']
                    print(f"[Boundary] Set section: {field.section}")
                if plss_data.get('township'):
                    field.township = plss_data['township']
                    print(f"[Boundary] Set township: {field.township}")
                if plss_data.get('range'):
                    field.range_value = plss_data['range']
                    print(f"[Boundary] Set range: {field.range_value}")
        except Exception as e:
            print(f"[Boundary] Error calculating centroid/PLSS: {e}")
            import traceback
            traceback.print_exc()
    
    if acres is not None:
        field.calculated_acres = acres
        field.total_acres = acres  # Also update main acres field
    
    field.save()
    
    return Response({
        'id': field.id,
        'name': field.name,
        'boundary_geojson': field.boundary_geojson,
        'calculated_acres': str(field.calculated_acres) if field.calculated_acres else None,
        'total_acres': str(field.total_acres) if field.total_acres else None,
        'gps_lat': str(field.gps_lat) if field.gps_lat else None,
        'gps_long': str(field.gps_long) if field.gps_long else None,
        'section': field.section,
        'township': field.township,
        'range': field.range_value,
        'message': 'Boundary saved successfully'
    })


@api_view(['POST'])
def get_plss(request):
    """
    Get Section, Township, Range from GPS coordinates.
    """
    lat = request.data.get('lat')
    lng = request.data.get('lng')
    
    if lat is None or lng is None:
        return Response({'error': 'lat and lng are required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        plss_data = get_plss_from_coordinates(float(lat), float(lng))
        return Response({
            'section': plss_data.get('section'),
            'township': plss_data.get('township'),
            'range': plss_data.get('range'),
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# -----------------------------------------------------------------------------
# HELPER FUNCTIONS
# -----------------------------------------------------------------------------

def get_current_water_year():
    """Get the current water year string (e.g., '2024-2025')."""
    today = date.today()
    if today.month >= 10:
        return f"{today.year}-{today.year + 1}"
    return f"{today.year - 1}-{today.year}"


def get_water_year_dates(water_year=None):
    """Get start and end dates for a water year."""
    if not water_year:
        water_year = get_current_water_year()
    
    start_year = int(water_year.split('-')[0])
    return {
        'start': date(start_year, 10, 1),
        'end': date(start_year + 1, 9, 30)
    }


def get_current_reporting_period():
    """Get the current semi-annual reporting period."""
    today = date.today()
    if today.month >= 10:
        # Oct-Mar = Period 1 of next year
        return {
            'period': f"{today.year + 1}-1",
            'type': 'semi_annual_1',
            'start': date(today.year, 10, 1),
            'end': date(today.year + 1, 3, 31)
        }
    elif today.month >= 4:
        # Apr-Sep = Period 2
        return {
            'period': f"{today.year}-2",
            'type': 'semi_annual_2',
            'start': date(today.year, 4, 1),
            'end': date(today.year, 9, 30)
        }
    else:
        # Jan-Mar = Period 1
        return {
            'period': f"{today.year}-1",
            'type': 'semi_annual_1',
            'start': date(today.year - 1, 10, 1),
            'end': date(today.year, 3, 31)
        }


# -----------------------------------------------------------------------------
# WELL VIEWSET
# -----------------------------------------------------------------------------

class WellViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    ViewSet for Well model.
    Provides CRUD operations plus custom actions for well management.
    """
    queryset = WaterSource.objects.filter(source_type="well").select_related(
        'water_source', 'water_source__farm'
    ).all()
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return WellListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return WellCreateSerializer
        return WellSerializer
    
    def get_queryset(self):
        """Filter wells by user's current company."""
        queryset = super().get_queryset()
        
        # Filter by company
        user = self.request.user
        if hasattr(user, 'current_company') and user.current_company:
            queryset = queryset.filter(
                water_source__farm__company=user.current_company
            )
        
        # Optional filters
        gsa = self.request.query_params.get('gsa')
        if gsa:
            queryset = queryset.filter(gsa=gsa)
        
        basin = self.request.query_params.get('basin')
        if basin:
            queryset = queryset.filter(basin=basin)
        
        farm_id = self.request.query_params.get('farm')
        if farm_id:
            queryset = queryset.filter(water_source__farm_id=farm_id)
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def readings(self, request, pk=None):
        """Get all readings for a specific well."""
        well = self.get_object()
        readings = water_source.readings.all()
        
        # Optional date filters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            readings = readings.filter(reading_date__gte=start_date)
        if end_date:
            readings = readings.filter(reading_date__lte=end_date)
        
        serializer = WellReadingListSerializer(readings, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def calibrations(self, request, pk=None):
        """Get calibration history for a specific well."""
        well = self.get_object()
        calibrations = water_source.calibrations.all()
        serializer = MeterCalibrationSerializer(calibrations, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def allocations(self, request, pk=None):
        """Get allocation history for a specific well."""
        well = self.get_object()
        allocations = water_source.allocations.all()
        
        water_year = request.query_params.get('water_year')
        if water_year:
            allocations = allocations.filter(water_year=water_year)
        
        serializer = WaterAllocationSerializer(allocations, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def extraction_summary(self, request, pk=None):
        """Get extraction summary for a well."""
        well = self.get_object()
        water_year = request.query_params.get('water_year', get_current_water_year())
        wy_dates = get_water_year_dates(water_year)
        
        # Get extraction
        extraction = water_source.readings.filter(
            reading_date__gte=wy_dates['start'],
            reading_date__lte=wy_dates['end']
        ).aggregate(
            total_af=Sum('extraction_acre_feet'),
            total_gallons=Sum('extraction_gallons'),
            reading_count=Count('id')
        )
        
        # Get allocation
        allocation = water_source.allocations.filter(
            water_year=water_year
        ).exclude(
            allocation_type='transferred_out'
        ).aggregate(
            total_af=Sum('allocated_acre_feet')
        )
        
        total_allocated = allocation['total_af'] or Decimal('0')
        total_extracted = extraction['total_af'] or Decimal('0')
        remaining = total_allocated - total_extracted
        
        return Response({
            'water_year': water_year,
            'well_id': well.id,
            'well_name': well.well_name or well.water_source.name,
            'total_allocated_af': float(total_allocated),
            'total_extracted_af': float(total_extracted),
            'remaining_af': float(remaining),
            'percent_used': float((total_extracted / total_allocated * 100) if total_allocated else 0),
            'is_over_allocation': remaining < 0,
            'reading_count': extraction['reading_count']
        })
    
    @action(detail=False, methods=['get'])
    def by_gsa(self, request):
        """Get wells grouped by GSA."""
        queryset = self.get_queryset()
        
        gsa_data = queryset.values('gsa').annotate(
            count=Count('id'),
            active_count=Count('id', filter=Q(status='active'))
        ).order_by('gsa')
        
        return Response(list(gsa_data))
    
    @action(detail=False, methods=['get'])
    def calibration_due(self, request):
        """Get wells with calibration due or overdue."""
        days = int(request.query_params.get('days', 30))
        warning_date = date.today() + timedelta(days=days)
        
        queryset = self.get_queryset().filter(
            Q(next_calibration_due__lte=warning_date) |
            Q(next_calibration_due__isnull=True),
            has_flowmeter=True,
            status='active'
        )
        
        serializer = WellListSerializer(queryset, many=True)
        return Response(serializer.data)


# -----------------------------------------------------------------------------
# WELL READING VIEWSET
# -----------------------------------------------------------------------------

class WellReadingViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """ViewSet for WellReading model."""

    queryset = WellReading.objects.select_related(
        'water_source', 'water_source__farm'
    ).all()
    permission_classes = [IsAuthenticated, HasCompanyAccess]

    def get_serializer_class(self):
        if self.action == 'list':
            return WellReadingListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return WellReadingCreateSerializer
        return WellReadingSerializer
    
    def get_queryset(self):
        """Filter by company and optional parameters."""
        queryset = WellReading.objects.select_related(
            'water_source', 'water_source__farm'
        )
        
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(
                water_source__farm__company=company
            )
        
        # Filter by water source (well)
        water_source_id = self.request.query_params.get('water_source') or self.request.query_params.get('well')
        if water_source_id:
            queryset = queryset.filter(water_source_id=water_source_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(reading_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(reading_date__lte=end_date)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def by_period(self, request):
        """Get readings for a specific reporting period."""
        period_type = request.query_params.get('period_type', 'semi_annual_1')
        water_year = request.query_params.get('water_year', get_current_water_year())
        
        wy_dates = get_water_year_dates(water_year)
        
        if period_type == 'semi_annual_1':
            start = wy_dates['start']
            end = date(wy_dates['start'].year + 1, 3, 31)
        elif period_type == 'semi_annual_2':
            start = date(wy_dates['end'].year, 4, 1)
            end = wy_dates['end']
        else:
            start = wy_dates['start']
            end = wy_dates['end']
        
        queryset = self.get_queryset().filter(
            reading_date__gte=start,
            reading_date__lte=end
        )
        
        serializer = WellReadingListSerializer(queryset, many=True)
        return Response({
            'period_type': period_type,
            'water_year': water_year,
            'start_date': start,
            'end_date': end,
            'readings': serializer.data
        })


# -----------------------------------------------------------------------------
# METER CALIBRATION VIEWSET
# -----------------------------------------------------------------------------

class MeterCalibrationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """ViewSet for MeterCalibration model."""
    
    queryset = MeterCalibration.objects.select_related(
        'water_source', 'water_source__farm'
    ).all()
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return MeterCalibrationCreateSerializer
        return MeterCalibrationSerializer
    
    def get_queryset(self):
        queryset = MeterCalibration.objects.select_related(
            'water_source', 'water_source__farm'
        )
        
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(
                water_source__farm__company=company
            )
        
        well_id = self.request.query_params.get('well')
        if well_id:
            queryset = queryset.filter(water_source_id=well_id)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def expiring(self, request):
        """Get calibrations expiring within specified days."""
        days = int(request.query_params.get('days', 90))
        expiry_date = date.today() + timedelta(days=days)
        
        queryset = self.get_queryset().filter(
            next_calibration_due__lte=expiry_date,
            next_calibration_due__gte=date.today()
        ).order_by('next_calibration_due')
        
        serializer = MeterCalibrationSerializer(queryset, many=True)
        return Response(serializer.data)


# -----------------------------------------------------------------------------
# WATER ALLOCATION VIEWSET
# -----------------------------------------------------------------------------

class WaterAllocationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """ViewSet for WaterAllocation model."""
    
    queryset = WaterAllocation.objects.select_related(
        'well', 'water_source', 'water_source__farm'
    ).all()
    serializer_class = WaterAllocationSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        user = self.request.user
        if hasattr(user, 'current_company') and user.current_company:
            queryset = queryset.filter(
                water_source__farm__company=user.current_company
            )
        
        well_id = self.request.query_params.get('well')
        if well_id:
            queryset = queryset.filter(water_source_id=well_id)
        
        water_year = self.request.query_params.get('water_year')
        if water_year:
            queryset = queryset.filter(water_year=water_year)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get allocation vs extraction summary for all wells."""
        water_year = request.query_params.get('water_year', get_current_water_year())
        wy_dates = get_water_year_dates(water_year)
        
        user = request.user
        if hasattr(user, 'current_company') and user.current_company:
            wells = WaterSource.objects.filter(source_type="well").filter(
                water_source__farm__company=user.current_company,
                status='active'
            )
        else:
            wells = WaterSource.objects.filter(source_type="well").filter(status='active')
        
        summaries = []
        for well in wells:
            allocation = water_source.allocations.filter(
                water_year=water_year
            ).exclude(
                allocation_type='transferred_out'
            ).aggregate(total=Sum('allocated_acre_feet'))['total'] or Decimal('0')
            
            extraction = water_source.readings.filter(
                reading_date__gte=wy_dates['start'],
                reading_date__lte=wy_dates['end']
            ).aggregate(total=Sum('extraction_acre_feet'))['total'] or Decimal('0')
            
            remaining = allocation - extraction
            
            summaries.append({
                'water_year': water_year,
                'well_id': well.id,
                'well_name': well.well_name or well.water_source.name,
                'gsa': well.gsa,
                'total_allocated_af': float(allocation),
                'total_extracted_af': float(extraction),
                'remaining_af': float(remaining),
                'percent_used': float((extraction / allocation * 100) if allocation else 0),
                'is_over_allocation': remaining < 0
            })
        
        return Response(summaries)


# -----------------------------------------------------------------------------
# EXTRACTION REPORT VIEWSET
# -----------------------------------------------------------------------------

class ExtractionReportViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """ViewSet for ExtractionReport model."""
    
    queryset = ExtractionReport.objects.select_related(
        'well', 'water_source', 'water_source__farm'
    ).all()
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ExtractionReportListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return ExtractionReportCreateSerializer
        return ExtractionReportSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        user = self.request.user
        if hasattr(user, 'current_company') and user.current_company:
            queryset = queryset.filter(
                water_source__farm__company=user.current_company
            )
        
        well_id = self.request.query_params.get('well')
        if well_id:
            queryset = queryset.filter(water_source_id=well_id)
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        gsa = self.request.query_params.get('gsa')
        if gsa:
            queryset = queryset.filter(well__gsa=gsa)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def generate(self, request):
        """Auto-generate extraction report from readings."""
        well_id = request.data.get('well_id')
        period_type = request.data.get('period_type', 'semi_annual_1')
        reporting_period = request.data.get('reporting_period')
        
        if not well_id:
            return Response(
                {'error': 'well_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            well = WaterSource.objects.filter(source_type="well").get(id=well_id)
        except Well.DoesNotExist:
            return Response(
                {'error': 'Well not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Determine period dates
        if not reporting_period:
            period_info = get_current_reporting_period()
            reporting_period = period_info['period']
            period_start = period_info['start']
            period_end = period_info['end']
        else:
            parts = reporting_period.split('-')
            year = int(parts[0])
            period_num = int(parts[1])
            
            if period_num == 1:
                period_start = date(year - 1, 10, 1)
                period_end = date(year, 3, 31)
            else:
                period_start = date(year, 4, 1)
                period_end = date(year, 9, 30)
        
        # Check if report already exists
        if ExtractionReport.objects.filter(
            well=well,
            reporting_period=reporting_period
        ).exists():
            return Response(
                {'error': f'Report for {reporting_period} already exists'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get readings for the period
        readings = WellReading.objects.filter(
            well=well,
            reading_date__gte=period_start,
            reading_date__lte=period_end
        ).order_by('reading_date', 'reading_time')
        
        if not readings.exists():
            return Response(
                {'error': 'No readings found for this period'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        first_reading = readings.first()
        last_reading = readings.last()
        
        pre_period_reading = WellReading.objects.filter(
            well=well,
            reading_date__lt=period_start
        ).order_by('-reading_date', '-reading_time').first()
        
        beginning_reading = pre_period_reading.meter_reading if pre_period_reading else first_reading.meter_reading
        beginning_date = pre_period_reading.reading_date if pre_period_reading else first_reading.reading_date
        
        water_year = f"{period_start.year}-{period_end.year}" if period_start.month >= 10 else f"{period_start.year - 1}-{period_start.year}"
        allocation = water_source.allocations.filter(
            water_year=water_year
        ).exclude(
            allocation_type='transferred_out'
        ).aggregate(total=Sum('allocated_acre_feet'))['total'] or Decimal('0')
        
        period_allocation = allocation / 2 if period_type.startswith('semi_annual') else allocation
        
        report = ExtractionReport.objects.create(
            well=well,
            period_type=period_type,
            reporting_period=reporting_period,
            period_start_date=period_start,
            period_end_date=period_end,
            beginning_meter_reading=beginning_reading,
            beginning_reading_date=beginning_date,
            ending_meter_reading=last_reading.meter_reading,
            ending_reading_date=last_reading.reading_date,
            period_allocation_af=period_allocation,
            status='draft'
        )
        
        serializer = ExtractionReportSerializer(report)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Mark report as submitted."""
        report = self.get_object()
        
        if report.status == 'submitted':
            return Response(
                {'error': 'Report already submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        report.status = 'submitted'
        report.submitted_date = date.today()
        report.save()
        
        serializer = ExtractionReportSerializer(report)
        return Response(serializer.data)


# -----------------------------------------------------------------------------
# IRRIGATION EVENT VIEWSET
# -----------------------------------------------------------------------------

class IrrigationEventViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """ViewSet for IrrigationEvent model."""
    
    queryset = IrrigationEvent.objects.select_related(
        'field', 'field__farm', 'well', 'water_source'
    ).all()
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return IrrigationEventListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return IrrigationEventCreateSerializer
        return IrrigationEventSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        user = self.request.user
        if hasattr(user, 'current_company') and user.current_company:
            queryset = queryset.filter(
                field__farm__company=user.current_company
            )
        
        field_id = self.request.query_params.get('field')
        if field_id:
            queryset = queryset.filter(field_id=field_id)
        
        well_id = self.request.query_params.get('well')
        if well_id:
            queryset = queryset.filter(water_source_id=well_id)
        
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(irrigation_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(irrigation_date__lte=end_date)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def by_field(self, request):
        """Get irrigation totals grouped by field."""
        queryset = self.get_queryset()
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(irrigation_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(irrigation_date__lte=end_date)
        
        field_data = queryset.values(
            'field__id', 'field__name', 'field__farm__name'
        ).annotate(
            total_af=Sum('water_applied_af'),
            total_gallons=Sum('water_applied_gallons'),
            event_count=Count('id'),
            total_hours=Sum('duration_hours')
        ).order_by('field__farm__name', 'field__name')
        
        return Response(list(field_data))


# -----------------------------------------------------------------------------
# SGMA DASHBOARD VIEW
# -----------------------------------------------------------------------------

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def sgma_dashboard(request):
    """
    Get SGMA compliance dashboard data.
    Returns summary statistics, alerts, and status for all wells.
    """
    user = request.user
    
    if hasattr(user, 'current_company') and user.current_company:
        wells = WaterSource.objects.filter(
            source_type="well",
            farm__company=user.current_company
        )
    else:
        wells = WaterSource.objects.filter(source_type="well")
    
    water_year = get_current_water_year()
    wy_dates = get_water_year_dates(water_year)
    current_period = get_current_reporting_period()
    
    # Well counts
    total_wells = wells.count()
    active_wells = wells.filter(active=True).count()
    wells_with_ami = wells.filter(has_ami=True).count()
    
    # YTD extraction
    ytd_extraction = WellReading.objects.filter(
        water_source__in=wells,
        reading_date__gte=wy_dates['start'],
        reading_date__lte=date.today()
    ).aggregate(total=Sum('extraction_acre_feet'))['total'] or Decimal('0')

    # YTD allocation
    ytd_allocation = WaterAllocation.objects.filter(
        water_source__in=wells,
        water_year=water_year
    ).exclude(
        allocation_type='transferred_out'
    ).aggregate(total=Sum('allocated_acre_feet'))['total'] or Decimal('0')
    
    allocation_remaining = ytd_allocation - ytd_extraction
    percent_used = (ytd_extraction / ytd_allocation * 100) if ytd_allocation else Decimal('0')
    
    # Current period extraction
    current_period_extraction = WellReading.objects.filter(
        water_source__in=wells,
        reading_date__gte=current_period['start'],
        reading_date__lte=min(current_period['end'], date.today())
    ).aggregate(total=Sum('extraction_acre_feet'))['total'] or Decimal('0')
    
    # Calibration status
    today = date.today()
    calibrations_current = wells.filter(
        meter_calibration_current=True,
        next_calibration_due__gt=today
    ).count()
    
    calibrations_due_soon = wells.filter(
        next_calibration_due__lte=today + timedelta(days=90),
        next_calibration_due__gt=today
    ).count()
    
    calibrations_overdue = wells.filter(
        Q(next_calibration_due__lte=today) |
        Q(next_calibration_due__isnull=True, has_flowmeter=True),
        active=True
    ).count()
    
    # Next deadlines
    next_calibration = wells.filter(
        next_calibration_due__gte=today
    ).order_by('next_calibration_due').values_list(
        'next_calibration_due', flat=True
    ).first()
    
    if today.month <= 3:
        next_report_due = date(today.year, 4, 1)
    elif today.month <= 9:
        next_report_due = date(today.year, 10, 1)
    else:
        next_report_due = date(today.year + 1, 4, 1)
    
    # Generate alerts
    alerts = []
    
    if calibrations_overdue > 0:
        alerts.append({
            'type': 'warning',
            'category': 'calibration',
            'message': f'{calibrations_overdue} well(s) have overdue meter calibrations',
            'action': 'Schedule calibration appointments'
        })
    
    if calibrations_due_soon > 0:
        alerts.append({
            'type': 'info',
            'category': 'calibration',
            'message': f'{calibrations_due_soon} well(s) have calibrations due within 90 days',
            'action': 'Plan upcoming calibrations'
        })
    
    if allocation_remaining < 0:
        alerts.append({
            'type': 'error',
            'category': 'allocation',
            'message': f'Over allocation by {abs(float(allocation_remaining)):.2f} AF',
            'action': 'Review extraction and consider purchasing additional allocation'
        })
    elif ytd_allocation > 0 and percent_used > 80:
        alerts.append({
            'type': 'warning',
            'category': 'allocation',
            'message': f'{float(percent_used):.1f}% of annual allocation used',
            'action': 'Monitor extraction closely'
        })
    
    # Wells by GSA
    wells_by_gsa_raw = wells.values('gsa').annotate(
        count=Count('id'),
        active_count=Count('id', filter=Q(active=True)),
        ytd_extraction=Sum(
            'readings__extraction_acre_feet',
            filter=Q(readings__reading_date__gte=wy_dates['start'])
        )
    ).order_by('gsa')

    # Convert Decimal to float for JSON serialization
    wells_by_gsa = []
    for item in wells_by_gsa_raw:
        wells_by_gsa.append({
            'gsa': item['gsa'],
            'count': item['count'],
            'active': item['active_count'],
            'ytd_extraction': float(item['ytd_extraction'] or 0)
        })
    
    # Recent readings
    recent_readings = WellReading.objects.filter(
        water_source__in=wells
    ).select_related('water_source').order_by('-reading_date', '-reading_time')[:10]
    
    recent_readings_data = WellReadingListSerializer(recent_readings, many=True).data
    
    return Response({
        'total_wells': total_wells,
        'active_wells': active_wells,
        'wells_with_ami': wells_with_ami,
        
        'ytd_extraction_af': float(ytd_extraction),
        'ytd_allocation_af': float(ytd_allocation),
        'allocation_remaining_af': float(allocation_remaining),
        'percent_allocation_used': float(percent_used),
        
        'current_period': current_period['period'],
        'current_period_extraction_af': float(current_period_extraction),
        'current_period_start': str(current_period['start']),
        'current_period_end': str(current_period['end']),
        
        'calibrations_current': calibrations_current,
        'calibrations_due_soon': calibrations_due_soon,
        'calibrations_overdue': calibrations_overdue,
        
        'next_report_due': str(next_report_due) if next_report_due else None,
        'next_calibration_due': str(next_calibration) if next_calibration else None,
        
        'alerts': alerts,
        'wells_by_gsa': wells_by_gsa,
        'recent_readings': recent_readings_data,
        
        'water_year': water_year,
        'as_of_date': str(date.today())
    })

# =============================================================================
# NUTRIENT MANAGEMENT VIEWS
# =============================================================================

class FertilizerProductViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """API endpoint for fertilizer products."""
    serializer_class = FertilizerProductSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'manufacturer', 'product_code']
    ordering_fields = ['name', 'nitrogen_pct', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        user = self.request.user
        company = getattr(user, 'current_company', None)
        queryset = FertilizerProduct.objects.filter(active=True)
        
        if company:
            queryset = queryset.filter(Q(company__isnull=True) | Q(company=company))
        else:
            queryset = queryset.filter(company__isnull=True)
        
        form = self.request.query_params.get('form')
        if form:
            queryset = queryset.filter(form=form)
        
        is_organic = self.request.query_params.get('is_organic')
        if is_organic is not None:
            queryset = queryset.filter(is_organic=is_organic.lower() == 'true')
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return FertilizerProductListSerializer
        return FertilizerProductSerializer
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        q = request.query_params.get('q', '')
        if len(q) < 2:
            return Response([])
        queryset = self.get_queryset().filter(
            Q(name__icontains=q) | Q(manufacturer__icontains=q)
        )[:20]
        return Response(FertilizerProductListSerializer(queryset, many=True).data)
    
    @action(detail=False, methods=['post'])
    def seed_common(self, request):
        from .models import get_common_fertilizers
        created, existing = 0, 0
        for data in get_common_fertilizers():
            _, was_created = FertilizerProduct.objects.get_or_create(name=data['name'], defaults=data)
            created += was_created
            existing += not was_created
        return Response({'created': created, 'existing': existing})


class NutrientApplicationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """API endpoint for nutrient applications."""
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['product__name', 'field__name', 'notes']
    ordering_fields = ['application_date', 'created_at', 'total_lbs_nitrogen']
    ordering = ['-application_date', '-created_at']
    
    def get_queryset(self):
        user = self.request.user
        company = getattr(user, 'current_company', None)
        queryset = NutrientApplication.objects.select_related('field', 'field__farm', 'product', 'water_source')
        
        if company:
            queryset = queryset.filter(field__farm__company=company)
        
        field_id = self.request.query_params.get('field')
        if field_id:
            queryset = queryset.filter(field_id=field_id)
        
        farm_id = self.request.query_params.get('farm')
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)
        
        year = self.request.query_params.get('year')
        if year:
            queryset = queryset.filter(application_date__year=year)
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return NutrientApplicationListSerializer
        return NutrientApplicationSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['get'])
    def by_field(self, request):
        year = request.query_params.get('year', date.today().year)
        queryset = self.get_queryset().filter(application_date__year=year)
        
        by_field = queryset.values(
            'field__id', 'field__name', 'field__farm__name', 'field__total_acres'
        ).annotate(
            application_count=Count('id'),
            total_lbs_nitrogen=Sum('total_lbs_nitrogen'),
            total_cost=Sum('total_cost')
        ).order_by('field__farm__name', 'field__name')
        
        results = []
        for item in by_field:
            acres = item['field__total_acres'] or Decimal('1')
            total_n = item['total_lbs_nitrogen'] or Decimal('0')
            results.append({
                'field_id': item['field__id'],
                'field_name': item['field__name'],
                'farm_name': item['field__farm__name'],
                'acres': float(acres),
                'application_count': item['application_count'],
                'total_lbs_nitrogen': float(total_n),
                'lbs_nitrogen_per_acre': float(total_n / acres) if acres else 0,
                'total_cost': float(item['total_cost']) if item['total_cost'] else None,
            })
        return Response(results)


class NutrientPlanViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """API endpoint for nitrogen management plans."""
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['year', 'field__name']
    ordering = ['-year', 'field__name']
    
    def get_queryset(self):
        user = self.request.user
        company = getattr(user, 'current_company', None)
        queryset = NutrientPlan.objects.select_related('field', 'field__farm')
        
        if company:
            queryset = queryset.filter(field__farm__company=company)
        
        year = self.request.query_params.get('year')
        if year:
            queryset = queryset.filter(year=year)
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return NutrientPlanListSerializer
        return NutrientPlanSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def nitrogen_summary(request):
    """Annual nitrogen summary by field for ILRP compliance."""
    user = request.user
    company = getattr(user, 'current_company', None)
    year = int(request.query_params.get('year', date.today().year))
    
    fields = Field.objects.filter(active=True).select_related('farm')
    if company:
        fields = fields.filter(farm__company=company)
    
    summary = []
    for field in fields:
        apps = field.nutrient_applications.filter(application_date__year=year)
        total_n = apps.aggregate(total=Sum('total_lbs_nitrogen'))['total'] or Decimal('0')
        acres = field.total_acres or Decimal('1')
        plan = field.nutrient_plans.filter(year=year).first()
        
        summary.append({
            'field_id': field.id,
            'field_name': field.name,
            'farm_name': field.farm.name if field.farm else '',
            'acres': float(acres),
            'crop': field.current_crop,
            'total_applications': apps.count(),
            'total_lbs_nitrogen': float(total_n),
            'lbs_nitrogen_per_acre': float(total_n / acres),
            'has_plan': plan is not None,
            'planned_nitrogen_lbs_acre': float(plan.net_planned_nitrogen) if plan else None,
            'variance_lbs_acre': float(plan.nitrogen_variance_per_acre) if plan else None,
        })
    return Response(summary)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def nitrogen_export(request):
    """Export nitrogen summary as Excel."""
    user = request.user
    company = getattr(user, 'current_company', None)
    year = int(request.query_params.get('year', date.today().year))
    
    fields = Field.objects.filter(active=True).select_related('farm')
    if company:
        fields = fields.filter(farm__company=company)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Nitrogen Report {year}"
    
    headers = ['Field', 'Farm', 'Acres', 'Crop', 'Total N (lbs)', 'N/Acre (lbs)', 'Applications']
    ws.append(headers)
    
    for field in fields:
        apps = field.nutrient_applications.filter(application_date__year=year)
        total_n = apps.aggregate(total=Sum('total_lbs_nitrogen'))['total'] or 0
        acres = float(field.total_acres or 1)
        
        ws.append([
            field.name,
            field.farm.name if field.farm else '',
            acres,
            field.current_crop,
            round(float(total_n), 1),
            round(float(total_n) / acres, 1),
            apps.count(),
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="nitrogen_report_{year}.xlsx"'
    return response


# =============================================================================
# QUARANTINE STATUS VIEWS
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def check_quarantine_status(request):
    """
    Check quarantine status for a farm or field.

    Query Parameters:
        farm_id: ID of the farm to check (mutually exclusive with field_id)
        field_id: ID of the field to check (mutually exclusive with farm_id)
        refresh: Set to 'true' to bypass cache and force a fresh check
        quarantine_type: Type of quarantine to check (default: 'HLB')

    Returns:
        QuarantineStatus object with in_quarantine, zone_name, last_checked, etc.
    """
    from .services.quarantine_service import CDFAQuarantineService
    from datetime import timedelta

    farm_id = request.query_params.get('farm_id')
    field_id = request.query_params.get('field_id')
    refresh = request.query_params.get('refresh', 'false').lower() == 'true'
    quarantine_type = request.query_params.get('quarantine_type', 'HLB')

    # Validate parameters
    if farm_id and field_id:
        return Response(
            {'error': 'Cannot specify both farm_id and field_id'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not farm_id and not field_id:
        return Response(
            {'error': 'Must specify either farm_id or field_id'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Get the farm or field (with company filtering for tenant isolation)
    company = get_user_company(request.user)
    farm = None
    field = None
    latitude = None
    longitude = None

    try:
        if farm_id:
            queryset = Farm.objects.filter(active=True)
            if company:
                queryset = queryset.filter(company=company)
            farm = queryset.get(id=farm_id)
            latitude = farm.gps_latitude
            longitude = farm.gps_longitude
        else:
            queryset = Field.objects.filter(active=True).select_related('farm')
            if company:
                queryset = queryset.filter(farm__company=company)
            field = queryset.get(id=field_id)
            latitude = field.gps_latitude
            longitude = field.gps_longitude
    except (Farm.DoesNotExist, Field.DoesNotExist):
        return Response(
            {'error': 'Farm or field not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Check for missing coordinates
    if latitude is None or longitude is None:
        return Response({
            'in_quarantine': None,
            'zone_name': None,
            'last_checked': None,
            'error': 'No GPS coordinates available for this location',
            'target_name': farm.name if farm else field.name,
            'target_type': 'farm' if farm else 'field',
        })

    # Try to get cached status (wrap in try-except for RLS issues)
    cache_filter = {
        'quarantine_type': quarantine_type,
    }
    if farm:
        cache_filter['farm'] = farm
    else:
        cache_filter['field'] = field

    try:
        cached_status = QuarantineStatus.objects.filter(**cache_filter).first()
    except Exception as db_error:
        # RLS policy may be blocking - log and continue without cache
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"RLS error querying QuarantineStatus: {db_error}")
        cached_status = None

    # Check if cache is valid (less than 24 hours old) and not forcing refresh
    if cached_status and not refresh:
        cache_age = timezone.now() - cached_status.last_checked
        if cache_age < timedelta(hours=24):
            serializer = QuarantineStatusSerializer(cached_status)
            return Response(serializer.data)

    # Query CDFA API
    service = CDFAQuarantineService()
    result = service.check_location(latitude, longitude)

    # Determine if status changed
    previous_status = cached_status.in_quarantine if cached_status else None
    status_changed = previous_status != result['in_quarantine'] and result['in_quarantine'] is not None

    # Save or update the status record
    status_data = {
        'in_quarantine': result['in_quarantine'],
        'zone_name': result['zone_name'] or '',
        'check_latitude': latitude,
        'check_longitude': longitude,
        'raw_response': result['raw_response'],
        'error_message': result['error'] or '',
    }

    if status_changed:
        status_data['last_changed'] = timezone.now()

    # Try to save status (wrap in try-except for RLS issues)
    try:
        if cached_status:
            for key, value in status_data.items():
                setattr(cached_status, key, value)
            cached_status.save()
            quarantine_status = cached_status
        else:
            quarantine_status = QuarantineStatus.objects.create(
                farm=farm,
                field=field,
                quarantine_type=quarantine_type,
                **status_data
            )
        serializer = QuarantineStatusSerializer(quarantine_status)
        return Response(serializer.data)
    except Exception as db_error:
        # RLS policy may be blocking saves - return result without caching
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"RLS error saving QuarantineStatus: {db_error}")

        # Return the CDFA result directly without database caching
        return Response({
            'in_quarantine': result['in_quarantine'],
            'zone_name': result['zone_name'] or '',
            'last_checked': timezone.now().isoformat(),
            'error_message': result['error'] or '',
            'target_name': farm.name if farm else field.name,
            'target_type': 'farm' if farm else 'field',
            '_cache_error': 'Unable to cache result - please run database migrations',
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def get_quarantine_boundaries(request):
    """
    Proxy endpoint for CDFA quarantine boundary GeoJSON.

    Returns GeoJSON FeatureCollection of all active quarantine zones.
    Caches the response for 1 hour to minimize external API calls.

    Query Parameters:
        refresh: Set to 'true' to bypass cache and fetch fresh data
    """
    from django.core.cache import cache

    CACHE_KEY = 'cdfa_quarantine_boundaries'
    CACHE_TIMEOUT = 3600  # 1 hour

    refresh = request.query_params.get('refresh', 'false').lower() == 'true'

    # Try to get from cache unless refreshing
    if not refresh:
        cached_data = cache.get(CACHE_KEY)
        if cached_data:
            return Response(cached_data)

    # Fetch from CDFA FeatureServer
    try:
        cdfa_url = (
            "https://gis2.cdfa.ca.gov/server/rest/services/Plant/ActiveQuarantines/FeatureServer/0/query"
        )
        params = {
            'where': '1=1',
            'outFields': '*',
            'returnGeometry': 'true',
            'outSR': '4326',  # WGS84 for Leaflet
            'f': 'geojson',
        }

        response = requests.get(cdfa_url, params=params, timeout=30)
        response.raise_for_status()

        geojson_data = response.json()

        # Validate it's a valid GeoJSON FeatureCollection
        if geojson_data.get('type') != 'FeatureCollection':
            return Response(
                {'error': 'Invalid response from CDFA API'},
                status=status.HTTP_502_BAD_GATEWAY
            )

        # Cache the response
        cache.set(CACHE_KEY, geojson_data, CACHE_TIMEOUT)

        return Response(geojson_data)

    except requests.exceptions.Timeout:
        return Response(
            {'error': 'CDFA API request timed out'},
            status=status.HTTP_504_GATEWAY_TIMEOUT
        )
    except requests.exceptions.RequestException as e:
        return Response(
            {'error': f'Failed to fetch quarantine boundaries: {str(e)}'},
            status=status.HTTP_502_BAD_GATEWAY
        )
    except ValueError as e:
        return Response(
            {'error': f'Invalid JSON response from CDFA API: {str(e)}'},
            status=status.HTTP_502_BAD_GATEWAY
        )


# =============================================================================
# IRRIGATION SCHEDULING VIEWS
# =============================================================================

class IrrigationZoneViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing irrigation zones.

    RLS NOTES:
    - Zones inherit company from their Field -> Farm
    - get_queryset filters by company through field.farm relationship
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'crop_type', 'field__name', 'field__farm__name']
    ordering_fields = ['name', 'acres', 'created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return IrrigationZoneListSerializer
        if self.action == 'retrieve':
            return IrrigationZoneDetailSerializer
        return IrrigationZoneSerializer

    def get_queryset(self):
        """Filter zones by current user's company through field.farm."""
        queryset = IrrigationZone.objects.filter(active=True).select_related(
            'field', 'field__farm', 'water_source'
        )
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(field__farm__company=company)

        # Optional filters
        field_id = self.request.query_params.get('field')
        if field_id:
            queryset = queryset.filter(field_id=field_id)

        farm_id = self.request.query_params.get('farm')
        if farm_id:
            queryset = queryset.filter(field__farm_id=farm_id)

        return queryset

    @action(detail=True, methods=['get'])
    def status(self, request, pk=None):
        """Get current irrigation status and recommendation for a zone."""
        zone = self.get_object()
        from .services.irrigation_scheduler import IrrigationScheduler

        try:
            scheduler = IrrigationScheduler(zone)
            status_data = scheduler.get_zone_status_summary()
            return Response(status_data)
        except Exception as e:
            return Response(
                {'error': f'Failed to calculate zone status: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def calculate(self, request, pk=None):
        """Calculate and optionally save a new irrigation recommendation."""
        zone = self.get_object()
        from .services.irrigation_scheduler import IrrigationScheduler

        save_recommendation = request.data.get('save', False)
        as_of_date = request.data.get('as_of_date')

        try:
            if as_of_date:
                as_of_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
            else:
                as_of_date = date.today()

            scheduler = IrrigationScheduler(zone)
            calculation = scheduler.calculate_recommendation(as_of_date)

            result = {
                'calculation': calculation,
                'saved': False,
            }

            if save_recommendation and calculation['recommended']:
                recommendation = scheduler.create_recommendation_record(calculation)
                result['recommendation_id'] = recommendation.id
                result['saved'] = True

            return Response(result)
        except Exception as e:
            return Response(
                {'error': f'Failed to calculate recommendation: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get', 'post'])
    def events(self, request, pk=None):
        """GET: List irrigation events, POST: Record new event."""
        zone = self.get_object()

        if request.method == 'GET':
            events = zone.irrigation_events.order_by('-date')[:20]
            serializer = IrrigationZoneEventSerializer(events, many=True)
            return Response(serializer.data)

        elif request.method == 'POST':
            data = request.data.copy()
            data['zone'] = zone.id
            serializer = IrrigationZoneEventCreateSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def recommendations(self, request, pk=None):
        """Get recommendations history for a zone."""
        zone = self.get_object()
        recommendations = zone.recommendations.order_by('-created_at')[:20]
        serializer = IrrigationRecommendationListSerializer(recommendations, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def weather(self, request, pk=None):
        """Get recent weather data for the zone's CIMIS target."""
        zone = self.get_object()

        if not zone.cimis_target:
            return Response(
                {'error': 'No CIMIS target configured for this zone'},
                status=status.HTTP_400_BAD_REQUEST
            )

        days = int(request.query_params.get('days', 7))
        from .services.cimis_service import CIMISService

        try:
            cimis = CIMISService()
            data = cimis.get_recent_data(
                zone.cimis_target,
                days=days,
                target_type=zone.cimis_target_type or 'station'
            )
            serializer = CIMISDataSerializer(
                [CIMISDataCache(**d) for d in data],
                many=True
            )
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': f'Failed to fetch weather data: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class IrrigationRecommendationViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing irrigation recommendations.

    RLS NOTES:
    - Recommendations inherit company from Zone -> Field -> Farm
    """
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['recommended_date', 'created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return IrrigationRecommendationListSerializer
        return IrrigationRecommendationSerializer

    def get_queryset(self):
        """Filter recommendations by current user's company."""
        queryset = IrrigationRecommendation.objects.select_related(
            'zone', 'zone__field', 'zone__field__farm'
        )
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(zone__field__farm__company=company)

        # Optional filters
        zone_id = self.request.query_params.get('zone')
        if zone_id:
            queryset = queryset.filter(zone_id=zone_id)

        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset.order_by('-created_at')

    @action(detail=True, methods=['post'])
    def apply(self, request, pk=None):
        """Mark recommendation as applied and record the irrigation event."""
        recommendation = self.get_object()

        if recommendation.status != 'pending':
            return Response(
                {'error': 'This recommendation has already been processed'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create irrigation event
        event_data = {
            'zone': recommendation.zone.id,
            'date': request.data.get('date', date.today().isoformat()),
            'depth_inches': request.data.get('depth_inches', recommendation.recommended_depth_inches),
            'duration_hours': request.data.get('duration_hours', recommendation.recommended_duration_hours),
            'method': request.data.get('method', 'scheduled'),
            'source': 'recommendation',
            'notes': request.data.get('notes', f'Applied recommendation #{recommendation.id}'),
        }

        serializer = IrrigationZoneEventCreateSerializer(data=event_data)
        serializer.is_valid(raise_exception=True)
        event = serializer.save()

        # Update recommendation status
        recommendation.status = 'applied'
        recommendation.save()

        return Response({
            'recommendation': IrrigationRecommendationSerializer(recommendation).data,
            'event': IrrigationZoneEventSerializer(event).data,
        })

    @action(detail=True, methods=['post'])
    def skip(self, request, pk=None):
        """Mark recommendation as skipped."""
        recommendation = self.get_object()

        if recommendation.status != 'pending':
            return Response(
                {'error': 'This recommendation has already been processed'},
                status=status.HTTP_400_BAD_REQUEST
            )

        recommendation.status = 'skipped'
        recommendation.save()

        return Response(IrrigationRecommendationSerializer(recommendation).data)


class CropCoefficientProfileViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for managing crop coefficient (Kc) profiles.

    System default profiles (zone=null) are read-only.
    Zone-specific profiles can be created/edited.
    """
    serializer_class = CropCoefficientProfileSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.SearchFilter]
    search_fields = ['crop_type', 'growth_stage']

    def get_queryset(self):
        """Get Kc profiles - system defaults and company-specific."""
        company = get_user_company(self.request.user)

        # Get system defaults (no zone)
        system_defaults = Q(zone__isnull=True)

        # Get company-specific profiles
        if company:
            company_profiles = Q(zone__field__farm__company=company)
            return CropCoefficientProfile.objects.filter(
                system_defaults | company_profiles
            ).select_related('zone', 'zone__field')
        else:
            return CropCoefficientProfile.objects.filter(system_defaults)

    def perform_create(self, serializer):
        """Ensure zone belongs to user's company."""
        zone = serializer.validated_data.get('zone')
        if zone:
            company = require_company(self.request.user)
            if zone.field.farm.company != company:
                raise serializers.ValidationError({
                    'zone': 'You can only create profiles for zones in your company.'
                })
        serializer.save()

    def perform_update(self, serializer):
        """Prevent editing system defaults."""
        if serializer.instance.zone is None:
            raise serializers.ValidationError(
                'System default profiles cannot be modified.'
            )
        serializer.save()

    def perform_destroy(self, instance):
        """Prevent deleting system defaults."""
        if instance.zone is None:
            raise serializers.ValidationError(
                'System default profiles cannot be deleted.'
            )
        instance.delete()


class SoilMoistureReadingViewSet(AuditLogMixin, viewsets.ModelViewSet):
    """
    API endpoint for soil moisture sensor readings.
    """
    serializer_class = SoilMoistureReadingSerializer
    permission_classes = [IsAuthenticated, HasCompanyAccess]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['reading_datetime']

    def get_queryset(self):
        """Filter readings by current user's company."""
        queryset = SoilMoistureReading.objects.select_related(
            'zone', 'zone__field', 'zone__field__farm'
        )
        company = get_user_company(self.request.user)
        if company:
            queryset = queryset.filter(zone__field__farm__company=company)

        zone_id = self.request.query_params.get('zone')
        if zone_id:
            queryset = queryset.filter(zone_id=zone_id)

        return queryset.order_by('-reading_datetime')


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def irrigation_dashboard(request):
    """
    Get irrigation dashboard summary data.

    Returns aggregate statistics, zone statuses, and pending recommendations.
    """
    company = get_user_company(request.user)
    if not company:
        return Response(
            {'error': 'No company associated with user'},
            status=status.HTTP_400_BAD_REQUEST
        )

    from .services.irrigation_scheduler import IrrigationScheduler

    # Get active zones
    zones = IrrigationZone.objects.filter(
        field__farm__company=company,
        active=True
    ).select_related('field', 'field__farm', 'water_source')

    total_zones = zones.count()
    total_acres = zones.aggregate(total=Sum('acres'))['total'] or Decimal('0')

    # Calculate status for each zone
    zone_statuses = []
    zones_needing = 0
    zones_soon = 0
    total_depletion = Decimal('0')
    zones_with_depletion = 0

    for zone in zones:
        try:
            scheduler = IrrigationScheduler(zone)
            zone_status = scheduler.get_zone_status_summary()
            zone_statuses.append(zone_status)

            if zone_status.get('status') == 'needs_irrigation':
                zones_needing += 1
            elif zone_status.get('status') == 'irrigation_soon':
                zones_soon += 1

            depletion = zone_status.get('depletion_pct')
            if depletion is not None:
                total_depletion += Decimal(str(depletion))
                zones_with_depletion += 1
        except Exception as e:
            zone_statuses.append({
                'zone_id': zone.id,
                'zone_name': zone.name,
                'status': 'error',
                'error': str(e)
            })

    avg_depletion = (total_depletion / zones_with_depletion) if zones_with_depletion > 0 else Decimal('0')

    # Get pending recommendations
    pending_recs = IrrigationRecommendation.objects.filter(
        zone__field__farm__company=company,
        status='pending'
    ).select_related('zone').order_by('recommended_date')[:10]

    # Get recent irrigation events
    recent_events = IrrigationEvent.objects.filter(
        zone__field__farm__company=company
    ).select_related('zone', 'zone__field').order_by('-date')[:10]

    # Try to get recent weather data (from first zone with CIMIS target)
    recent_eto = None
    recent_rain = None
    zone_with_cimis = zones.exclude(cimis_target__isnull=True).exclude(cimis_target='').first()
    if zone_with_cimis:
        try:
            from .services.cimis_service import CIMISService
            cimis = CIMISService()
            weather_data = cimis.get_recent_data(
                zone_with_cimis.cimis_target,
                days=7,
                target_type=zone_with_cimis.cimis_target_type or 'station'
            )
            for d in weather_data:
                if d.get('eto'):
                    recent_eto = (recent_eto or Decimal('0')) + d['eto']
                if d.get('precipitation'):
                    recent_rain = (recent_rain or Decimal('0')) + d['precipitation']
        except Exception:
            pass

    # Organize zones by status
    zones_by_status = {
        'needs_irrigation': [z for z in zone_statuses if z.get('status') == 'needs_irrigation'],
        'irrigation_soon': [z for z in zone_statuses if z.get('status') == 'irrigation_soon'],
        'ok': [z for z in zone_statuses if z.get('status') == 'ok'],
        'error': [z for z in zone_statuses if z.get('status') == 'error'],
    }

    return Response({
        'total_zones': total_zones,
        'active_zones': total_zones,  # Already filtered to active
        'zones_needing_irrigation': zones_needing,
        'zones_irrigation_soon': zones_soon,
        'total_acres': float(total_acres),
        'avg_depletion_pct': float(round(avg_depletion, 1)),
        'recent_eto_total': float(recent_eto) if recent_eto else None,
        'recent_rainfall_total': float(recent_rain) if recent_rain else None,
        'zones': IrrigationZoneListSerializer(zones, many=True).data,
        'zones_by_status': zones_by_status,
        'pending_recommendations': IrrigationRecommendationListSerializer(pending_recs, many=True).data,
        'recent_events': IrrigationZoneEventSerializer(recent_events, many=True).data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def cimis_stations(request):
    """
    Get list of nearby CIMIS stations.

    Query Parameters:
        lat: Latitude
        lng: Longitude
        limit: Max number of stations (default 5)
    """
    lat = request.query_params.get('lat')
    lng = request.query_params.get('lng')
    limit = int(request.query_params.get('limit', 5))

    if not lat or not lng:
        return Response(
            {'error': 'lat and lng parameters are required'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # For now, return a static list of common California CIMIS stations
    # In production, this could query the CIMIS station list API
    common_stations = [
        {'id': '2', 'name': 'Five Points', 'county': 'Fresno'},
        {'id': '5', 'name': 'Shafter', 'county': 'Kern'},
        {'id': '6', 'name': 'Bishop', 'county': 'Inyo'},
        {'id': '7', 'name': 'Firebaugh', 'county': 'Fresno'},
        {'id': '15', 'name': 'Stratford', 'county': 'Kings'},
        {'id': '39', 'name': 'Parlier', 'county': 'Fresno'},
        {'id': '54', 'name': 'Arvin/Edison', 'county': 'Kern'},
        {'id': '56', 'name': 'Castroville', 'county': 'Monterey'},
        {'id': '80', 'name': 'Fresno State', 'county': 'Fresno'},
        {'id': '105', 'name': 'Lindcove', 'county': 'Tulare'},
        {'id': '106', 'name': 'Meloland', 'county': 'Imperial'},
        {'id': '116', 'name': 'Salinas South', 'county': 'Monterey'},
        {'id': '125', 'name': 'Twitchell Island', 'county': 'Sacramento'},
        {'id': '131', 'name': 'Ripon', 'county': 'San Joaquin'},
        {'id': '170', 'name': 'Goleta Foothills', 'county': 'Santa Barbara'},
    ]

    return Response({
        'stations': common_stations[:limit],
        'note': 'Use station ID as cimis_target with target_type="station"',
    })


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, HasCompanyAccess])
def load_water_data_api(request):
    """
    API endpoint to load water wells and readings from the fixtures file.

    GET: Returns info about what will be imported (dry run)
    POST: Actually performs the import

    Query Parameters:
        farm_id: Farm ID to assign wells to (defaults to first farm)
    """
    import json
    import os
    from decimal import Decimal as D

    # Get farm
    farm_id = request.query_params.get('farm_id')
    if farm_id:
        try:
            farm = Farm.objects.get(id=int(farm_id))
        except Farm.DoesNotExist:
            return Response({'error': f'Farm {farm_id} not found'}, status=400)
    else:
        farm = Farm.objects.first()
        if not farm:
            return Response({'error': 'No farms exist'}, status=400)

    # Load fixture file
    fixture_path = os.path.join(os.path.dirname(__file__), '..', 'fixtures', 'water_data_export.json')
    if not os.path.exists(fixture_path):
        return Response({'error': 'Fixture file not found', 'path': fixture_path}, status=404)

    with open(fixture_path, 'r') as f:
        data = json.load(f)

    water_sources = [item for item in data if item.get('model') == 'api.watersource']
    well_readings = [item for item in data if item.get('model') == 'api.wellreading']

    # Check existing
    existing_wells = set(WaterSource.objects.filter(
        source_type='well',
        state_well_number__isnull=False
    ).exclude(state_well_number='').values_list('state_well_number', flat=True))

    new_sources = [s for s in water_sources if s['fields'].get('state_well_number') not in existing_wells]

    if request.method == 'GET':
        return Response({
            'farm': {'id': farm.id, 'name': farm.name},
            'fixture_file': fixture_path,
            'total_sources_in_file': len(water_sources),
            'total_readings_in_file': len(well_readings),
            'existing_wells_in_db': len(existing_wells),
            'new_wells_to_create': len(new_sources),
            'message': 'Use POST to actually import the data'
        })

    # POST - do the import
    source_pk_map = {}
    sources_created = 0
    sources_skipped = 0

    for item in water_sources:
        old_pk = item['pk']
        fields = item['fields']
        state_well = fields.get('state_well_number', '')

        # Check if exists
        existing = WaterSource.objects.filter(state_well_number=state_well).first() if state_well else None
        if existing:
            source_pk_map[old_pk] = existing.pk
            sources_skipped += 1
            continue

        try:
            ws = WaterSource(
                farm=farm,
                name=fields.get('name', ''),
                source_type='well',
                well_name=fields.get('well_name', ''),
                state_well_number=state_well,
                gsa=fields.get('gsa', ''),
                owner_code=fields.get('owner_code', ''),
                base_extraction_rate=D(fields['base_extraction_rate']) if fields.get('base_extraction_rate') else None,
                gsp_rate=D(fields['gsp_rate']) if fields.get('gsp_rate') else None,
                domestic_rate=D(fields['domestic_rate']) if fields.get('domestic_rate') else None,
                fixed_quarterly_fee=D(fields['fixed_quarterly_fee']) if fields.get('fixed_quarterly_fee') else None,
                is_domestic_well=fields.get('is_domestic_well', False),
                has_flowmeter=fields.get('has_flowmeter', True),
                flowmeter_units=fields.get('flowmeter_units', 'acre_feet'),
                flowmeter_multiplier=D(fields.get('flowmeter_multiplier', '1.0')),
                well_status=fields.get('well_status', 'active'),
                active=fields.get('active', True),
                used_for_irrigation=fields.get('used_for_irrigation', True),
                notes=fields.get('notes', ''),
            )
            ws.save()
            source_pk_map[old_pk] = ws.pk
            sources_created += 1
        except Exception as e:
            return Response({'error': f'Failed to create well: {e}'}, status=500)

    # Import readings
    readings_created = 0
    readings_skipped = 0

    for item in well_readings:
        fields = item['fields']
        old_source_pk = fields.get('water_source')

        if old_source_pk not in source_pk_map:
            readings_skipped += 1
            continue

        new_source_pk = source_pk_map[old_source_pk]
        reading_date = fields.get('reading_date')

        # Check duplicate
        if WellReading.objects.filter(water_source_id=new_source_pk, reading_date=reading_date).exists():
            readings_skipped += 1
            continue

        # Skip if no meter reading
        if not fields.get('meter_reading'):
            readings_skipped += 1
            continue

        try:
            wr = WellReading(
                water_source_id=new_source_pk,
                reading_date=reading_date,
                meter_reading=D(fields['meter_reading']),
                reading_type=fields.get('reading_type', 'manual'),
                extraction_acre_feet=D(fields['extraction_acre_feet']) if fields.get('extraction_acre_feet') else None,
                notes=fields.get('notes', ''),
            )
            wr.save()
            readings_created += 1
        except Exception as e:
            readings_skipped += 1

    return Response({
        'success': True,
        'farm': {'id': farm.id, 'name': farm.name},
        'wells_created': sources_created,
        'wells_skipped': sources_skipped,
        'readings_created': readings_created,
        'readings_skipped': readings_skipped,
    })







































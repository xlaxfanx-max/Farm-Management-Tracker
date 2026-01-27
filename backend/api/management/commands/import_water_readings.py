"""
Management command to import water meter readings from Excel files.

Supports UWCD and OBGMA pivot-style formats where:
- Wells are columns
- Dates are rows
- Each date block has: Meter Reading, Usage, Multiplier, Acre-Ft Extracted, etc.

Usage:
    python manage.py import_water_readings --file="path/to/file.xlsx" --gsa=uwcd --dry-run
    python manage.py import_water_readings --file="path/to/file.xlsx" --gsa=obgma
"""

import os
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from api.models import WaterSource, WellReading


class Command(BaseCommand):
    help = 'Import water meter readings from Excel files (UWCD or OBGMA pivot format)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to the Excel file to import'
        )
        parser.add_argument(
            '--gsa',
            type=str,
            choices=['uwcd', 'obgma'],
            required=True,
            help='GSA format (uwcd or obgma)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without saving to database'
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            default=True,
            help='Skip readings that already exist (by date + well)'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        gsa = options['gsa']
        dry_run = options['dry_run']
        skip_existing = options['skip_existing']

        if not os.path.exists(file_path):
            raise CommandError(f'File not found: {file_path}')

        self.stdout.write(f'\n{"=" * 60}')
        self.stdout.write(f'WATER READINGS IMPORT - {gsa.upper()}')
        self.stdout.write(f'{"=" * 60}')
        self.stdout.write(f'File: {file_path}')
        self.stdout.write(f'Mode: {"DRY RUN (no changes)" if dry_run else "LIVE IMPORT"}')
        self.stdout.write(f'{"=" * 60}\n')

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f'Failed to open Excel file: {e}')

        # Parse pivot format
        readings = self.parse_pivot_format(wb, gsa)

        if not readings:
            self.stdout.write(self.style.WARNING('No readings found in file'))
            return

        # Match wells and validate
        self.stdout.write(f'\nFound {len(readings)} readings to process\n')

        matched, unmatched, skipped, errors = self.validate_readings(readings, skip_existing)

        # Report unmatched wells
        if unmatched:
            self.stdout.write(self.style.WARNING(f'\n[!] UNMATCHED WELLS ({len(unmatched)} readings):'))
            unmatched_wells = {}
            for r in unmatched:
                key = r['state_well_number']
                if key not in unmatched_wells:
                    unmatched_wells[key] = {'count': 0, 'name': r.get('well_description', '')}
                unmatched_wells[key]['count'] += 1

            for well_num, info in sorted(unmatched_wells.items()):
                self.stdout.write(f'   - {well_num} ({info["name"][:30]}) - {info["count"]} readings')
            self.stdout.write('')

        # Report skipped (already exist)
        if skipped:
            self.stdout.write(self.style.NOTICE(f'[i] SKIPPED (already exist): {len(skipped)} readings\n'))

        # Report errors
        if errors:
            self.stdout.write(self.style.ERROR(f'\n[X] ERRORS ({len(errors)}):'))
            for err in errors[:10]:
                self.stdout.write(f'   - {err}')
            if len(errors) > 10:
                self.stdout.write(f'   ... and {len(errors) - 10} more')
            self.stdout.write('')

        # Report what will be imported
        self.stdout.write(self.style.SUCCESS(f'\n[OK] READY TO IMPORT: {len(matched)} readings\n'))

        if matched:
            # Show preview
            self.stdout.write('Preview (first 15):')
            self.stdout.write('-' * 90)
            for r in matched[:15]:
                well_name = r.get('well_name', r.get('well_description', 'Unknown'))[:30]
                meter = f"{r['meter_reading']:.2f}" if r.get('meter_reading') else 'N/A'
                extract = f"{r['extraction_af']:.4f}" if r.get('extraction_af') else 'N/A'
                self.stdout.write(
                    f"  {r['reading_date']} | {well_name:30} | "
                    f"Meter: {meter:>12} | "
                    f"Extract: {extract:>10} AF"
                )
            if len(matched) > 15:
                self.stdout.write(f'  ... and {len(matched) - 15} more')
            self.stdout.write('-' * 90)

        if dry_run:
            self.stdout.write(self.style.WARNING('\n[*] DRY RUN - No changes made. Remove --dry-run to import.\n'))
            return

        if not matched:
            self.stdout.write(self.style.WARNING('\nNo readings to import.'))
            return

        # Confirm import
        self.stdout.write('')
        confirm = input(f'Import {len(matched)} readings? [y/N]: ')
        if confirm.lower() != 'y':
            self.stdout.write(self.style.WARNING('Import cancelled.'))
            return

        # Do the import
        imported_count = self.do_import(matched)
        self.stdout.write(self.style.SUCCESS(f'\n[OK] Successfully imported {imported_count} readings!\n'))

    def parse_pivot_format(self, wb, gsa):
        """
        Parse pivot-style Excel where:
        - Row 3 (or similar): State Well Numbers across columns
        - Row 4-6: Well descriptions, owner codes
        - Data rows: Date in col A, field type in col B, values in well columns
        """
        readings = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.stdout.write(f'Processing sheet: {sheet_name}')

            # Find the row with state well numbers (look for pattern like 04N22W...)
            # Pick the row with the MOST well numbers (not just the first with 2+)
            well_row = None
            well_columns = {}  # col_idx -> state_well_number
            best_count = 0

            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                row_wells = {}
                for col_idx, val in enumerate(row):
                    if val and self._looks_like_state_well(str(val)):
                        row_wells[col_idx] = str(val).strip()

                if len(row_wells) > best_count:
                    best_count = len(row_wells)
                    well_row = row_idx
                    well_columns = row_wells

            if not well_row:
                self.stdout.write(f'  - No well number row found, skipping sheet')
                continue

            self.stdout.write(f'  - Found {len(well_columns)} wells in row {well_row}')

            # Get well descriptions from rows below well numbers
            well_info = {col: {'state_well_number': num} for col, num in well_columns.items()}

            # Look at rows 4, 5, 6 for descriptions and owner codes
            for row_idx in range(well_row + 1, min(well_row + 4, 10)):
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                for col_idx in well_columns:
                    if col_idx < len(row) and row[col_idx]:
                        val = str(row[col_idx]).strip()
                        if 'description' not in well_info[col_idx]:
                            well_info[col_idx]['description'] = val
                        elif 'owner_code' not in well_info[col_idx]:
                            well_info[col_idx]['owner_code'] = val

            # Now parse data rows - look for dates in column A
            current_date = None
            current_data = {}  # col_idx -> {meter_reading, extraction_af, etc.}

            for row in sheet.iter_rows(min_row=well_row + 4, values_only=True):
                if not any(row):
                    continue

                col_a = row[0] if len(row) > 0 else None
                col_b = str(row[1]).lower().strip() if len(row) > 1 and row[1] else ''

                # Check if this is a date row (new reading block)
                parsed_date = self._parse_date(col_a)
                if parsed_date:
                    # Save previous block if exists
                    if current_date and current_data:
                        for col_idx, data in current_data.items():
                            if data.get('meter_reading') is not None or data.get('extraction_af') is not None:
                                info = well_info.get(col_idx, {})
                                readings.append({
                                    'state_well_number': info.get('state_well_number', ''),
                                    'well_description': info.get('description', ''),
                                    'owner_code': info.get('owner_code', ''),
                                    'reading_date': current_date,
                                    'meter_reading': data.get('meter_reading'),
                                    'extraction_af': data.get('extraction_af'),
                                    'usage_raw': data.get('usage_raw'),
                                    'multiplier': data.get('multiplier', Decimal('1')),
                                    'gsa': gsa,
                                    'source_sheet': sheet_name
                                })

                    current_date = parsed_date
                    current_data = {col: {} for col in well_columns}

                # Parse field type rows
                if current_date and col_b:
                    for col_idx in well_columns:
                        if col_idx < len(row):
                            val = self._parse_decimal(row[col_idx])
                            if val is not None:
                                if 'meter' in col_b and 'read' in col_b:
                                    current_data[col_idx]['meter_reading'] = val
                                elif 'usage' in col_b or 'period' in col_b:
                                    current_data[col_idx]['usage_raw'] = val
                                elif 'multiplier' in col_b or 'mult' in col_b:
                                    current_data[col_idx]['multiplier'] = val
                                elif 'acre' in col_b and ('ft' in col_b or 'feet' in col_b):
                                    current_data[col_idx]['extraction_af'] = val
                                elif col_b == 'af' or 'extracted' in col_b:
                                    current_data[col_idx]['extraction_af'] = val

            # Don't forget the last block
            if current_date and current_data:
                for col_idx, data in current_data.items():
                    if data.get('meter_reading') is not None or data.get('extraction_af') is not None:
                        info = well_info.get(col_idx, {})
                        readings.append({
                            'state_well_number': info.get('state_well_number', ''),
                            'well_description': info.get('description', ''),
                            'owner_code': info.get('owner_code', ''),
                            'reading_date': current_date,
                            'meter_reading': data.get('meter_reading'),
                            'extraction_af': data.get('extraction_af'),
                            'usage_raw': data.get('usage_raw'),
                            'multiplier': data.get('multiplier', Decimal('1')),
                            'gsa': gsa,
                            'source_sheet': sheet_name
                        })

            self.stdout.write(f'  - Parsed {len([r for r in readings if r["source_sheet"] == sheet_name])} readings from this sheet')

        return readings

    def _looks_like_state_well(self, val):
        """Check if value looks like a CA state well number (e.g., 04N22W06K12S)."""
        if not val or len(val) < 10:
            return False
        val = val.upper().strip()
        # Pattern: digits + N/S + digits + W/E + more chars
        import re
        pattern = r'^\d{2}[NS]\d{2}[WE]\d{2}[A-Z]\d{2}[A-Z]?$'
        return bool(re.match(pattern, val))

    def _parse_date(self, val):
        """Parse various date formats."""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val

        val_str = str(val).strip()
        if not val_str:
            return None

        # Skip if it looks like a year total row
        if 'total' in val_str.lower():
            return None

        formats = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d-%b-%Y', '%d-%b-%y']
        for fmt in formats:
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_decimal(self, val):
        """Parse a decimal value."""
        if val is None:
            return None
        try:
            if isinstance(val, (int, float)):
                return Decimal(str(val))
            val_str = str(val).strip().replace(',', '')
            if not val_str or val_str == '-' or val_str.lower() == 'none':
                return None
            return Decimal(val_str)
        except (InvalidOperation, ValueError):
            return None

    def validate_readings(self, readings, skip_existing):
        """Validate readings and match to wells in database."""
        matched = []
        unmatched = []
        skipped = []
        errors = []

        # Build cache of wells by state well number
        wells_cache = {}
        for ws in WaterSource.objects.filter(source_type='well'):
            if ws.state_well_number:
                key = ws.state_well_number.upper().strip()
                wells_cache[key] = ws

        self.stdout.write(f'Found {len(wells_cache)} wells in database with state well numbers')

        # Build cache of existing readings
        existing_readings = set()
        if skip_existing:
            for wr in WellReading.objects.values_list('water_source_id', 'reading_date'):
                existing_readings.add((wr[0], wr[1]))

        for r in readings:
            if not r.get('state_well_number'):
                continue

            state_well = r['state_well_number'].upper().strip()

            # Try to find matching well
            well = wells_cache.get(state_well)

            if not well:
                unmatched.append(r)
                continue

            r['water_source'] = well
            r['well_name'] = well.name

            # Check if already exists
            if skip_existing and (well.id, r['reading_date']) in existing_readings:
                skipped.append(r)
                continue

            # Validate data
            if r.get('meter_reading') is None and r.get('extraction_af') is None:
                errors.append(f"{state_well} on {r['reading_date']}: No meter reading or extraction value")
                continue

            matched.append(r)

        return matched, unmatched, skipped, errors

    def do_import(self, readings):
        """Import validated readings to database."""
        imported = 0

        with transaction.atomic():
            for r in readings:
                try:
                    well = r['water_source']

                    reading = WellReading(
                        water_source=well,
                        reading_date=r['reading_date'],
                        meter_reading=r.get('meter_reading'),
                        reading_type='manual',
                        notes=f"Imported from {r['gsa'].upper()} spreadsheet ({r['source_sheet']})"
                    )

                    # Store pre-calculated extraction if available
                    if r.get('extraction_af') is not None:
                        reading.extraction_acre_feet = r['extraction_af']

                    reading.save()
                    imported += 1

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error importing {r['state_well_number']} {r['reading_date']}: {e}"))

        return imported

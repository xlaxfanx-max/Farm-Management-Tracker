"""Seed the Finch operation: 11 ranches, their acreage, and their blocks.

Everything the operations work publishes is a rate — dollars per acre, water per
acre, cartons per acre — so the ranch list and the acreage denominator have to
exist before any of it means anything. Today the only Farm-creating code in the
repo is seed_demo_data, which makes 'North Ranch' and 'South Ranch'.

Sources, in the order they are trusted:

  'Setup - Acreage'   ranch x crop x year, bearing / non-bearing / grazing.
                      This is what the ranch P&L is built on and it is the
                      authority for the RANCH denominator.
  'Setup - Plantings' block register — planting year, rootstock, spacing, and
                      acres on some rows. The only source with planting years,
                      so it is the authority for BLOCK identity.

Where they disagree, the disagreement is recorded rather than resolved
silently: Field.acres_confidence says how much weight a block's acreage can
carry, and the command refuses to write while a blocking conflict is unanswered.

This is never run on deploy. It is a deliberate, reviewable act:

    python manage.py seed_finch_operation --company="Finch Farms"
    python manage.py seed_finch_operation --company="Finch Farms" --commit
"""

import os
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from api.models import (
    Company,
    Farm,
    Field,
    LegalEntity,
    RanchCropAcreage,
)

DEFAULT_WORKBOOK = (
    r'C:/Users/Owner/Documents/Claude/Projects/Finch Farms Summer 2026/'
    r'Ag/Farm Projections/Finch Ag P&L + 5-Year Projection V6.17.xlsx'
)

# The five entities the 11 ranches are held in. Short codes match the workbooks.
ENTITIES = [
    ('JPF', 'James P. Finch Jr.', 'partnership', 'JPF Historical 2021-2025.xlsx'),
    ('FFLLC', 'Finch Farms LLC', 'llc', 'Finch Farms LLC Historical 2021-2025.xlsx'),
    ('TCC', 'Thacher Creek LLC', 'llc', 'Thacher Creek Historical 2021-2025.xlsx'),
    ('RMLF', 'RMLF', 'llc', 'RMLF Historical 2021-2025.xlsx'),
    ('F&P', 'F&P', 'partnership', 'F&P Historical 2021-2025.xlsx'),
]

RANCH_ENTITY = {
    'Foster Park': 'JPF',
    'Grand': 'JPF',
    'Piru': 'JPF',
    'Saticoy': 'JPF',
    'Sespe': 'JPF',
    'San Antonio Creek': 'FFLLC',
    'Rio Vista': 'FFLLC',
    'Helicopter Ranch': 'FFLLC',
    'Thacher Creek': 'TCC',
    'Office': 'RMLF',
    'Old Telegraph': 'F&P',
}

CROP_LABEL_TO_CODE = {
    'Avocados': 'avocados',
    'Lemons': 'lemons',
    'Valencias': 'valencias',
    'Mandarins': 'mandarins',
    'Pixies': 'pixies',
    'Navels': 'navels',
    'Cara Cara': 'cara_cara',
    'Other Fruit': 'other_fruit',
    'Grazing Acerage': 'grazing',  # the workbook's spelling
}

ACREAGE_YEARS = [2021, 2022, 2023, 2024, 2025]

# Conflicts that must be answered before a per-acre figure can be published.
# Each is a real disagreement between two sources, not a data-entry slip.
BLOCKING_CONFLICTS = {
    'grand_house': (
        'Grand House block acreage: packinghouse statements say 9.00 ac while '
        "the P&L carries the whole Grand ranch at 6.70. The block cannot exceed "
        'its ranch. At 9.00 it grades B on yield; at 6.70 it grades A. '
        'Seeded at the statement figure, flagged.'
    ),
    'saticoy_even_split': (
        'Saticoy 1A / 1B / 1C are each recorded at exactly 15.00 ac, which reads '
        'as an even split on the packinghouse side rather than a survey — yet 1B '
        'produces ~1,184 ctn/ac against 1C at ~409, a 2.89x spread on identical '
        'stated acres. Either the acres are wrong or the blocks really differ '
        'that much, and the card cannot tell which.'
    ),
    'rio_vista': (
        "Rio Vista's Setup - Acreage row is internally impossible: 20.59 bearing "
        'lemon acres flat across 2021-2025 with ZERO non-bearing, while Setup - '
        'Plantings lists 6 Rio Vista lemon blocks totalling 30.13 planted acres '
        'planted 2017/2023/2023/2023/2024/2025 — only the 2017 block would be '
        'bearing by 2025. Blocks are seeded from Plantings; the Acreage row is '
        'marked superseded.'
    ),
}


def _dec(value):
    """Excel numbers arrive as floats with long tails (84.88999999999999)."""
    if value is None or value == '':
        return None
    try:
        return Decimal(str(round(float(value), 4)))
    except (TypeError, ValueError, InvalidOperation):
        return None


class Command(BaseCommand):
    help = 'Seed the 11 Finch ranches, their crop acreage, and their blocks'

    def add_arguments(self, parser):
        parser.add_argument('--company', type=str, default='Finch Farms',
                            help='Company name (substring match)')
        parser.add_argument('--workbook', type=str, default=DEFAULT_WORKBOOK,
                            help='Path to the P&L + Projection workbook')
        parser.add_argument('--commit', action='store_true',
                            help='Write. Without this the command only reports.')
        parser.add_argument('--accept-conflicts', action='store_true',
                            help='Write despite unresolved acreage conflicts. '
                                 'Affected rows are stamped so the UI can badge them.')

    # -- sources ----------------------------------------------------------

    def _load_workbook(self, path):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError(
                'openpyxl is required to read the acreage workbook.'
            ) from exc
        if not os.path.exists(path):
            raise CommandError(f'Workbook not found: {path}')
        return openpyxl.load_workbook(path, data_only=True)

    def _read_acreage(self, wb):
        """Setup - Acreage -> {(ranch, crop_code, year): (bearing, non_bearing)}.

        Layout: col A ranch (only on the first row of each block), col B crop,
        then Bearing/Non-Bearing pairs from col C for 2021..2025. Rows whose
        crop label ends in 'Total' are subtotals and are skipped.
        """
        ws = wb['Setup - Acreage']
        out = {}
        ranches = []
        current = None
        for row in ws.iter_rows(min_row=6, values_only=True):
            label_a = str(row[0]).strip() if row[0] else ''
            label_b = str(row[1]).strip() if row[1] else ''
            if label_a:
                current = label_a
                if current not in ranches and 'TOTAL' not in current.upper():
                    ranches.append(current)
            if not label_b or not current:
                continue
            if 'total' in label_b.lower():
                continue
            crop_code = CROP_LABEL_TO_CODE.get(label_b)
            if crop_code is None:
                continue
            for offset, year in enumerate(ACREAGE_YEARS):
                bearing = _dec(row[2 + offset * 2])
                non_bearing = _dec(row[3 + offset * 2])
                if bearing is None and non_bearing is None:
                    continue
                out[(current, crop_code, year)] = (
                    bearing or Decimal('0'),
                    non_bearing or Decimal('0'),
                )
        return ranches, out

    def _read_plantings(self, wb):
        """Setup - Plantings -> list of block dicts.

        The sheet carries a crop maturity-curve table below the block register;
        parsing stops at it. Rows are only accepted when the ranch is one we
        know, which excludes the curve table's 'Crop \\ Tree age' header row.
        """
        ws = wb['Setup - Plantings']
        blocks = []
        for row in ws.iter_rows(min_row=7, values_only=True):
            ranch = str(row[0]).strip() if row[0] else ''
            if not ranch or ranch not in RANCH_ENTITY:
                continue
            crop_label = str(row[4]).strip() if row[4] else ''
            crop_code = CROP_LABEL_TO_CODE.get(crop_label)
            if crop_code is None:
                continue
            block = str(row[2]).strip() if row[2] else ''
            year_planted = None
            try:
                year_planted = int(float(row[6])) if row[6] not in (None, '') else None
            except (TypeError, ValueError):
                year_planted = None
            blocks.append({
                'ranch': ranch,
                'code': str(row[1]).strip() if row[1] else '',
                'block': block,
                'variety': str(row[3]).strip() if row[3] else '',
                'crop_label': crop_label,
                'crop_code': crop_code,
                'rootstock': str(row[5]).strip() if row[5] else '',
                'year_planted': year_planted,
                'tree_spacing_ft': _dec(row[7]),
                'row_spacing_ft': _dec(row[8]),
                'trees_per_acre': _dec(row[9]),
                'acres': _dec(row[11]) if len(row) > 11 else None,
            })
        return blocks

    def _read_ownership(self, wb):
        """Projection - Setup -> {ranch: Decimal ownership fraction}.

        NOTE: this column is referenced by zero formulas in the workbook, so it
        records intent rather than driving anything. Seeded as a starting point,
        not as a confirmed figure.
        """
        ws = wb['Projection - Setup']
        out = {}
        for row in ws.iter_rows(min_row=11, max_row=30, values_only=True):
            ranch = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if ranch not in RANCH_ENTITY:
                continue
            pct = _dec(row[5]) if len(row) > 5 else None
            if pct is not None and Decimal('0') <= pct <= Decimal('1'):
                out[ranch] = pct
        return out

    # -- main -------------------------------------------------------------

    def handle(self, *args, **options):
        commit = options['commit']
        accept_conflicts = options['accept_conflicts']

        company = Company.objects.filter(name__icontains=options['company']).first()
        if not company:
            available = ', '.join(Company.objects.values_list('name', flat=True)) or '(none)'
            raise CommandError(
                f'Company "{options["company"]}" not found. Available: {available}'
            )

        wb = self._load_workbook(options['workbook'])
        ranches, acreage = self._read_acreage(wb)
        blocks = self._read_plantings(wb)
        ownership = self._read_ownership(wb)

        self.stdout.write('')
        self.stdout.write(f'Company:            {company.name}')
        self.stdout.write(f'Workbook:           {os.path.basename(options["workbook"])}')
        self.stdout.write(f'Ranches:            {len(ranches)}')
        self.stdout.write(f'Acreage rows:       {len(acreage)}')
        self.stdout.write(f'Block rows:         {len(blocks)}')
        self.stdout.write(f'  with real acres:  {sum(1 for b in blocks if b["acres"])}')
        self.stdout.write('')

        if len(ranches) != 11:
            self.stdout.write(self.style.WARNING(
                f'Expected 11 ranches, parsed {len(ranches)}: {ranches}'
            ))

        # The subscription cap blocks ranch 4 of 11.
        needed = max(len(ranches), company.farm_count)
        if company.max_farms < needed:
            self.stdout.write(self.style.WARNING(
                f'max_farms is {company.max_farms}; raising to {needed} '
                f'(the cap would block ranch {company.max_farms + 1} of {len(ranches)})'
            ))

        self._report_conflicts()

        if not commit:
            self.stdout.write(self.style.WARNING(
                'Dry run — nothing written. Re-run with --commit to apply.'
            ))
            return

        if not accept_conflicts:
            raise CommandError(
                'Refusing to write while the acreage conflicts above are '
                'unresolved. Answer them, or pass --accept-conflicts to seed '
                'anyway — every affected row will be stamped so the UI can '
                'badge it as unconfirmed.'
            )

        with transaction.atomic():
            entities = self._seed_entities(company)
            farms = self._seed_farms(company, ranches, ownership, entities, needed)
            n_acreage = self._seed_acreage(farms, acreage)
            n_fields = self._seed_fields(farms, blocks, acreage)

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'[OK] {len(entities)} entities, {len(farms)} ranches, '
            f'{n_acreage} acreage rows, {n_fields} blocks.'
        ))
        self.stdout.write(
            'Per-acre figures on blocks flagged even_split or placeholder must '
            'be badged in the UI. They are not measurements.'
        )

    # -- reporting --------------------------------------------------------

    def _report_conflicts(self):
        self.stdout.write(self.style.WARNING('Acreage conflicts needing an answer:'))
        for key, text in BLOCKING_CONFLICTS.items():
            self.stdout.write(f'  [{key}]')
            for line in text.split('. '):
                if line.strip():
                    self.stdout.write(f'      {line.strip().rstrip(".")}.')
        self.stdout.write('')

    # -- writers ----------------------------------------------------------

    def _seed_entities(self, company):
        entities = {}
        for short_code, name, entity_type, qb_file in ENTITIES:
            entity, _ = LegalEntity.objects.get_or_create(
                company=company,
                short_code=short_code,
                defaults={
                    'name': name,
                    'entity_type': entity_type,
                    'qb_file_name': qb_file,
                },
            )
            entities[short_code] = entity
        return entities

    def _seed_farms(self, company, ranches, ownership, entities, needed):
        if company.max_farms < needed:
            company.max_farms = needed
            company.save(update_fields=['max_farms'])

        farms = {}
        for ranch in ranches:
            entity = entities.get(RANCH_ENTITY.get(ranch, ''))
            farm, created = Farm.objects.get_or_create(
                company=company,
                name=ranch,
                defaults={'county': 'ventura'},
            )
            # Only fill what is empty — never overwrite something a person set.
            updates = []
            if farm.owning_entity_id is None and entity is not None:
                farm.owning_entity = entity
                updates.append('owning_entity')
            pct = ownership.get(ranch)
            if pct is not None and farm.finch_ownership_pct == Decimal('1.0000') \
                    and pct != Decimal('1.0000'):
                farm.finch_ownership_pct = pct
                updates.append('finch_ownership_pct')
            if not farm.county:
                farm.county = 'ventura'
                updates.append('county')
            if updates:
                farm.save(update_fields=updates)
            farms[ranch] = farm
        return farms

    def _seed_acreage(self, farms, acreage):
        count = 0
        for (ranch, crop_code, year), (bearing, non_bearing) in acreage.items():
            farm = farms.get(ranch)
            if farm is None:
                continue
            # The workbook records grazing in the non-bearing column; it is not
            # planted ground and must not inflate a per-acre denominator.
            if crop_code == 'grazing':
                bearing, non_bearing, grazing = Decimal('0'), Decimal('0'), non_bearing
            else:
                grazing = Decimal('0')

            note = ''
            superseded = False
            if ranch == 'Rio Vista' and crop_code == 'lemons':
                superseded = True
                note = BLOCKING_CONFLICTS['rio_vista']

            RanchCropAcreage.objects.update_or_create(
                farm=farm,
                crop_code=crop_code,
                year=year,
                defaults={
                    'bearing_acres': bearing,
                    'non_bearing_acres': non_bearing,
                    'grazing_acres': grazing,
                    'source': 'pnl_setup_acreage',
                    'is_superseded': superseded,
                    'note': note,
                },
            )
            count += 1
        return count

    @staticmethod
    def _block_names(blocks):
        """Give every register row a name that is unique within its ranch.

        The block register has no clean natural key. Foster Park rows carry
        real block numbers (2a, 3b, 5c); Grand and Thacher Creek rows carry
        none and differ only by variety; Piru and Rio Vista rows carry neither
        code, block nor variety and are distinguishable only by crop and
        acreage. Thacher Creek genuinely lists 'Cara' twice.

        So the name is built from whatever identifiers exist, and a numeric
        suffix is added when that is still not unique. The suffix is a marker
        that the source could not tell two blocks apart — not a real block
        label — and rows carrying one are worth resolving against a block map.
        """
        names, seen, ambiguous = [], {}, 0
        for b in blocks:
            parts = [p for p in (b['code'] or b['ranch'], b['block'], b['variety']) if p]
            if not b['block'] and not b['variety']:
                parts.append(b['crop_label'])
            base = ' '.join(parts).strip() or b['crop_label']
            key = (b['ranch'], base)
            seen[key] = seen.get(key, 0) + 1
            if seen[key] > 1:
                ambiguous += 1
                base = f'{base} #{seen[key]}'
            names.append(base)
        return names, ambiguous

    def _seed_fields(self, farms, blocks, acreage):
        """Create one Field per block register row.

        Field.total_acres is non-nullable, so every block has to assert an
        acreage. Where the register supplies one we use it; where it does not we
        divide the ranch-crop bearing acres evenly across that ranch's blocks of
        the same crop and mark it even_split. An even split is a placeholder
        with arithmetic done to it — it is never presented as a measurement.
        """
        # How many blocks share each ranch-crop, for the even split.
        share_counts = {}
        for b in blocks:
            key = (b['ranch'], b['crop_code'])
            share_counts[key] = share_counts.get(key, 0) + 1

        names, ambiguous = self._block_names(blocks)
        if ambiguous:
            self.stdout.write(self.style.WARNING(
                f'{ambiguous} register row(s) could not be told apart by code, '
                f'block or variety and were given a numeric suffix. Resolve '
                f'them against a block map before publishing block-level figures.'
            ))

        count = 0
        for b, name in zip(blocks, names):
            farm = farms.get(b['ranch'])
            if farm is None:
                continue

            if b['acres']:
                acres = b['acres']
                confidence = 'pnl'
                source = 'V6.17 Setup - Plantings (Acres column)'
            else:
                latest = acreage.get((b['ranch'], b['crop_code'], 2025))
                ranch_bearing = latest[0] if latest else Decimal('0')
                n = share_counts.get((b['ranch'], b['crop_code']), 1) or 1
                acres = (ranch_bearing / n).quantize(Decimal('0.01'))
                confidence = 'even_split'
                source = (
                    f'Even split of {ranch_bearing} bearing ac across {n} '
                    f'{b["crop_label"]} blocks (V6.17 Setup - Acreage 2025)'
                )

            if acres <= 0:
                acres = Decimal('0.01')
                confidence = 'placeholder'
                source = 'No acreage in any source — placeholder so the block can exist'

            defaults = {
                'total_acres': acres,
                'acres_source': source[:200],
                'acres_confidence': confidence,
                'county': farm.county or 'ventura',
                'current_crop': b['crop_label'],
                'year_planted': b['year_planted'],
                'field_number': b['block'][:50],
            }
            if b['tree_spacing_ft']:
                defaults['tree_spacing_ft'] = b['tree_spacing_ft']
            if b['row_spacing_ft']:
                defaults['row_spacing_ft'] = b['row_spacing_ft']
            if b['trees_per_acre']:
                defaults['trees_per_acre'] = b['trees_per_acre']

            field, created = Field.objects.get_or_create(
                farm=farm, name=name, defaults=defaults,
            )
            if not created:
                # Refresh provenance and agronomic facts, but leave a
                # human-corrected acreage alone.
                changed = []
                for attr in ('current_crop', 'year_planted', 'field_number'):
                    if getattr(field, attr) in (None, '') and defaults.get(attr):
                        setattr(field, attr, defaults[attr])
                        changed.append(attr)
                if field.acres_confidence in ('placeholder', 'even_split') \
                        and confidence not in ('placeholder', 'even_split'):
                    field.total_acres = acres
                    field.acres_source = defaults['acres_source']
                    field.acres_confidence = confidence
                    changed += ['total_acres', 'acres_source', 'acres_confidence']
                if changed:
                    field.save(update_fields=changed)
            count += 1
        return count

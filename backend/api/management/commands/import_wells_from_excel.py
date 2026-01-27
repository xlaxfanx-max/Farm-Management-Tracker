"""
Management command to auto-create wells from Excel water reading files.

This reads the UWCD/OBGMA pivot-style Excel files and extracts well information
(State Well Number, description, owner code) to create WaterSource records.

Usage:
    python manage.py import_wells_from_excel --file="path/to/file.xlsx" --gsa=uwcd --farm=1 --dry-run
    python manage.py import_wells_from_excel --file="path/to/file.xlsx" --gsa=obgma --farm=1
"""

import os
import re
from decimal import Decimal
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from api.models import WaterSource, Farm


# GSA Fee Defaults
GSA_FEE_DEFAULTS = {
    'uwcd': {
        'display_name': 'United Water Conservation District',
        'base_extraction_rate': Decimal('192.34'),
        'gsp_rate': None,
        'domestic_rate': Decimal('214.22'),
        'fixed_quarterly_fee': None,
    },
    'obgma': {
        'display_name': 'Ojai Basin GMA',
        'base_extraction_rate': Decimal('25.00'),
        'gsp_rate': Decimal('100.00'),
        'domestic_rate': None,
        'fixed_quarterly_fee': Decimal('70.00'),
    },
}


class Command(BaseCommand):
    help = 'Import wells from Excel water reading files (creates WaterSource records)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to the Excel file to read well info from'
        )
        parser.add_argument(
            '--gsa',
            type=str,
            choices=['uwcd', 'obgma'],
            required=True,
            help='GSA format (uwcd or obgma)'
        )
        parser.add_argument(
            '--farm',
            type=int,
            required=False,
            default=None,
            help='Farm ID to associate wells with (optional - will use first farm if not specified)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without saving to database'
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            default=True,
            help='Skip wells that already exist (by state well number)'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        gsa = options['gsa']
        farm_id = options['farm']
        dry_run = options['dry_run']
        skip_existing = options['skip_existing']

        if not os.path.exists(file_path):
            raise CommandError(f'File not found: {file_path}')

        # Get or create a farm to associate wells with
        if farm_id:
            try:
                farm = Farm.objects.get(id=farm_id)
            except Farm.DoesNotExist:
                raise CommandError(f'Farm with ID {farm_id} not found')
        else:
            # Use first available farm as placeholder (user will reassign later)
            farm = Farm.objects.first()
            if not farm:
                raise CommandError('No farms exist in database. Create a farm first.')

        self.stdout.write(f'\n{"=" * 60}')
        self.stdout.write(f'WELL IMPORT FROM EXCEL - {gsa.upper()}')
        self.stdout.write(f'{"=" * 60}')
        self.stdout.write(f'File: {file_path}')
        self.stdout.write(f'Farm: {farm.name} (ID: {farm.id}) - placeholder, reassign in UI later')
        self.stdout.write(f'Mode: {"DRY RUN (no changes)" if dry_run else "LIVE IMPORT"}')
        self.stdout.write(f'{"=" * 60}\n')

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f'Failed to open Excel file: {e}')

        # Extract well info from Excel
        wells = self.extract_wells_from_excel(wb, gsa)

        if not wells:
            self.stdout.write(self.style.WARNING('No wells found in file'))
            return

        self.stdout.write(f'\nFound {len(wells)} wells in Excel file\n')

        # Check for existing wells
        existing_wells = set()
        if skip_existing:
            for ws in WaterSource.objects.filter(source_type='well'):
                if ws.state_well_number:
                    existing_wells.add(ws.state_well_number.upper().strip())

        self.stdout.write(f'Found {len(existing_wells)} existing wells with state well numbers in database\n')

        # Categorize wells
        to_create = []
        to_skip = []

        for well in wells:
            state_well = well['state_well_number'].upper().strip()
            if state_well in existing_wells:
                to_skip.append(well)
            else:
                to_create.append(well)

        # Report skipped
        if to_skip:
            self.stdout.write(self.style.NOTICE(f'[i] SKIPPING (already exist): {len(to_skip)} wells'))
            for w in to_skip:
                self.stdout.write(f'   - {w["state_well_number"]} ({w["description"][:40]})')
            self.stdout.write('')

        # Report what will be created
        self.stdout.write(self.style.SUCCESS(f'\n[OK] WELLS TO CREATE: {len(to_create)}\n'))

        if to_create:
            self.stdout.write('Wells to be created:')
            self.stdout.write('-' * 90)

            gsa_defaults = GSA_FEE_DEFAULTS.get(gsa, {})

            for w in to_create:
                self.stdout.write(
                    f"  {w['state_well_number']:15} | "
                    f"{w['description'][:35]:35} | "
                    f"Owner: {w.get('owner_code', 'N/A'):8}"
                )
            self.stdout.write('-' * 90)

            self.stdout.write(f'\nGSA Fee Defaults to apply ({gsa.upper()}):')
            self.stdout.write(f'  - Base Rate: ${gsa_defaults.get("base_extraction_rate", "N/A")}/AF')
            if gsa_defaults.get('gsp_rate'):
                self.stdout.write(f'  - GSP Rate: ${gsa_defaults.get("gsp_rate")}/AF')
            if gsa_defaults.get('domestic_rate'):
                self.stdout.write(f'  - Domestic Rate: ${gsa_defaults.get("domestic_rate")}/AF')
            if gsa_defaults.get('fixed_quarterly_fee'):
                self.stdout.write(f'  - Fixed Quarterly: ${gsa_defaults.get("fixed_quarterly_fee")}')

        if dry_run:
            self.stdout.write(self.style.WARNING('\n[*] DRY RUN - No changes made. Remove --dry-run to create wells.\n'))
            return

        if not to_create:
            self.stdout.write(self.style.WARNING('\nNo wells to create.'))
            return

        # Confirm creation
        self.stdout.write('')
        confirm = input(f'Create {len(to_create)} wells? [y/N]: ')
        if confirm.lower() != 'y':
            self.stdout.write(self.style.WARNING('Import cancelled.'))
            return

        # Create the wells
        created_count = self.create_wells(to_create, farm, gsa)
        self.stdout.write(self.style.SUCCESS(f'\n[OK] Successfully created {created_count} wells!\n'))

    def extract_wells_from_excel(self, wb, gsa):
        """Extract well information from pivot-style Excel."""
        wells = []
        seen_wells = set()  # Track unique wells across sheets

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.stdout.write(f'Processing sheet: {sheet_name}')

            # Find the row with state well numbers
            well_row = None
            well_columns = {}
            best_count = 0

            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                row_wells = {}
                for col_idx, val in enumerate(row):
                    if val and self._looks_like_state_well(str(val)):
                        row_wells[col_idx] = str(val).strip().upper()

                if len(row_wells) > best_count:
                    best_count = len(row_wells)
                    well_row = row_idx
                    well_columns = row_wells

            if not well_row:
                self.stdout.write(f'  - No well number row found, skipping sheet')
                continue

            self.stdout.write(f'  - Found {len(well_columns)} wells in row {well_row}')

            # Get well descriptions and owner codes from rows below
            for col_idx, state_well_number in well_columns.items():
                if state_well_number in seen_wells:
                    continue
                seen_wells.add(state_well_number)

                well_info = {
                    'state_well_number': state_well_number,
                    'description': '',
                    'owner_code': '',
                    'source_sheet': sheet_name,
                }

                # Look at rows below well number for description and owner code
                for row_idx in range(well_row + 1, min(well_row + 4, 10)):
                    try:
                        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                        if col_idx < len(row) and row[col_idx]:
                            val = str(row[col_idx]).strip()
                            if not well_info['description']:
                                well_info['description'] = val
                            elif not well_info['owner_code'] and len(val) <= 10:
                                # Owner codes are typically short (e.g., JPF, FF, RMLF)
                                well_info['owner_code'] = val
                    except (IndexError, TypeError):
                        pass

                # Generate a name from description or state well number
                if well_info['description']:
                    # Clean up description for use as name
                    name = well_info['description']
                    # Truncate if too long
                    if len(name) > 50:
                        name = name[:47] + '...'
                    well_info['name'] = name
                else:
                    well_info['name'] = f"Well {state_well_number}"

                wells.append(well_info)

            self.stdout.write(f'  - Extracted {len([w for w in wells if w["source_sheet"] == sheet_name])} unique wells from this sheet')

        return wells

    def _looks_like_state_well(self, val):
        """Check if value looks like a CA state well number (e.g., 04N22W06K12S)."""
        if not val or len(val) < 10:
            return False
        val = val.upper().strip()
        pattern = r'^\d{2}[NS]\d{2}[WE]\d{2}[A-Z]\d{2}[A-Z]?$'
        return bool(re.match(pattern, val))

    def create_wells(self, wells, farm, gsa):
        """Create WaterSource records for wells."""
        created = 0
        gsa_defaults = GSA_FEE_DEFAULTS.get(gsa, {})

        with transaction.atomic():
            for w in wells:
                try:
                    water_source = WaterSource(
                        farm=farm,
                        name=w['name'],
                        source_type='well',

                        # Well identification
                        well_name=w['name'],
                        state_well_number=w['state_well_number'],

                        # GSA info
                        gsa=gsa,
                        owner_code=w.get('owner_code', ''),

                        # Fee configuration from GSA defaults
                        base_extraction_rate=gsa_defaults.get('base_extraction_rate'),
                        gsp_rate=gsa_defaults.get('gsp_rate'),
                        domestic_rate=gsa_defaults.get('domestic_rate'),
                        fixed_quarterly_fee=gsa_defaults.get('fixed_quarterly_fee'),

                        # Default settings
                        has_flowmeter=True,
                        flowmeter_units='acre_feet',
                        flowmeter_multiplier=Decimal('1.0'),
                        well_status='active',
                        active=True,
                        used_for_irrigation=True,

                        notes=f"Auto-imported from {gsa.upper()} Excel file. Original description: {w['description']}"
                    )
                    water_source.save()
                    created += 1
                    self.stdout.write(f'  Created: {w["state_well_number"]} - {w["name"]}')

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'  Error creating {w["state_well_number"]}: {e}'))

        return created
